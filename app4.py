import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, timedelta
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io

# Database Operations
def migrate_db():
    """Add created_at column to existing attendance table with proper SQLite handling"""
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    
    try:
        # Check if created_at column exists
        cursor.execute("PRAGMA table_info(attendance)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'created_at' not in columns:
            # Step 1: Add column without default value
            cursor.execute("ALTER TABLE attendance ADD COLUMN created_at TEXT")
            
            # Step 2: Update existing records with current timestamp
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("UPDATE attendance SET created_at = ?", (current_time,))
            
            # Step 3: Create temporary table with desired schema
            cursor.execute("""
                CREATE TABLE attendance_new (
                    id INTEGER PRIMARY KEY,
                    ht_number TEXT,
                    date TEXT,
                    period TEXT,
                    status TEXT,
                    faculty TEXT,
                    subject TEXT,
                    lesson_plan TEXT,
                    created_at TEXT DEFAULT (datetime('now', 'localtime'))
                )
            """)
            
            # Step 4: Copy data to new table
            cursor.execute("""
                INSERT INTO attendance_new 
                SELECT id, ht_number, date, period, status, faculty, subject, lesson_plan, created_at
                FROM attendance
            """)
            
            # Step 5: Drop old table and rename new table
            cursor.execute("DROP TABLE attendance")
            cursor.execute("ALTER TABLE attendance_new RENAME TO attendance")
            
            conn.commit()
            print("Database migration completed successfully")
    except Exception as e:
        print(f"Migration error: {str(e)}")
        conn.rollback()
    finally:
        conn.close()

def init_db():
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    
    # Create Students table
    c.execute('''CREATE TABLE IF NOT EXISTS students
                 (ht_number TEXT PRIMARY KEY, name TEXT, original_section TEXT, merged_section TEXT)''')
    
    # Create Faculty table
    c.execute('''CREATE TABLE IF NOT EXISTS faculty
                 (id INTEGER PRIMARY KEY, name TEXT, username TEXT UNIQUE, password TEXT)''')
    
    # Create Attendance table with created_at column
    c.execute('''CREATE TABLE IF NOT EXISTS attendance
                 (id INTEGER PRIMARY KEY, 
                  ht_number TEXT, 
                  date TEXT, 
                  period TEXT, 
                  status TEXT, 
                  faculty TEXT, 
                  subject TEXT, 
                  lesson_plan TEXT,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    # Create Section-Subject-Mapping table
    c.execute('''CREATE TABLE IF NOT EXISTS section_subject_mapping
                 (id INTEGER PRIMARY KEY, section TEXT, subject_name TEXT)''')
    
    # Add default admin user if not exists
    c.execute("INSERT OR IGNORE INTO faculty (username, password, name) VALUES (?, ?, ?)", 
             ('admin', 'admin', 'Administrator'))
             
    conn.commit()
    conn.close()
    
    # Run migration to ensure created_at column exists
    migrate_db()

def add_student(ht_number, name, original_section, merged_section):
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO students VALUES (?, ?, ?, ?)", 
              (ht_number, name, original_section, merged_section))
    conn.commit()
    conn.close()

def add_faculty(name, username, password):
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO faculty (name, username, password) VALUES (?, ?, ?)", 
              (name, username, password))
    conn.commit()
    conn.close()


def get_students(section=None):
    conn = sqlite3.connect('attendance.db')
    if section:
        df = pd.read_sql_query("SELECT * FROM students WHERE merged_section = ?", conn, params=(section,))
    else:
        df = pd.read_sql_query("SELECT * FROM students", conn)
    conn.close()
    return df

def get_faculty():
    conn = sqlite3.connect('attendance.db')
    df = pd.read_sql_query("SELECT * FROM faculty", conn)
    conn.close()
    return df

def get_attendance(start_date, end_date=None, section=None):
    conn = sqlite3.connect('attendance.db')
    query = "SELECT * FROM attendance WHERE date >= ?"
    params = [start_date]
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    if section:
        query += " AND ht_number IN (SELECT ht_number FROM students WHERE merged_section = ?)"
        params.append(section)
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def get_section_subjects(section):
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT subject_name FROM section_subject_mapping WHERE section = ? ORDER BY subject_name", (section,))
    subjects = cursor.fetchall()
    conn.close()
    return [subject[0] for subject in subjects]

# Page Functions
def login():
    st.title("Login")
    
    # Center the login form
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.image("https://rvit.edu.in/rvitlogo_f.jpg", width=200)
        login_type = st.radio("Select Login Type", ["Faculty", "Admin"])
        username = st.text_input("Username").strip()  # Remove whitespace
        password = st.text_input("Password", type="password").strip()  # Remove whitespace
        
        if st.button("Login", type="primary", use_container_width=True):
            conn = sqlite3.connect('attendance.db')
            cursor = conn.cursor()
            
            # Use parameterized query to prevent SQL injection
            cursor.execute("""
                SELECT * FROM faculty 
                WHERE username = ? AND password = ?
            """, (username, password))
            
            user = cursor.fetchone()
            conn.close()
            
            if user:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.session_state.faculty_name = user[1]  # faculty name is in index 1
                
                # Check if the user is an admin
                if login_type == "Admin" and username == "rvk":
                    st.session_state.is_admin = True
                else:
                    st.session_state.is_admin = False
                
                st.rerun()
            else:
                st.error("Invalid credentials")



def get_last_attendance_pattern(section, faculty=None):
    """Get the last attendance pattern for a section regardless of faculty"""
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    try:
        # Get the most recent date and period for this section
        # Removed faculty filter to allow patterns from any faculty
        query = """
            SELECT DISTINCT a.date, a.period, a.faculty
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE s.merged_section = ? 
            ORDER BY a.date DESC, a.period DESC
            LIMIT 1
        """
        cursor.execute(query, (section,))
        last_record = cursor.fetchone()
        
        if not last_record:
            return {}
            
        date, period, marking_faculty = last_record
        
        # Get the attendance pattern for that date and period
        query = """
            SELECT a.ht_number, a.status
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE s.merged_section = ?
            AND a.date = ?
            AND a.period = ?
        """
        cursor.execute(query, (section, date, period))
        pattern = {row[0]: row[1] for row in cursor.fetchall()}
        
        # If pattern found, log which faculty's pattern is being used
        if pattern:
            st.info(f"Using attendance pattern from {marking_faculty}'s class on {date}")
            
        return pattern
    except Exception as e:
        st.error(f"Error getting attendance pattern: {str(e)}")
        return {}
    finally:
        conn.close()



def check_attendance_exists(section, date, period):
    """Check if attendance has already been marked for given section, date and period"""
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    try:
        # Modified query to also get the faculty name who marked attendance
        query = """
            SELECT DISTINCT a.faculty 
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE s.merged_section = ? 
            AND a.date = ? 
            AND a.period = ?
            LIMIT 1
        """
        cursor.execute(query, (section, date, period))
        result = cursor.fetchone()
        if result:
            return True, result[0]  # Return both status and faculty name
        return False, None
    finally:
        conn.close()



def mark_attendance_page():
    """Enhanced attendance marking page with improved UI and mobile responsiveness"""
    
    # Custom CSS for better mobile UI
    st.markdown("""
        <style>
        /* Current Session Header */
        .current-session {
            background: linear-gradient(135deg, #6B46C1 0%, #553C9A 100%);
            padding: 1.5rem;
            border-radius: 1rem;
            margin-bottom: 1.5rem;
            color: white;
        }
        
        /* Force columns to stay side by side */
        div.row-widget.stHorizontal > div[data-testid="column"] {
            width: auto !important;
            flex: none !important;
        }
        
        div.row-widget.stHorizontal > div[data-testid="column"]:first-child {
            width: 70% !important;
            flex: 0 0 70% !important;
        }
        
        div.row-widget.stHorizontal > div[data-testid="column"]:last-child {
            width: 30% !important;
            flex: 0 0 30% !important;
        }
        
        /* Student Entry */
        .student-entry {
            background: #1E1E1E;
            border-bottom: 1px solid #2D2D2D;
            padding: 0.75rem;
            height: 100%;
            display: flex;
            flex-direction: column;
            justify-content: center;
        }
        
        .student-info {
            display: flex;
            flex-direction: column;
            gap: 0.25rem;
        }
        
        .student-name {
            color: #FF1493;
            font-size: 0.8rem;
            font-weight: 500;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        
        .student-id {
            color: #FFA500;
            font-size: 1.0rem;
        }
        
        .section-label {
            color: #808080;
            font-size: 0.75rem;
        }
        
        /* Checkbox styling */
        div[data-testid="column"] [data-testid="stCheckbox"] {
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
            height: 100% !important;
            margin: 0 !important;
            padding: 0.75rem !important;
        }
        
        .stCheckbox > label {
            min-width: 44px;
            min-height: 44px;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }
        
        /* Remove default padding and gaps */
        div[data-testid="column"] {
            padding: 0 !important;
        }
        
        div.row-widget.stHorizontal {
            flex-wrap: nowrap !important;
            gap: 0 !important;
        }
        
        /* Ensure form elements don't break layout */
        .stForm > div[data-testid="stForm"] {
            width: 100% !important;
        }
        
        /* Force mobile layout to match desktop */
        @media (max-width: 768px) {
            div.row-widget.stHorizontal {
                display: flex !important;
                flex-direction: row !important;
            }
            
            div.row-widget.stHorizontal > div[data-testid="column"] {
                min-width: 0 !important;
            }
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Get unique sections
    students_df = get_students()
    if students_df.empty:
        st.warning("No students found in the database. Please add students first.")
        return
        
    sections = students_df['merged_section'].unique()
    
    # Selection controls in a card
    st.markdown('<div class="current-session">', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        section = st.selectbox(
            "Select Section",
            [""] + list(sections),  # Add empty string as first option
            index=0,  # Set index to 0 to select the empty option
            key="section_select"
        )
    with col2:
        period = st.selectbox(
            "Select Period",
            [""] + ['P1', 'P2', 'P3', 'P4', 'P5', 'P6'],
            index=0,
            key="period_select"
        )
    with col3:
        # Only show subjects if a section is selected
        if section:
            subject = st.selectbox(
                "Select Subject",
                [""] + get_section_subjects(section),
                index=0,
                key="subject_select"
            )
        else:
            subject = st.selectbox(
                "Select Subject",
                [""],  # Only show empty option if no section selected
                index=0,
                key="subject_select"
            )
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Check if attendance already exists
    today = datetime.now().strftime('%Y-%m-%d')
    exists, faculty = check_attendance_exists(section, today, period)
    if exists:
        st.error(f"Attendance has already been marked for {section} - {period} today by {faculty}!")
        return
    
    if section and period and subject:
        students = get_students(section)
        
        # Action buttons
        col1, col2, col3 = st.columns(3)
        with col1:
            mark_all_present = st.button("✓ Mark All Present", 
                                       type="primary",
                                       key="mark_all_present",
                                       use_container_width=True)
        with col2:
            mark_all_absent = st.button("✗ Mark All Absent", 
                                      key="mark_all_absent",
                                      use_container_width=True)
        with col3:
            use_last_pattern = st.button("↺ Use Last Pattern", 
                                       key="use_last_pattern",
                                       use_container_width=True)
        
        # Initialize attendance data
        if 'attendance_data' not in st.session_state:
            st.session_state.attendance_data = {}
        
        # Update attendance data based on button clicks
        if mark_all_present:
            for student in students.itertuples():
                st.session_state.attendance_data[student.ht_number] = 'P'
        elif mark_all_absent:
            for student in students.itertuples():
                st.session_state.attendance_data[student.ht_number] = 'A'
        elif use_last_pattern:
            pattern = get_last_attendance_pattern(section, st.session_state.faculty_name)
            if pattern:
                st.session_state.attendance_data = pattern
                st.success("Last attendance pattern applied!")
            else:
                st.info("No previous attendance pattern found for this section.")
        
        # Create a form for student attendance
        with st.form(key='attendance_form'):
            for student in students.itertuples():
                # Create two columns with 70-30 split
                cols = st.columns([0.7, 0.3])
                
                # Student info in first column
                with cols[0]:
                    st.markdown(f"""
                        <div class="student-entry">
                            <div class="student-info">
                                <span class="student-name">{student.name}</span>
                                <span class="student-id">{student.ht_number}</span>
                                <span class="section-label">{student.original_section}</span>
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
                
                # Attendance checkbox in second column
                with cols[1]:
                    default_value = st.session_state.attendance_data.get(student.ht_number, 'P') == 'P'
                    status = st.checkbox("Present", 
                                      value=default_value,
                                      key=f"attendance_{student.ht_number}")
                    st.session_state.attendance_data[student.ht_number] = 'P' if status else 'A'
            
            # Lesson plan input
            st.markdown("<br>", unsafe_allow_html=True)
            lesson_plan = st.text_area("📝 Lesson Plan", 
                                     placeholder="Enter the topic covered in this class...",
                                     height=100)
            
            # Submit button
            submit_button = st.form_submit_button("Submit Attendance", 
                                                type="primary",
                                                use_container_width=True)
        
        if submit_button:
            if not lesson_plan.strip():
                st.error("Please enter a lesson plan before submitting attendance.")
            else:
                # Double check for duplicate attendance
                exists, faculty = check_attendance_exists(section, today, period)
                if exists:
                    st.error(f"Attendance has already been marked for {section} - {period} today by {faculty}!")
                    return
                    
                date = datetime.now().strftime('%Y-%m-%d')
                for ht_number, status in st.session_state.attendance_data.items():
                    mark_attendance(ht_number, date, period, status, 
                                 st.session_state.faculty_name, subject, lesson_plan)
                st.success("Attendance marked successfully!")
                st.session_state.attendance_data = {}
                st.rerun()







# Add these functions after your existing database functions

def get_overall_statistics(start_date, end_date, sections=None):
    """Enhanced overall statistics with detailed analytics"""
    conn = sqlite3.connect('attendance.db')
    try:
        # Base query with proper date filtering and section handling
        query = """
            WITH DateRangeAttendance AS (
                SELECT 
                    s.ht_number,
                    s.name,
                    s.original_section,
                    a.subject,
                    a.date,
                    a.period,
                    a.status,
                    strftime('%H', a.created_at) as hour,
                    a.created_at
                FROM students s
                JOIN attendance a ON s.ht_number = a.ht_number
                WHERE date(a.date) BETWEEN date(?) AND date(?)
        """
        
        params = [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')]
        
        # Add section filter if sections are provided
        if sections:
            section_placeholders = ','.join('?' * len(sections))
            query += f" AND s.original_section IN ({section_placeholders})"
            params.extend(sections)
            
        # Close the CTE and select from it
        query += """
            )
            SELECT * FROM DateRangeAttendance
        """
        
        # Execute query and create DataFrame
        df = pd.read_sql_query(query, conn, params=params)
        
        if df.empty:
            return {
                'student_stats': pd.DataFrame(),
                'total_students': 0,
                'avg_attendance': 0,
                'below_75': 0,
                'day_wise_stats': pd.DataFrame(),
                'subject_wise_stats': pd.DataFrame(),
                'time_slot_analysis': pd.DataFrame(),
                'trend_analysis': pd.DataFrame(),
                'section_comparison': pd.DataFrame()
            }
            
        # 1. Student-wise Statistics
        student_stats = df.groupby(['ht_number', 'name', 'original_section']).agg({
            'status': lambda x: sum(x == 'P'),
            'date': 'count'
        }).reset_index()
        
        student_stats.columns = ['HT Number', 'Student Name', 'Section', 'Present', 'Total']
        student_stats['Percentage'] = (student_stats['Present'] / student_stats['Total'] * 100).round(2)
        student_stats['Classes (A/C)'] = student_stats.apply(
            lambda x: f"{int(x['Present'])}/{int(x['Total'])}", 
            axis=1
        )
        
        # 2. Day-wise Analysis
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        df['day_name'] = pd.to_datetime(df['date']).dt.day_name()
        day_wise = df.groupby('day_name').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2)
        }).reindex(days).reset_index()
        day_wise.columns = ['Day', 'Attendance %']
        
        # 3. Subject-wise Analysis
        subject_wise = df.groupby('subject').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2),
            'ht_number': 'count'
        }).reset_index()
        subject_wise.columns = ['Subject', 'Attendance %', 'Total Classes']
        
        # 4. Time Slot Analysis - Fixed to use proper hour column
        df['hour'] = pd.to_numeric(df['hour'], errors='coerce')  # Convert hour to numeric
        time_slot = df.groupby(['period', 'hour']).agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2)
        }).reset_index()
        time_slot.columns = ['Period', 'Hour', 'Attendance %']
        
        # 5. Trend Analysis
        trend = df.groupby('date').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2)
        }).reset_index()
        trend.columns = ['Date', 'Attendance %']
        
        # 6. Section Comparison
        section_comp = df.groupby('original_section').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2),
            'ht_number': lambda x: len(set(x))
        }).reset_index()
        section_comp.columns = ['Section', 'Attendance %', 'Student Count']
        
        # Calculate summary statistics
        total_students = len(student_stats)
        avg_attendance = student_stats['Percentage'].mean()
        below_75 = len(student_stats[student_stats['Percentage'] < 75])
        
        return {
            'student_stats': student_stats,
            'total_students': total_students,
            'avg_attendance': avg_attendance,
            'below_75': below_75,
            'day_wise_stats': day_wise,
            'subject_wise_stats': subject_wise,
            'time_slot_analysis': time_slot,
            'trend_analysis': trend,
            'section_comparison': section_comp
        }
        
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return None
    finally:
        conn.close()



def get_student_report(ht_number, start_date, end_date):
    """Modified to properly handle date range filtering"""
    conn = sqlite3.connect('attendance.db')
    try:
        # Convert dates to proper format
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
        
        # Get student details
        student_query = "SELECT * FROM students WHERE ht_number = ?"
        student_df = pd.read_sql_query(student_query, conn, params=[ht_number])
        
        # Get attendance details with proper date filtering
        attendance_query = """
            SELECT 
                subject,
                COUNT(DISTINCT date || period) as total_classes,
                SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) as present_count,
                ROUND(CAST(SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) AS FLOAT) / 
                    NULLIF(COUNT(*), 0) * 100, 2) as attendance_percentage
            FROM attendance 
            WHERE ht_number = ? 
            AND date(date) BETWEEN date(?) AND date(?)
            GROUP BY subject
        """
        attendance_df = pd.read_sql_query(
            attendance_query, 
            conn, 
            params=[
                ht_number, 
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            ]
        )
        
        # Get daily attendance with proper date filtering
        daily_query = """
            SELECT 
                date,
                period,
                subject,
                status,
                faculty,
                lesson_plan,
                created_at as submission_time
            FROM attendance 
            WHERE ht_number = ? 
            AND date(date) BETWEEN date(?) AND date(?)
            ORDER BY date DESC, period ASC
        """
        daily_df = pd.read_sql_query(
            daily_query, 
            conn, 
            params=[
                ht_number,
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            ]
        )
        
        return student_df, attendance_df, daily_df
        
    except Exception as e:
        st.error(f"Error generating student report: {str(e)}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    finally:
        conn.close()

def get_subject_analysis(start_date, end_date, section=None):
    """Modified to properly handle date range filtering"""
    conn = sqlite3.connect('attendance.db')
    try:
        # Convert dates to proper format
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
        
        query = """
            SELECT 
                subject,
                COUNT(DISTINCT date || period) as total_classes,
                SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) as present_count,
                COUNT(*) as total_attendance,
                ROUND(CAST(SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) AS FLOAT) / 
                    NULLIF(COUNT(*), 0) * 100, 2) as attendance_percentage
            FROM attendance 
            WHERE date(date) BETWEEN date(?) AND date(?)
        """
        params = [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')]
        
        if section and section != 'All':
            query += " AND ht_number IN (SELECT ht_number FROM students WHERE merged_section = ?)"
            params.append(section)
        
        query += " GROUP BY subject ORDER BY subject"
        
        df = pd.read_sql_query(query, conn, params=params)
        return df
        
    except Exception as e:
        st.error(f"Error generating subject analysis: {str(e)}")
        return pd.DataFrame()
    finally:
        conn.close()



def export_to_excel(df, filename):
    """Export DataFrame to Excel with proper formatting"""
    # Create Excel writer object
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance Report')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Attendance Report']
        
        # Format headers
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        
        # Apply formatting
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 15)
    
    return output.getvalue()

def view_statistics_page():
    """Enhanced statistics page with course-wise filtering and Excel export"""
    
    # Tab layout
    tab1, tab2, tab3 = st.tabs(["📊 Overall Statistics", "👤 Student Reports", "📚 Subject Analysis"])
    
    with tab1:
        st.subheader("Overall Attendance Statistics")
        
        # Get unique courses and sections
        students_df = get_students()
        if students_df.empty:
            st.warning("No students found in the database.")
            return
        
        # Course selection (e.g., "B.Tech-II")
        courses = sorted(list(set([
            section.split('-')[0] 
            for section in students_df['original_section'].unique()
        ])))
        selected_course = st.selectbox("Select Course", courses)
        
        # Get sections for selected course
        sections = sorted(list(students_df[
            students_df['original_section'].str.startswith(selected_course)
        ]['original_section'].unique()))
        
        # Add "All" option at the beginning of the list
        sections = ['All'] + sections
        
        # Multi-select for sections with "All" as default
        selected_sections = st.multiselect(
            "Select Sections",
            sections,
            default=['All'],
            key="section_multiselect"
        )
        
        # Handle "All" selection
        if 'All' in selected_sections:
            selected_sections = sections[1:]  # All sections except "All"
        
        # Date range selection
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("From Date", 
                                     value=datetime.now() - timedelta(days=30),
                                     key="start_date_overall")
        with col2:
            end_date = st.date_input("To Date", 
                                   value=datetime.now(),
                                   key="end_date_overall")
        
        if selected_sections and st.button("Generate Report", type="primary"):
            try:
                conn = sqlite3.connect('attendance.db')
                
                # Modified query to ensure correct attendance counting
                query = """
                    WITH AllStudents AS (
                        -- Get all students in selected sections
                        SELECT ht_number, name, original_section
                        FROM students
                        WHERE original_section IN ({})
                    ),
                    SectionClasses AS (
                        -- Get conducted classes per section
                        SELECT 
                            s.original_section,
                            a.date,
                            a.period,
                            a.subject
                        FROM attendance a
                        JOIN students s ON a.ht_number = s.ht_number
                        WHERE a.date BETWEEN ? AND ?
                        AND s.original_section IN ({})
                        GROUP BY s.original_section, a.date, a.period, a.subject
                    ),
                    StudentAttendance AS (
                        -- Calculate attendance per student
                        SELECT 
                            s.ht_number,
                            s.name,
                            s.original_section as section,
                            COUNT(DISTINCT sc.date || sc.period) as conducted,
                            COUNT(DISTINCT CASE WHEN a.status = 'P' 
                                THEN a.date || a.period 
                                ELSE NULL END) as attended
                        FROM AllStudents s
                        LEFT JOIN SectionClasses sc ON s.original_section = sc.original_section
                        LEFT JOIN attendance a ON s.ht_number = a.ht_number 
                            AND sc.date = a.date 
                            AND sc.period = a.period
                        GROUP BY s.ht_number, s.name, s.original_section
                    )
                    SELECT 
                        ht_number as "HT Number",
                        name as "Student Name",
                        section as "Section",
                        attended || '/' || conducted as "Classes (A/C)",
                        ROUND(CAST(attended AS FLOAT) / NULLIF(conducted, 0) * 100, 2) as "Overall %"
                    FROM StudentAttendance
                    ORDER BY "Section", "HT Number"
                """.format(
                    ','.join(['?'] * len(selected_sections)),
                    ','.join(['?'] * len(selected_sections))
                )
                
                params = [
                    *selected_sections,
                    start_date.strftime('%Y-%m-%d'),
                    end_date.strftime('%Y-%m-%d'),
                    *selected_sections
                ]
                
                df = pd.read_sql_query(query, conn, params=params)
                
                if not df.empty:
                    # Calculate summary metrics
                    total_students = len(df)
                    avg_attendance = df['Overall %'].mean()
                    below_75 = len(df[df['Overall %'] < 75])
                    
                    # Display metrics
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Students", total_students)
                    with col2:
                        st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                    with col3:
                        st.metric("Students Below 75%", below_75)
                    
                    # Display student-wise statistics
                    st.subheader("Student-wise Statistics")
                    st.dataframe(df)
                    
                    # Export to Excel
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Attendance Report', index=False)
                        
                        workbook = writer.book
                        worksheet = writer.sheets['Attendance Report']
                        
                        # Format headers
                        for col_num, value in enumerate(df.columns.values):
                            cell = worksheet.cell(1, col_num + 1)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color='D3D3D3',
                                end_color='D3D3D3',
                                fill_type='solid'
                            )
                        
                        # Adjust column widths
                        for idx, col in enumerate(df.columns):
                            worksheet.column_dimensions[get_column_letter(idx + 1)].width = 15
                    
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=output.getvalue(),
                        file_name=f"attendance_report_{selected_course}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.info("No attendance records found for the selected criteria.")
                
            except Exception as e:
                st.error(f"Error generating report: {str(e)}")
            finally:
                conn.close()


    with tab2:
        st.subheader("Student Reports")
        
        # Student selection
        student_ht = st.selectbox(
            "Select Student",
            students_df['ht_number'].tolist(),
            format_func=lambda x: (
                f"{x} - {students_df[students_df['ht_number'] == x]['name'].iloc[0]}"
            )
        )
        
        # Date range for student report
        col1, col2 = st.columns(2)
        with col1:
            student_start_date = st.date_input(
                "From Date",
                value=datetime.now() - timedelta(days=30),
                key="start_date_student"
            )
        with col2:
            student_end_date = st.date_input(
                "To Date",
                value=datetime.now(),
                key="end_date_student"
            )
        
        if st.button("Generate Student Report", type="primary"):
            student_df, attendance_df, daily_df = get_student_report(
                student_ht,
                student_start_date.strftime('%Y-%m-%d'),
                student_end_date.strftime('%Y-%m-%d')
            )
            
            if not student_df.empty:
                # Student Information
                st.write("### Student Information")
                st.write(f"Name: {student_df['name'].iloc[0]}")
                st.write(f"Section: {student_df['original_section'].iloc[0]}")
                
                if not attendance_df.empty:
                    # Format attendance data
                    attendance_df['Classes'] = attendance_df.apply(
                        lambda x: f"{int(x['present_count'])}/{int(x['total_classes'])}",
                        axis=1
                    )
                    
                    # Display subject-wise attendance
                    st.write("### Subject-wise Attendance")
                    st.dataframe(attendance_df[
                        ['subject', 'Classes', 'attendance_percentage']
                    ].rename(columns={
                        'Classes': 'Classes (Attended/Conducted)',
                        'attendance_percentage': 'Attendance %'
                    }))
                
                if not daily_df.empty:
                    st.write("### Daily Attendance Log")
                    st.dataframe(daily_df)
                    
                    # Export student report to Excel
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        daily_df.to_excel(
                            writer,
                            sheet_name='Daily Log',
                            index=False
                        )
                        attendance_df.to_excel(
                            writer,
                            sheet_name='Subject Summary',
                            index=False
                        )
                    
                    st.download_button(
                        label="📥 Download Student Report",
                        data=output.getvalue(),
                        file_name=(
                            f"student_report_{student_ht}_"
                            f"{datetime.now().strftime('%Y%m%d')}.xlsx"
                        ),
                        mime=(
                            "application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet"
                        )
                    )
                else:
                    st.info("No attendance records found for the selected date range.")
            else:
                st.error("Student not found in the database.")
    
    with tab3:
        st.subheader("Subject Analysis")
        
        # Date range for subject analysis
        col1, col2 = st.columns(2)
        with col1:
            subject_start_date = st.date_input(
                "From Date",
                value=datetime.now() - timedelta(days=30),
                key="start_date_subject"
            )
        with col2:
            subject_end_date = st.date_input(
                "To Date",
                value=datetime.now(),
                key="end_date_subject"
            )
        
        # Section selection for subject analysis
        subject_section = st.selectbox(
            "Select Section",
            ['All'] + list(students_df['original_section'].unique()),
            key="section_subject"
        )
        
        if st.button("Generate Subject Analysis", type="primary"):
            subject_df = get_subject_analysis(
                subject_start_date.strftime('%Y-%m-%d'),
                subject_end_date.strftime('%Y-%m-%d'),
                None if subject_section == 'All' else subject_section
            )
            
            if not subject_df.empty:
                # Display subject-wise analysis
                st.write("### Subject-wise Attendance Analysis")
                
                # Create bar chart
                st.bar_chart(subject_df.set_index('subject')['attendance_percentage'])
                
                # Display detailed statistics
                st.write("### Detailed Subject Statistics")
                st.dataframe(subject_df)
                
                # Export subject analysis to Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    subject_df.to_excel(
                        writer,
                        sheet_name='Subject Analysis',
                        index=False
                    )
                
                st.download_button(
                    label="📥 Download Subject Analysis",
                    data=output.getvalue(),
                    file_name=(
                        f"subject_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
                    ),
                    mime=(
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    )
                )
            else:
                st.info("No attendance records found for subject analysis.")





def reset_credentials_page():
    st.subheader("Reset Credentials")
    
    with st.form("reset_credentials_form"):
        current_password = st.text_input("Current Password", type="password")
        new_password = st.text_input("New Password", type="password")
        confirm_password = st.text_input("Confirm New Password", type="password")
        
        submitted = st.form_submit_button("Reset Password", type="primary")
        
        if submitted:
            if not all([current_password, new_password, confirm_password]):
                st.error("All fields are required")
                return
                
            if new_password != confirm_password:
                st.error("New passwords do not match")
                return
                
            conn = sqlite3.connect('attendance.db')
            try:
                cursor = conn.cursor()
                
                # Verify current password
                cursor.execute("""
                    SELECT * FROM faculty 
                    WHERE username = ? AND password = ?
                """, (st.session_state.username, current_password))
                
                if cursor.fetchone():
                    # Update password
                    cursor.execute("""
                        UPDATE faculty 
                        SET password = ? 
                        WHERE username = ?
                    """, (new_password, st.session_state.username))
                    
                    conn.commit()
                    st.success("Password updated successfully! Please login again.")
                    
                    # Clear session state to force re-login
                    for key in list(st.session_state.keys()):
                        del st.session_state[key]
                    st.rerun()
                else:
                    st.error("Current password is incorrect")
                    
            except Exception as e:
                st.error(f"Error resetting password: {str(e)}")
            finally:
                conn.close()
                

def mark_attendance(ht_number, date, period, status, faculty, subject, lesson_plan):
    """Modified to properly handle timestamp creation"""
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    c.execute("""
        INSERT INTO attendance 
        (ht_number, date, period, status, faculty, subject, lesson_plan, created_at) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (ht_number, date, period, status, faculty, subject, lesson_plan, current_time))
    
    conn.commit()
    conn.close()
  
def get_faculty_classload(username, start_date=None, end_date=None):
    """Modified to handle cases where created_at might not exist"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Get faculty name
        cursor.execute("SELECT name FROM faculty WHERE username = ?", (username,))
        faculty_name = cursor.fetchone()[0]
        
        # Base parameters
        params = [faculty_name]
        
        # Prepare date filters if provided
        date_filter = ""
        if start_date and end_date:
            date_filter = "AND a.date >= ? AND a.date <= ?"
            params.extend([start_date, end_date])
        
        # Modified query to handle missing created_at
        query = f"""
            WITH PeriodGroups AS (
                SELECT 
                    date,
                    period,
                    COUNT(DISTINCT subject) as subject_count
                FROM (
                    SELECT DISTINCT
                        a.date,
                        a.period,
                        a.subject
                    FROM attendance a
                    WHERE a.faculty = ?
                    {date_filter}
                )
                GROUP BY date, period
            ),
            DetailedClasses AS (
                SELECT 
                    a.date,
                    a.period,
                    a.subject,
                    GROUP_CONCAT(DISTINCT s.merged_section) as sections,
                    COUNT(DISTINCT s.merged_section) as section_count,
                    a.lesson_plan,
                    MIN(a.id) as first_entry_id,
                    COUNT(DISTINCT a.ht_number) as student_count,
                    pg.subject_count,
                    COALESCE(MIN(a.created_at), datetime('now', 'localtime')) as submission_time
                FROM attendance a
                JOIN students s ON a.ht_number = s.ht_number
                JOIN PeriodGroups pg ON a.date = pg.date AND a.period = pg.period
                WHERE a.faculty = ?
                {date_filter}
                GROUP BY a.date, a.period, a.subject
            )
            SELECT 
                d.date,
                d.period,
                d.subject,
                d.sections,
                d.section_count,
                ROUND(1.0/d.subject_count, 2) as distributed_load,
                d.lesson_plan,
                d.student_count,
                time(d.submission_time) as submission_time
            FROM DetailedClasses d
            ORDER BY d.date DESC, d.period, d.subject
        """
        
        cursor.execute(query, params + params)
        rows = cursor.fetchall()
        
        if not rows:
            return pd.DataFrame()
            
        # Create DataFrame with submission times
        classload_data = []
        for row in rows:
            date, period, subject, sections, section_count, dist_load, lesson_plan, student_count, submission_time = row
            
            entry = {
                'Date': date,
                'Period': period,
                'Subject': subject,
                'Section': sections,
                'Combined Sections': sections,
                'Distributed Load': dist_load,
                'Student Count': student_count,
                'Submission Time': submission_time or 'Not recorded',
                'Lesson Plan': lesson_plan
            }
            classload_data.append(entry)
        
        df = pd.DataFrame(classload_data)
        
        if not df.empty:
            df['Total Load'] = df.groupby(['Date', 'Period'])['Distributed Load'].transform('sum')
            df['Day'] = pd.to_datetime(df['Date']).dt.strftime('%A')
            df = df.sort_values(['Date', 'Period', 'Subject'])
            df['Cumulative Load'] = df['Distributed Load'].cumsum()
        
        return df
        
    except Exception as e:
        print(f"Error calculating faculty classload: {str(e)}")
        return pd.DataFrame()
    finally:
        if conn:
            conn.close()

def show_class_timetable_page():
    """Modified to handle cases where created_at might not exist"""
    st.subheader("Class Timetable")
    
    col1, col2 = st.columns(2)
    with col1:
        selected_date = st.date_input(
            "Select Date",
            datetime.now(),
            key="class_timetable_date"
        )
    with col2:
        view_type = st.selectbox(
            "View Type",
            ["Single Day", "Date Range"],
            key="class_view_type"
        )
    
    if view_type == "Date Range":
        end_date = st.date_input(
            "End Date",
            selected_date,
            key="class_timetable_end_date"
        )
    else:
        end_date = selected_date
    
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Modified query to handle missing created_at
        query = """
            SELECT DISTINCT
                a.date,
                a.period,
                a.faculty,
                a.subject,
                s.merged_section as section,
                MIN(a.id) as first_entry_id,
                COUNT(DISTINCT a.ht_number) as student_count,
                COALESCE(MIN(a.created_at), datetime('now', 'localtime')) as submission_time
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE a.date BETWEEN ? AND ?
            GROUP BY a.date, a.period, a.faculty, a.subject, s.merged_section
            ORDER BY a.date, a.period, s.merged_section
        """
        
        cursor.execute(query, (
            selected_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d')
        ))
        
        classes_data = []
        for row in cursor.fetchall():
            date, period, faculty, subject, section, entry_id, student_count, submission_time = row
            
            # Format submission time
            try:
                time_str = datetime.strptime(submission_time, '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
            except (ValueError, TypeError):
                time_str = 'Not recorded'
            
            classes_data.append({
                'Date': date,
                'Period': period,
                'Faculty': faculty,
                'Subject': subject,
                'Section': section,
                'Students': student_count,
                'Submission Time': time_str,
                'Status': 'Completed'
            })
        
        if classes_data:
            df_classes = pd.DataFrame(classes_data)
            
            # Display summary metrics
            st.write("### Daily Summary")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Classes", len(df_classes))
            with col2:
                st.metric("Faculty Engaged", df_classes['Faculty'].nunique())
            with col3:
                st.metric("Sections Covered", df_classes['Section'].nunique())
            
            # Display timetable with submission time
            st.write("### Class Timetable")
            st.dataframe(
                df_classes,
                column_config={
                    'Date': st.column_config.TextColumn('Date', width=100),
                    'Period': st.column_config.TextColumn('Period', width=70),
                    'Faculty': st.column_config.TextColumn('Faculty', width=150),
                    'Subject': st.column_config.TextColumn('Subject', width=120),
                    'Section': st.column_config.TextColumn('Section', width=120),
                    'Students': st.column_config.NumberColumn('Students', width=100),
                    'Submission Time': st.column_config.TextColumn('Time', width=100),
                    'Status': st.column_config.TextColumn('Status', width=100)
                },
                hide_index=True,
                use_container_width=True
            )
            
        else:
            if view_type == "Date Range":
                st.info(f"No classes recorded between {selected_date.strftime('%Y-%m-%d')} and {end_date.strftime('%Y-%m-%d')}")
            else:
                st.info(f"No classes recorded on {selected_date.strftime('%Y-%m-%d')}")
            
    except Exception as e:
        st.error(f"Error loading class timetable: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

    
                

def view_faculty_worksheet_stats():
    st.subheader("Faculty Worksheet Statistics")

    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", datetime.now() - timedelta(days=30))
    with col2:
        end_date = st.date_input("End Date", datetime.now())

    if st.button("Generate Report", type="primary"):
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()

        query = """
        SELECT 
            f.name as Faculty,
            COUNT(DISTINCT a.date || a.period) as Classes_Conducted,
            COUNT(DISTINCT fa.id) as Other_Activities,
            ROUND(COUNT(DISTINCT a.date || a.period) + COUNT(DISTINCT fa.id), 2) as Total_Load
        FROM 
            faculty f
        LEFT JOIN 
            attendance a ON f.name = a.faculty
        LEFT JOIN 
            faculty_activities fa ON f.name = fa.faculty
        WHERE 
            (a.date BETWEEN ? AND ?) OR (fa.date BETWEEN ? AND ?)
        GROUP BY 
            f.name
        ORDER BY 
            Total_Load DESC
        """

        cursor.execute(query, (start_date, end_date, start_date, end_date))
        results = cursor.fetchall()

        if results:
            df = pd.DataFrame(results, columns=['Faculty', 'Classes Conducted', 'Other Activities', 'Total Load'])
            st.dataframe(df)

            # Export to Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Faculty Worksheet Stats', index=False)
            
            st.download_button(
                label="Download Excel Report",
                data=output.getvalue(),
                file_name=f"faculty_worksheet_stats_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No data found for the selected date range.")

        conn.close()

def update_faculty_activity(username, date, period, activity_type, description):
    """Update faculty activity with enhanced error handling"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Create faculty_activities table if not exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS faculty_activities (
                id INTEGER PRIMARY KEY,
                faculty TEXT,
                date TEXT,
                period TEXT,
                activity_type TEXT,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Get faculty name
        cursor.execute("SELECT name FROM faculty WHERE username = ?", (username,))
        faculty_name = cursor.fetchone()[0]
        
        # Format date properly
        if isinstance(date, datetime):
            date = date.strftime('%Y-%m-%d')
            
        # Check if activity exists
        cursor.execute("""
            SELECT id FROM faculty_activities 
            WHERE faculty = ? AND date = ? AND period = ?
        """, (faculty_name, date, period))
        
        existing = cursor.fetchone()
        
        if existing:
            # Update existing activity
            cursor.execute("""
                UPDATE faculty_activities 
                SET activity_type = ?, description = ?
                WHERE id = ?
            """, (activity_type, description, existing[0]))
        else:
            # Insert new activity
            cursor.execute("""
                INSERT INTO faculty_activities (faculty, date, period, activity_type, description)
                VALUES (?, ?, ?, ?, ?)
            """, (faculty_name, date, period, activity_type, description))
            
        conn.commit()
        return True
        
    except Exception as e:
        st.error(f"Error updating activity: {str(e)}")
        return False
    finally:
        conn.close()





def get_faculty_worksheet_data(username, start_date, end_date=None, view_type="Daily"):
    """Get faculty worksheet data including all periods and activities with proper load distribution"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Get faculty name
        cursor.execute("SELECT name FROM faculty WHERE username = ?", (username,))
        faculty_name = cursor.fetchone()[0]
        
        # Initialize worksheet data structure
        worksheet_data = {
            'total_classes': 0,
            'total_hours': 0,
            'unique_sections': set(),
            'periods': {
                f'P{i}': {
                    'time': f'Period {i}',
                    'subject': 'No Activity',
                    'section': '-',
                    'topic': '-',
                    'status': 'Pending',
                    'activity_type': 'None'
                } for i in range(1, 7)
            }
        }
        
        # Convert dates to proper format
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date and isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        if not end_date:
            end_date = start_date
            
        # Modified query to properly handle load distribution
        query = """
            WITH PeriodGroups AS (
                SELECT 
                    date,
                    period,
                    COUNT(DISTINCT subject) as subject_count
                FROM (
                    SELECT DISTINCT
                        a.date,
                        a.period,
                        a.subject
                    FROM attendance a
                    WHERE a.faculty = ?
                    AND a.date BETWEEN ? AND ?
                )
                GROUP BY date, period
            ),
            DetailedClasses AS (
                SELECT 
                    a.date,
                    a.period,
                    a.subject,
                    GROUP_CONCAT(DISTINCT s.merged_section) as sections,
                    COUNT(DISTINCT s.merged_section) as section_count,
                    a.lesson_plan,
                    pg.subject_count,
                    ROUND(1.0/pg.subject_count, 2) as distributed_load
                FROM attendance a
                JOIN students s ON a.ht_number = s.ht_number
                JOIN PeriodGroups pg ON a.date = pg.date AND a.period = pg.period
                WHERE a.faculty = ?
                AND a.date BETWEEN ? AND ?
                GROUP BY a.date, a.period, a.subject
            )
            SELECT 
                period,
                subject,
                sections,
                lesson_plan,
                distributed_load
            FROM DetailedClasses
        """
        
        cursor.execute(query, (
            faculty_name, 
            start_date.strftime('%Y-%m-%d'), 
            end_date.strftime('%Y-%m-%d'),
            faculty_name,
            start_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d')
        ))
        
        total_load = 0
        for row in cursor.fetchall():
            period, subject, sections, lesson_plan, dist_load = row
            worksheet_data['periods'][period] = {
                'time': f'Period {period[1]}',
                'subject': subject,
                'section': sections,
                'topic': lesson_plan,
                'status': 'Completed',
                'activity_type': 'Class',
                'load': dist_load
            }
            total_load += dist_load
            worksheet_data['unique_sections'].update(sections.split(','))
            
        worksheet_data['total_classes'] = total_load
            
        # Get other activities
        query = """
            SELECT period, activity_type, description
            FROM faculty_activities
            WHERE faculty = ? AND date BETWEEN ? AND ?
        """
        
        cursor.execute(query, (faculty_name, start_date.strftime('%Y-%m-%d'),
                             end_date.strftime('%Y-%m-%d')))
                             
        for row in cursor.fetchall():
            period, activity_type, description = row
            if period in worksheet_data['periods']:
                worksheet_data['periods'][period] = {
                    'time': f'Period {period[1]}',
                    'subject': activity_type,
                    'section': '-',
                    'topic': description,
                    'status': 'Completed',
                    'activity_type': 'Other',
                    'load': 1.0
                }
                worksheet_data['total_hours'] += 1
        
        return worksheet_data
        
    except Exception as e:
        st.error(f"Error getting worksheet data: {str(e)}")
        return None
    finally:
        conn.close()



def show_faculty_classload():
    """Enhanced faculty classload page with worksheet capabilities"""
    st.subheader("Faculty Workload Dashboard")
    
    # Create tabs for different views
    tab1, tab2 = st.tabs(["📊 Overall Workload", "📝 Daily Worksheet"])
    
    # Tab 1: Overall Workload
    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("From Date", datetime.now())
        with col2:
            end_date = st.date_input("To Date", datetime.now())
            
        if st.button("Generate Workload Report", type="primary"):
            df = get_faculty_classload(
                st.session_state.username,
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            )
            
            if not df.empty:
                # Calculate summary metrics
                total_load = df['Distributed Load'].sum()
                unique_subjects = df['Subject'].nunique()
                unique_sections = df['Section'].nunique()
                
                # Display summary metrics
                st.write("### Workload Summary")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Net Classes", f"{total_load:.2f}")
                with col2:
                    st.metric("Unique Subjects", unique_subjects)
                with col3:
                    st.metric("Unique Sections", unique_sections)
                
                # Subject-wise distribution
                st.write("### Subject-wise Distribution")
                subject_dist = df.groupby('Subject')['Distributed Load'].sum().reset_index()
                subject_dist = subject_dist.sort_values('Distributed Load', ascending=False)
                st.bar_chart(subject_dist.set_index('Subject'))
                
                # Detailed workload table
                st.write("### Detailed Workload Report")
                st.dataframe(
                    df.sort_values(['Date', 'Period']),
                    column_config={
                        'Date': st.column_config.TextColumn('Date', width=100),
                        'Period': st.column_config.TextColumn('Period', width=80),
                        'Subject': st.column_config.TextColumn('Subject', width=150),
                        'Section': st.column_config.TextColumn('Section', width=120),
                        'Combined Sections': st.column_config.TextColumn('Combined Sections', width=200),
                        'Distributed Load': st.column_config.NumberColumn('Load', format="%.2f", width=80),
                        'Entry Time': st.column_config.TextColumn('Time', width=100),
                        'Lesson Plan': st.column_config.TextColumn('Lesson Plan', width=300)
                    },
                    hide_index=True
                )
                
                # Export options
                st.write("### Export Options")
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Detailed Workload', index=False)
                    
                    # Write summary sheet
                    summary_df = pd.DataFrame([{
                        'Metric': 'Net Classes',
                        'Value': total_load
                    }, {
                        'Metric': 'Unique Subjects',
                        'Value': unique_subjects
                    }, {
                        'Metric': 'Unique Sections',
                        'Value': unique_sections
                    }])
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Write subject distribution
                    subject_dist.to_excel(writer, sheet_name='Subject Distribution', index=False)
                    
                    # Format worksheets
                    for sheet in writer.sheets.values():
                        for column in sheet.columns:
                            max_length = max(len(str(cell.value or '')) for cell in column)
                            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(50, max(12, max_length + 2))
                
                st.download_button(
                    label="📥 Download Complete Report",
                    data=buffer.getvalue(),
                    file_name=f"faculty_workload_{st.session_state.username}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("No workload data found for the selected date range")
    
    # Tab 2: Daily Worksheet
    with tab2:
        col1, col2 = st.columns(2)
        with col1:
            view_start_date = st.date_input("From Date", datetime.now(), key="worksheet_start_date")
        with col2:
            view_end_date = st.date_input("To Date", datetime.now(), key="worksheet_end_date")
        
        today = datetime.now().date()
        selected_date = st.date_input("Select Date for Worksheet Entry", today, key="worksheet_entry_date")
        
        worksheet_data = get_faculty_worksheet_data(
            st.session_state.username,
            view_start_date,
            view_end_date
        )
        
        if worksheet_data:
            # Display summary metrics
            st.write("### Daily Summary")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Net Class Conducted", worksheet_data['total_classes'])
            with col2:
                st.metric("Other Activities", worksheet_data['total_hours'])
            with col3:
                total_load = worksheet_data['total_classes'] + worksheet_data['total_hours']
                st.metric("Total Load", f"{total_load}/6")
            
            # Display period-wise activities
            st.write("### Period-wise Activities")
            for period, data in worksheet_data['periods'].items():
                with st.expander(f"Period {period[1]} - {data['time']}", expanded=False):
                    if data['activity_type'] == 'Class':
                        # Display class details
                        st.info(f"""
                            **Subject:** {data['subject']}  
                            **Section(s):** {data['section']}  
                            **Topic:** {data['topic']}  
                            **Load:** {data.get('load', 1.0):.2f}
                        """)
                    else:
                        # Only show activity input form if it's today
                        if selected_date == today:
                            with st.form(key=f"activity_form_{period}"):
                                activity_type = st.selectbox(
                                    "Activity Type",
                                    [
                                        " ",
                                        "Diploma Classwork",
                                        "MCA Claswork",
                                        "BCA Classwork",
                                        "Absentees Phone Follow-up",
                                        "Administrative Work",
                                        "AICTE",
                                        "APSCHE",
                                        "B.Tech / MCA / M.Tech Project / Internship",
                                        "Department Work",
                                        "Department/HoD's Meeting",
                                        "DST",
                                        "Exam Duty",
                                        "Floor Duty",
                                        "IQAC",
                                        "JNTUK",
                                        "MSME",
                                        "NAAC",
                                        "NBA",
                                        "NIRF",
                                        "Parent's Meeting",
                                        "Placement",
                                        "PMKVY",
                                        "SBTET",
                                        "Student Counselling / Mentoring",
                                        "UGC",
                                        "Other"
                                    ],
                                    key=f"activity_type_{period}"
                                )
                                
                                description = st.text_area(
                                    "Description",
                                    value=data['topic'] if data['topic'] != '-' else '',
                                    key=f"description_{period}",
                                    height=100
                                )
                                
                                submitted = st.form_submit_button("Save Activity")
                                if submitted:
                                    if activity_type != "Select Activity" and description:
                                        if update_faculty_activity(
                                            st.session_state.username,
                                            selected_date,
                                            period,
                                            activity_type,
                                            description
                                        ):
                                            st.success("Activity updated successfully!")
                                            st.rerun()
                                    else:
                                        st.error("Please fill in all fields")
                        else:
                            if data['activity_type'] != 'None':
                                st.info(f"""
                                    **Activity:** {data['subject']}  
                                    **Description:** {data['topic']}  
                                    **Load:** {data.get('load', 1.0):.2f}
                                """)
            
            # Download worksheet
            if st.button("Download Daily Worksheet"):
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    worksheet_df = pd.DataFrame([
                        {
                            'Period': period,
                            'Time': data['time'],
                            'Activity': data['subject'],
                            'Section': data['section'],
                            'Topic/Description': data['topic'],
                            'Type': data['activity_type'],
                            'Load': data.get('load', 1.0)
                        }
                        for period, data in worksheet_data['periods'].items()
                    ])
                    
                    worksheet_df.to_excel(writer, sheet_name='Daily Worksheet', index=False)
                    
                    worksheet = writer.sheets['Daily Worksheet']
                    for column in worksheet.columns:
                        max_length = max(len(str(cell.value or '')) for cell in column)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(50, max(12, max_length + 2))
                
                st.download_button(
                    label="📥 Download Worksheet",
                    data=buffer.getvalue(),
                    file_name=f"daily_worksheet_{view_start_date.strftime('%Y%m%d')}_{view_end_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("No worksheet data found for the selected date range")

def view_faculty_worksheet_stats():
    """Display faculty worksheet statistics with filtering and export capabilities"""
    st.subheader("Faculty Worksheet Statistics")
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input(
            "From Date",
            value=datetime.now() - timedelta(days=30),
            key="faculty_stats_start_date"
        )
    with col2:
        end_date = st.date_input(
            "To Date",
            value=datetime.now(),
            key="faculty_stats_end_date"
        )
    
    if st.button("Generate Report", type="primary"):
        try:
            conn = sqlite3.connect('attendance.db')
            cursor = conn.cursor()
            
            # Query to get faculty workload statistics
            query = """
            WITH FacultyClasses AS (
                SELECT 
                    f.name as faculty_name,
                    COUNT(DISTINCT a.date || a.period) as classes_conducted,
                    COUNT(DISTINCT s.merged_section) as unique_sections,
                    COUNT(DISTINCT a.subject) as unique_subjects,
                    SUM(CASE 
                        WHEN a.status = 'P' THEN 1.0 
                        ELSE 0.0 
                    END) / COUNT(*) * 100 as attendance_percentage
                FROM faculty f
                LEFT JOIN attendance a ON f.name = a.faculty
                LEFT JOIN students s ON a.ht_number = s.ht_number
                WHERE a.date BETWEEN ? AND ?
                GROUP BY f.name
            ),
            FacultyActivities AS (
                SELECT 
                    faculty,
                    COUNT(*) as other_activities,
                    COUNT(DISTINCT activity_type) as unique_activities
                FROM faculty_activities
                WHERE date BETWEEN ? AND ?
                GROUP BY faculty
            )
            SELECT 
                fc.faculty_name,
                COALESCE(fc.classes_conducted, 0) as classes_conducted,
                COALESCE(fa.other_activities, 0) as other_activities,
                COALESCE(fc.unique_sections, 0) as unique_sections,
                COALESCE(fc.unique_subjects, 0) as unique_subjects,
                COALESCE(fa.unique_activities, 0) as unique_activities,
                ROUND(COALESCE(fc.attendance_percentage, 0), 2) as avg_attendance,
                COALESCE(fc.classes_conducted, 0) + COALESCE(fa.other_activities, 0) as total_load
            FROM FacultyClasses fc
            LEFT JOIN FacultyActivities fa ON fc.faculty_name = fa.faculty
            ORDER BY total_load DESC
            """
            
            cursor.execute(query, (
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d'),
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            ))
            
            results = cursor.fetchall()
            
            if results:
                # Create DataFrame
                df = pd.DataFrame(results, columns=[
                    'Faculty Name',
                    'Classes Conducted',
                    'Other Activities',
                    'Unique Sections',
                    'Unique Subjects',
                    'Unique Activity Types',
                    'Avg Attendance %',
                    'Total Load'
                ])
                
                # Display summary metrics
                st.write("### Overall Summary")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Faculty", len(df))
                with col2:
                    st.metric("Total Classes", df['Classes Conducted'].sum())
                with col3:
                    st.metric("Total Activities", df['Other Activities'].sum())
                with col4:
                    st.metric("Avg Load", f"{df['Total Load'].mean():.2f}")
                
                # Display detailed statistics
                st.write("### Faculty-wise Statistics")
                st.dataframe(
                    df,
                    column_config={
                        'Faculty Name': st.column_config.TextColumn('Faculty Name', width=200),
                        'Classes Conducted': st.column_config.NumberColumn('Classes', width=100),
                        'Other Activities': st.column_config.NumberColumn('Activities', width=100),
                        'Unique Sections': st.column_config.NumberColumn('Sections', width=100),
                        'Unique Subjects': st.column_config.NumberColumn('Subjects', width=100),
                        'Unique Activity Types': st.column_config.NumberColumn('Activity Types', width=120),
                        'Avg Attendance %': st.column_config.NumberColumn('Attendance %', format="%.2f%%", width=120),
                        'Total Load': st.column_config.NumberColumn('Total Load', format="%.2f", width=100)
                    },
                    hide_index=True,
                    use_container_width=True
                )
                
                # Export to Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Write main statistics
                    df.to_excel(writer, sheet_name='Faculty Statistics', index=False)
                    
                    # Add summary sheet
                    summary_df = pd.DataFrame([{
                        'Metric': 'Total Faculty',
                        'Value': len(df)
                    }, {
                        'Metric': 'Total Classes',
                        'Value': df['Classes Conducted'].sum()
                    }, {
                        'Metric': 'Total Activities',
                        'Value': df['Other Activities'].sum()
                    }, {
                        'Metric': 'Average Load',
                        'Value': df['Total Load'].mean()
                    }])
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Format worksheets
                    for sheet in writer.sheets.values():
                        for column in sheet.columns:
                            max_length = max(len(str(cell.value or '')) for cell in column)
                            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(50, max(12, max_length + 2))
                
                st.download_button(
                    label="📥 Download Complete Report",
                    data=output.getvalue(),
                    file_name=f"faculty_worksheet_stats_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Display visualization
                st.write("### Workload Distribution")
                fig = {
                    'data': [{
                        'type': 'bar',
                        'name': 'Classes',
                        'x': df['Faculty Name'],
                        'y': df['Classes Conducted'],
                    }, {
                        'type': 'bar',
                        'name': 'Activities',
                        'x': df['Faculty Name'],
                        'y': df['Other Activities'],
                    }],
                    'layout': {
                        'barmode': 'stack',
                        'title': 'Faculty Workload Distribution',
                        'xaxis': {'title': 'Faculty'},
                        'yaxis': {'title': 'Count'}
                    }
                }
                st.plotly_chart(fig, use_container_width=True)
                
            else:
                st.info("No data found for the selected date range")
            
        except Exception as e:
            st.error(f"Error generating faculty statistics: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()

import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ... (other imports and functions remain unchanged)

def admin_page():
    st.title("Admin Dashboard")
    
    tab1, tab2, tab3, tab4 = st.tabs(["Manage Students", "Manage Faculty", "View Data", "Faculty Worksheet Stats"])
    
    with tab1:
        st.subheader("Add/Edit Student")
        with st.form("student_form"):
            ht_number = st.text_input("HT Number")
            name = st.text_input("Student Name")
            original_section = st.text_input("Original Section")
            merged_section = st.text_input("Merged Section")
            
            if st.form_submit_button("Add Student"):
                if all([ht_number, name, original_section, merged_section]):
                    add_student(ht_number, name, original_section, merged_section)
                    st.success("Student added successfully!")
                else:
                    st.error("All fields are required!")
    
    with tab2:
        st.subheader("Add/Edit Faculty")
        with st.form("faculty_form"):
            name = st.text_input("Faculty Name")
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            
            if st.form_submit_button("Add Faculty"):
                if all([name, username, password]):
                    add_faculty(name, username, password)
                    st.success("Faculty added successfully!")
                else:
                    st.error("All fields are required!")
    
    with tab3:
        st.subheader("View Database")
        if st.button("View All Students"):
            st.write(get_students())
        if st.button("View All Faculty"):
            faculty_df = get_faculty()
            st.write(faculty_df)  # Now showing all columns including password
    
    with tab4:
        view_faculty_worksheet_stats()

    st.divider()
    st.subheader("Download Database and Export Data")
    col1, col2 = st.columns(2)

    with col1:
        # Download database file
        with open('attendance.db', 'rb') as f:
            db_bytes = f.read()
        st.download_button(
            label="📥 Download Database File",
            data=db_bytes,
            file_name=f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
            mime="application/x-sqlite3"
        )

    with col2:
        # Download complete Excel workbook including sensitive data
        if st.button("📊 Export All Data to Excel (Including Sensitive Info)"):
            try:
                conn = sqlite3.connect('attendance.db')
                output = BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Export students
                    pd.read_sql_query("SELECT * FROM students", conn).to_excel(
                        writer, sheet_name='Students', index=False)
                    
                    # Export faculty (including passwords)
                    pd.read_sql_query("SELECT * FROM faculty", conn).to_excel(
                        writer, sheet_name='Faculty', index=False)
                    
                    # Export attendance
                    pd.read_sql_query("SELECT * FROM attendance", conn).to_excel(
                        writer, sheet_name='Attendance', index=False)
                    
                    # Export section-subject mapping
                    pd.read_sql_query("SELECT * FROM section_subject_mapping", conn).to_excel(
                        writer, sheet_name='Subject Mapping', index=False)
                    
                    # Export faculty activities
                    pd.read_sql_query("SELECT * FROM faculty_activities", conn).to_excel(
                        writer, sheet_name='Faculty Activities', index=False)
                
                st.warning("⚠️ This export contains sensitive information including passwords. Handle with extreme caution!")
                st.download_button(
                    label="📥 Download Complete Excel Export (Sensitive)",
                    data=output.getvalue(),
                    file_name=f"attendance_export_sensitive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"Error exporting data: {str(e)}")
            finally:
                if 'conn' in locals():
                    conn.close()

def main():
    # Initialize the database
    init_db()
    
    # Set page config
    st.set_page_config(
        page_title="RVIT - Attendance Management System",
        page_icon="https://rvit.edu.in/rvitlogo_f.jpg",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Add custom CSS
    st.markdown("""
        <style>
        .stApp {
            max-width: 100%;
            padding: 1rem;
        }
        .stButton button {
            width: 100%;
            padding: 0.8rem !important;
            border-radius: 10px !important;
            margin: 0.5rem 0 !important;
        }
        .student-row {
            background-color: #f0f2f6;
            padding: 1rem;
            border-radius: 10px;
            margin: 0.5rem 0;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Session state management
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
    
    # Main application logic
    if not st.session_state.logged_in:
        login()
    else:
        # Sidebar for navigation
        with st.sidebar:
            st.image("https://rvit.edu.in/rvitlogo_f.jpg", width=100)
            st.write(f"Welcome, {st.session_state.faculty_name}")
            st.divider()
            
            if st.session_state.get('is_admin', False):
                page = "Admin Dashboard"
            else:
                page = st.radio("Navigation", ["Mark Attendance", "View Statistics", 
                                               "Classes Timetable", "My Work Tracker", "Reset Credentials"])
            
            if st.button("Logout", type="primary"):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
        
        # Main content
        if page == "Admin Dashboard" and st.session_state.get('is_admin', False):
            admin_page()
        elif page == "Mark Attendance":
            mark_attendance_page()
        elif page == "View Statistics":
            view_statistics_page()
        elif page == "Classes Timetable":
            show_class_timetable_page()
        elif page == "My Work Tracker":
            show_faculty_classload()
        elif page == "Reset Credentials":
            reset_credentials_page()
        else:
            st.error("You don't have permission to access this page.")

if __name__ == "__main__":
    main()

