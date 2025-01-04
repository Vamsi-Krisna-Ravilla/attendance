import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, timedelta

from io import BytesIO
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io
import pytz
from io import BytesIO




def date_range_selector(key_prefix: str) -> tuple[datetime, datetime]:
    """
    A reusable component for date range selection with single/range toggle
    
    Args:
        key_prefix: Unique prefix for session state keys
    
    Returns:
        tuple: (start_date, end_date)
    """
    col1, col2 = st.columns(2)
    with col1:
        selected_date = st.date_input(
            "Select Date",
            datetime.now(),
            key=f"{key_prefix}_date"
        )
    with col2:
        view_type = st.selectbox(
            "View Type",
            ["Single Day", "Date Range"],
            key=f"{key_prefix}_view_type"
        )
    
    if view_type == "Date Range":
        end_date = st.date_input(
            "End Date",
            selected_date,
            key=f"{key_prefix}_end_date"
        )
    else:
        end_date = selected_date
        
    return selected_date, end_date




# Database Operations
def migrate_db():
    """Add created_at column to existing attendance table with proper SQLite handling"""
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    
    try:
        # Check if created_at column exists
        cursor.execute("PRAGMA table_info(attendance)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'created_at' not in columns:
            # Step 1: Add column without default value
            cursor.execute("ALTER TABLE attendance ADD COLUMN created_at TEXT")
            
            # Step 2: Update existing records with current timestamp
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("UPDATE attendance SET created_at = ?", (current_time,))
            
            # Step 3: Create temporary table with desired schema
            cursor.execute("""
                CREATE TABLE attendance_new (
                    id INTEGER PRIMARY KEY,
                    ht_number TEXT,
                    date TEXT,
                    period TEXT,
                    status TEXT,
                    faculty TEXT,
                    subject TEXT,
                    lesson_plan TEXT,
                    created_at TEXT DEFAULT (datetime('now', 'localtime'))
                )
            """)
            
            # Step 4: Copy data to new table
            cursor.execute("""
                INSERT INTO attendance_new 
                SELECT id, ht_number, date, period, status, faculty, subject, lesson_plan, created_at
                FROM attendance
            """)
            
            # Step 5: Drop old table and rename new table
            cursor.execute("DROP TABLE attendance")
            cursor.execute("ALTER TABLE attendance_new RENAME TO attendance")
            
            conn.commit()
            print("Database migration completed successfully")
    except Exception as e:
        print(f"Migration error: {str(e)}")
        conn.rollback()
    finally:
        conn.close()


def add_student(ht_number, name, original_section, merged_section):
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO students VALUES (?, ?, ?, ?)", 
              (ht_number, name, original_section, merged_section))
    conn.commit()
    conn.close()

def add_faculty(name, username, password):
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO faculty (name, username, password) VALUES (?, ?, ?)", 
              (name, username, password))
    conn.commit()
    conn.close()


def get_students(section=None):
    conn = sqlite3.connect('attendance.db')
    if section:
        df = pd.read_sql_query("SELECT * FROM students WHERE merged_section = ?", conn, params=(section,))
    else:
        df = pd.read_sql_query("SELECT * FROM students", conn)
    conn.close()
    return df

def get_faculty():
    conn = sqlite3.connect('attendance.db')
    df = pd.read_sql_query("SELECT * FROM faculty", conn)
    conn.close()
    return df

def get_attendance(start_date, end_date=None, section=None):
    conn = sqlite3.connect('attendance.db')
    query = "SELECT * FROM attendance WHERE date >= ?"
    params = [start_date]
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    if section:
        query += " AND ht_number IN (SELECT ht_number FROM students WHERE merged_section = ?)"
        params.append(section)
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def get_section_subjects(section):
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT subject_name FROM section_subject_mapping WHERE section = ? ORDER BY subject_name", (section,))
    subjects = cursor.fetchall()
    conn.close()
    return [subject[0] for subject in subjects]

# Page Functions
def login():
    st.title("Login")
    
    # Center the login form
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.image("https://rvit.edu.in/rvitlogo_f.jpg", width=200)
        login_type = st.radio("Select Login Type", ["Faculty", "Admin"])
        username = st.text_input("Username").strip()  # Remove whitespace
        password = st.text_input("Password", type="password").strip()  # Remove whitespace
        
        if st.button("Login", type="primary", use_container_width=True):
            conn = sqlite3.connect('attendance.db')
            cursor = conn.cursor()
            
            # Use parameterized query to prevent SQL injection
            cursor.execute("""
                SELECT * FROM faculty 
                WHERE username = ? AND password = ?
            """, (username, password))
            
            user = cursor.fetchone()
            conn.close()
            
            if user:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.session_state.faculty_name = user[1]  # faculty name is in index 1
                
                # Check if the user is an admin
                if login_type == "Admin" and username == "rvk":
                    st.session_state.is_admin = True
                else:
                    st.session_state.is_admin = False
                
                st.rerun()
            else:
                st.error("Invalid credentials")



def get_last_attendance_pattern(section, faculty=None):
    """Get the last attendance pattern for a section regardless of faculty"""
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    try:
        # Get the most recent date and period for this section
        # Removed faculty filter to allow patterns from any faculty
        query = """
            SELECT DISTINCT a.date, a.period, a.faculty
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE s.merged_section = ? 
            ORDER BY a.date DESC, a.period DESC
            LIMIT 1
        """
        cursor.execute(query, (section,))
        last_record = cursor.fetchone()
        
        if not last_record:
            return {}
            
        date, period, marking_faculty = last_record
        
        # Get the attendance pattern for that date and period
        query = """
            SELECT a.ht_number, a.status
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE s.merged_section = ?
            AND a.date = ?
            AND a.period = ?
        """
        cursor.execute(query, (section, date, period))
        pattern = {row[0]: row[1] for row in cursor.fetchall()}
        
        # If pattern found, log which faculty's pattern is being used
        if pattern:
            st.info(f"Using attendance pattern from {marking_faculty}'s class on {date}")
            
        return pattern
    except Exception as e:
        st.error(f"Error getting attendance pattern: {str(e)}")
        return {}
    finally:
        conn.close()



def check_attendance_exists(section, date, period):
    """Check if attendance has already been marked for given section, date and period"""
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    try:
        # Modified query to also get the faculty name who marked attendance
        query = """
            SELECT DISTINCT a.faculty 
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE s.merged_section = ? 
            AND a.date = ? 
            AND a.period = ?
            LIMIT 1
        """
        cursor.execute(query, (section, date, period))
        result = cursor.fetchone()
        if result:
            return True, result[0]  # Return both status and faculty name
        return False, None
    finally:
        conn.close()



def mark_attendance_page():
    """Enhanced attendance marking page with improved UI and mobile responsiveness"""
    """Page for marking attendance"""
    st.title("Mark Attendance")
    
    # Get current time in IST for the attendance page
    current_time = get_current_time_ist()
    
    # Custom CSS for better mobile UI
    st.markdown("""
        <style>
        /* Current Session Header */
        .current-session {
            background: linear-gradient(135deg, #6B46C1 0%, #553C9A 100%);
            padding: 1.5rem;
            border-radius: 1rem;
            margin-bottom: 1.5rem;
            color: white;
        }
        
        /* Force columns to stay side by side */
        div.row-widget.stHorizontal > div[data-testid="column"] {
            width: auto !important;
            flex: none !important;
        }
        
        div.row-widget.stHorizontal > div[data-testid="column"]:first-child {
            width: 70% !important;
            flex: 0 0 70% !important;
        }
        
        div.row-widget.stHorizontal > div[data-testid="column"]:last-child {
            width: 30% !important;
            flex: 0 0 30% !important;
        }
        
        /* Student Entry */
        .student-entry {
            background: #1E1E1E;
            border-bottom: 1px solid #2D2D2D;
            padding: 0.75rem;
            height: 100%;
            display: flex;
            flex-direction: column;
            justify-content: center;
        }
        
        .student-info {
            display: flex;
            flex-direction: column;
            gap: 0.25rem;
        }
        
        .student-name {
            color: #FF1493;
            font-size: 0.8rem;
            font-weight: 500;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        
        .student-id {
            color: #FFA500;
            font-size: 1.0rem;
        }
        
        .section-label {
            color: #808080;
            font-size: 0.75rem;
        }
        
        /* Checkbox styling */
        div[data-testid="column"] [data-testid="stCheckbox"] {
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
            height: 100% !important;
            margin: 0 !important;
            padding: 0.75rem !important;
        }
        
        .stCheckbox > label {
            min-width: 44px;
            min-height: 44px;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }
        
        /* Remove default padding and gaps */
        div[data-testid="column"] {
            padding: 0 !important;
        }
        
        div.row-widget.stHorizontal {
            flex-wrap: nowrap !important;
            gap: 0 !important;
        }
        
        /* Ensure form elements don't break layout */
        .stForm > div[data-testid="stForm"] {
            width: 100% !important;
        }
        
        /* Force mobile layout to match desktop */
        @media (max-width: 768px) {
            div.row-widget.stHorizontal {
                display: flex !important;
                flex-direction: row !important;
            }
            
            div.row-widget.stHorizontal > div[data-testid="column"] {
                min-width: 0 !important;
            }
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Get unique sections
    students_df = get_students()
    if students_df.empty:
        st.warning("No students found in the database. Please add students first.")
        return
        
    sections = students_df['merged_section'].unique()
    
    # Selection controls in a card
    st.markdown('<div class="current-session">', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        section = st.selectbox(
            "Select Section",
            [""] + list(sections),  # Add empty string as first option
            index=0,  # Set index to 0 to select the empty option
            key="section_select"
        )
    with col2:
        period = st.selectbox(
            "Select Period",
            [""] + ['P1', 'P2', 'P3', 'P4', 'P5', 'P6'],
            index=0,
            key="period_select"
        )
    with col3:
        # Only show subjects if a section is selected
        if section:
            subject = st.selectbox(
                "Select Subject",
                [""] + get_section_subjects(section),
                index=0,
                key="subject_select"
            )
        else:
            subject = st.selectbox(
                "Select Subject",
                [""],  # Only show empty option if no section selected
                index=0,
                key="subject_select"
            )
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Check if attendance already exists
    today = datetime.now().strftime('%Y-%m-%d')
    exists, faculty = check_attendance_exists(section, today, period)
    if exists:
        st.error(f"Attendance has already been marked for {section} - {period} today by {faculty}!")
        return
    
    if section and period and subject:
        students = get_students(section)
        
        # Action buttons
        col1, col2, col3 = st.columns(3)
        with col1:
            mark_all_present = st.button("✓ Mark All Present", 
                                       type="primary",
                                       key="mark_all_present",
                                       use_container_width=True)
        with col2:
            mark_all_absent = st.button("✗ Mark All Absent", 
                                      key="mark_all_absent",
                                      use_container_width=True)
        with col3:
            use_last_pattern = st.button("↺ Use Last Pattern", 
                                       key="use_last_pattern",
                                       use_container_width=True)
        
        # Initialize attendance data
        if 'attendance_data' not in st.session_state:
            st.session_state.attendance_data = {}
        
        # Update attendance data based on button clicks
        if mark_all_present:
            for student in students.itertuples():
                st.session_state.attendance_data[student.ht_number] = 'P'
        elif mark_all_absent:
            for student in students.itertuples():
                st.session_state.attendance_data[student.ht_number] = 'A'
        elif use_last_pattern:
            pattern = get_last_attendance_pattern(section, st.session_state.faculty_name)
            if pattern:
                st.session_state.attendance_data = pattern
                st.success("Last attendance pattern applied!")
            else:
                st.info("No previous attendance pattern found for this section.")
        
        # Create a form for student attendance
        with st.form(key='attendance_form'):
            for student in students.itertuples():
                # Create two columns with 70-30 split
                cols = st.columns([0.7, 0.3])
                
                # Student info in first column
                with cols[0]:
                    st.markdown(f"""
                        <div class="student-entry">
                            <div class="student-info">
                                <span class="student-name">{student.name}</span>
                                <span class="student-id">{student.ht_number}</span>
                                <span class="section-label">{student.original_section}</span>
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
                
                # Attendance checkbox in second column
                with cols[1]:
                    default_value = st.session_state.attendance_data.get(student.ht_number, 'P') == 'P'
                    status = st.checkbox("Present", 
                                      value=default_value,
                                      key=f"attendance_{student.ht_number}")
                    st.session_state.attendance_data[student.ht_number] = 'P' if status else 'A'
            
            # Lesson plan input
            st.markdown("<br>", unsafe_allow_html=True)
            lesson_plan = st.text_area("📝 Lesson Plan", 
                                     placeholder="Enter the topic covered in this class...",
                                     height=100)
            
            # Submit button
            submit_button = st.form_submit_button("Submit Attendance", 
                                                type="primary",
                                                use_container_width=True)
        
        if submit_button:
            if not lesson_plan.strip():
                st.error("Please enter a lesson plan before submitting attendance.")
            else:
                # Double check for duplicate attendance
                exists, faculty = check_attendance_exists(section, today, period)
                if exists:
                    st.error(f"Attendance has already been marked for {section} - {period} today by {faculty}!")
                    return
                    
                date = datetime.now().strftime('%Y-%m-%d')
                for ht_number, status in st.session_state.attendance_data.items():
                    mark_attendance(ht_number, date, period, status, 
                                 st.session_state.faculty_name, subject, lesson_plan)
                st.success("Attendance marked successfully!")
                st.session_state.attendance_data = {}
                st.rerun()







# Add these functions after your existing database functions

def get_overall_statistics(start_date, end_date, sections=None):
    """Enhanced overall statistics with detailed analytics"""
    conn = sqlite3.connect('attendance.db')
    try:
        # Base query with proper date filtering and section handling
        query = """
            WITH DateRangeAttendance AS (
                SELECT 
                    s.ht_number,
                    s.name,
                    s.original_section,
                    a.subject,
                    a.date,
                    a.period,
                    a.status,
                    strftime('%H', a.created_at) as hour,
                    a.created_at
                FROM students s
                JOIN attendance a ON s.ht_number = a.ht_number
                WHERE date(a.date) BETWEEN date(?) AND date(?)
        """
        
        params = [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')]
        
        # Add section filter if sections are provided
        if sections:
            section_placeholders = ','.join('?' * len(sections))
            query += f" AND s.original_section IN ({section_placeholders})"
            params.extend(sections)
            
        # Close the CTE and select from it
        query += """
            )
            SELECT * FROM DateRangeAttendance
        """
        
        # Execute query and create DataFrame
        df = pd.read_sql_query(query, conn, params=params)
        
        if df.empty:
            return {
                'student_stats': pd.DataFrame(),
                'total_students': 0,
                'avg_attendance': 0,
                'below_75': 0,
                'day_wise_stats': pd.DataFrame(),
                'subject_wise_stats': pd.DataFrame(),
                'time_slot_analysis': pd.DataFrame(),
                'trend_analysis': pd.DataFrame(),
                'section_comparison': pd.DataFrame()
            }
            
        # 1. Student-wise Statistics
        student_stats = df.groupby(['ht_number', 'name', 'original_section']).agg({
            'status': lambda x: sum(x == 'P'),
            'date': 'count'
        }).reset_index()
        
        student_stats.columns = ['HT Number', 'Student Name', 'Section', 'Present', 'Total']
        student_stats['Percentage'] = (student_stats['Present'] / student_stats['Total'] * 100).round(2)
        student_stats['Classes (A/C)'] = student_stats.apply(
            lambda x: f"{int(x['Present'])}/{int(x['Total'])}", 
            axis=1
        )
        
        # 2. Day-wise Analysis
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        df['day_name'] = pd.to_datetime(df['date']).dt.day_name()
        day_wise = df.groupby('day_name').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2)
        }).reindex(days).reset_index()
        day_wise.columns = ['Day', 'Attendance %']
        
        # 3. Subject-wise Analysis
        subject_wise = df.groupby('subject').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2),
            'ht_number': 'count'
        }).reset_index()
        subject_wise.columns = ['Subject', 'Attendance %', 'Total Classes']
        
        # 4. Time Slot Analysis - Fixed to use proper hour column
        df['hour'] = pd.to_numeric(df['hour'], errors='coerce')  # Convert hour to numeric
        time_slot = df.groupby(['period', 'hour']).agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2)
        }).reset_index()
        time_slot.columns = ['Period', 'Hour', 'Attendance %']
        
        # 5. Trend Analysis
        trend = df.groupby('date').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2)
        }).reset_index()
        trend.columns = ['Date', 'Attendance %']
        
        # 6. Section Comparison
        section_comp = df.groupby('original_section').agg({
            'status': lambda x: (sum(x == 'P') / len(x) * 100).round(2),
            'ht_number': lambda x: len(set(x))
        }).reset_index()
        section_comp.columns = ['Section', 'Attendance %', 'Student Count']
        
        # Calculate summary statistics
        total_students = len(student_stats)
        avg_attendance = student_stats['Percentage'].mean()
        below_75 = len(student_stats[student_stats['Percentage'] < 75])
        
        return {
            'student_stats': student_stats,
            'total_students': total_students,
            'avg_attendance': avg_attendance,
            'below_75': below_75,
            'day_wise_stats': day_wise,
            'subject_wise_stats': subject_wise,
            'time_slot_analysis': time_slot,
            'trend_analysis': trend,
            'section_comparison': section_comp
        }
        
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return None
    finally:
        conn.close()



def get_student_report(ht_number, start_date, end_date):
    """Modified to properly handle date range filtering"""
    conn = sqlite3.connect('attendance.db')
    try:
        # Convert dates to proper format
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
        
        # Get student details
        student_query = "SELECT * FROM students WHERE ht_number = ?"
        student_df = pd.read_sql_query(student_query, conn, params=[ht_number])
        
        # Get attendance details with proper date filtering
        attendance_query = """
            SELECT 
                subject,
                COUNT(DISTINCT date || period) as total_classes,
                SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) as present_count,
                ROUND(CAST(SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) AS FLOAT) / 
                    NULLIF(COUNT(*), 0) * 100, 2) as attendance_percentage
            FROM attendance 
            WHERE ht_number = ? 
            AND date(date) BETWEEN date(?) AND date(?)
            GROUP BY subject
        """
        attendance_df = pd.read_sql_query(
            attendance_query, 
            conn, 
            params=[
                ht_number, 
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            ]
        )
        
        # Get daily attendance with proper date filtering
        daily_query = """
            SELECT 
                date,
                period,
                subject,
                status,
                faculty,
                lesson_plan,
                created_at as submission_time
            FROM attendance 
            WHERE ht_number = ? 
            AND date(date) BETWEEN date(?) AND date(?)
            ORDER BY date DESC, period ASC
        """
        daily_df = pd.read_sql_query(
            daily_query, 
            conn, 
            params=[
                ht_number,
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            ]
        )
        
        return student_df, attendance_df, daily_df
        
    except Exception as e:
        st.error(f"Error generating student report: {str(e)}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    finally:
        conn.close()

def get_subject_analysis(start_date, end_date, section=None):
    """Modified to properly handle date range filtering"""
    conn = sqlite3.connect('attendance.db')
    try:
        # Convert dates to proper format
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
        
        query = """
            SELECT 
                subject,
                COUNT(DISTINCT date || period) as total_classes,
                SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) as present_count,
                COUNT(*) as total_attendance,
                ROUND(CAST(SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) AS FLOAT) / 
                    NULLIF(COUNT(*), 0) * 100, 2) as attendance_percentage
            FROM attendance 
            WHERE date(date) BETWEEN date(?) AND date(?)
        """
        params = [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')]
        
        if section and section != 'All':
            query += " AND ht_number IN (SELECT ht_number FROM students WHERE merged_section = ?)"
            params.append(section)
        
        query += " GROUP BY subject ORDER BY subject"
        
        df = pd.read_sql_query(query, conn, params=params)
        return df
        
    except Exception as e:
        st.error(f"Error generating subject analysis: {str(e)}")
        return pd.DataFrame()
    finally:
        conn.close()



def export_to_excel(df, filename):
    """Export DataFrame to Excel with proper formatting"""
    # Create Excel writer object
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance Report')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Attendance Report']
        
        # Format headers
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        
        # Apply formatting
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 15)
    
    return output.getvalue()

def view_statistics_page():
    """Enhanced statistics page with course-wise filtering and Excel export"""
    
    # Tab layout
    tab1, tab2, tab3 = st.tabs(["📊 Overall Statistics", "👤 Student Reports", "📚 Subject Analysis"])
    
    with tab1:
        st.subheader("Overall Attendance Statistics")
        
        # Use the new date range selector
        selected_date, end_date = date_range_selector("stats")
        
        # Get unique courses and sections
        students_df = get_students()
        if students_df.empty:
            st.warning("No students found in the database.")
            return
        
        # Course selection (e.g., "B.Tech-II")
        courses = sorted(list(set([
            section.split('-')[0] 
            for section in students_df['original_section'].unique()
        ])))
        selected_course = st.selectbox("Select Course", courses)
        
        # Get sections for selected course
        sections = sorted(list(students_df[
            students_df['original_section'].str.startswith(selected_course)
        ]['original_section'].unique()))
        
        # Add "All" option at the beginning of the list
        sections = ['All'] + sections
        
        # Multi-select for sections with "All" as default
        selected_sections = st.multiselect(
            "Select Sections",
            sections,
            default=['All'],
            key="section_multiselect"
        )
        
        # Handle "All" selection
        if 'All' in selected_sections:
            selected_sections = sections[1:]  # All sections except "All"
        
        if selected_sections and st.button("Generate Report", type="primary"):
            try:
                conn = sqlite3.connect('attendance.db')
                
                # Modified query to ensure correct attendance counting
                query = """
                    WITH AllStudents AS (
                        -- Get all students in selected sections
                        SELECT ht_number, name, original_section
                        FROM students
                        WHERE original_section IN ({})
                    ),
                    SectionClasses AS (
                        -- Get conducted classes per section
                        SELECT 
                            s.original_section,
                            a.date,
                            a.period,
                            a.subject
                        FROM attendance a
                        JOIN students s ON a.ht_number = s.ht_number
                        WHERE a.date BETWEEN ? AND ?
                        AND s.original_section IN ({})
                        GROUP BY s.original_section, a.date, a.period, a.subject
                    ),
                    StudentAttendance AS (
                        -- Calculate attendance per student
                        SELECT 
                            s.ht_number,
                            s.name,
                            s.original_section as section,
                            COUNT(DISTINCT sc.date || sc.period) as conducted,
                            COUNT(DISTINCT CASE WHEN a.status = 'P' 
                                THEN a.date || a.period 
                                ELSE NULL END) as attended
                        FROM AllStudents s
                        LEFT JOIN SectionClasses sc ON s.original_section = sc.original_section
                        LEFT JOIN attendance a ON s.ht_number = a.ht_number 
                            AND sc.date = a.date 
                            AND sc.period = a.period
                        GROUP BY s.ht_number, s.name, s.original_section
                    )
                    SELECT 
                        ht_number as "HT Number",
                        name as "Student Name",
                        section as "Section",
                        attended || '/' || conducted as "Classes (A/C)",
                        ROUND(CAST(attended AS FLOAT) / NULLIF(conducted, 0) * 100, 2) as "Overall %"
                    FROM StudentAttendance
                    ORDER BY "Section", "HT Number"
                """.format(
                    ','.join(['?'] * len(selected_sections)),
                    ','.join(['?'] * len(selected_sections))
                )
                
                params = [
                    *selected_sections,
                    selected_date.strftime('%Y-%m-%d'),
                    end_date.strftime('%Y-%m-%d'),
                    *selected_sections
                ]
                
                df = pd.read_sql_query(query, conn, params=params)
                
                if not df.empty:
                    # Calculate summary metrics
                    total_students = len(df)
                    avg_attendance = df['Overall %'].mean()
                    below_75 = len(df[df['Overall %'] < 75])
                    
                    # Display metrics
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Students", total_students)
                    with col2:
                        st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                    with col3:
                        st.metric("Students Below 75%", below_75)
                    
                    # Display student-wise statistics
                    st.subheader("Student-wise Statistics")
                    st.dataframe(df)
                    
                    # Export to Excel
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Attendance Report', index=False)
                        
                        workbook = writer.book
                        worksheet = writer.sheets['Attendance Report']
                        
                        # Format headers
                        for col_num, value in enumerate(df.columns.values):
                            cell = worksheet.cell(1, col_num + 1)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color='D3D3D3',
                                end_color='D3D3D3',
                                fill_type='solid'
                            )
                        
                        # Adjust column widths
                        for idx, col in enumerate(df.columns):
                            worksheet.column_dimensions[get_column_letter(idx + 1)].width = 15
                    
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=output.getvalue(),
                        file_name=f"attendance_report_{selected_course}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.info("No attendance records found for the selected criteria.")
                
            except Exception as e:
                st.error(f"Error generating report: {str(e)}")
            finally:
                conn.close()

    with tab2:
        st.subheader("Student Reports")
        
        # Student selection
        student_ht = st.selectbox(
            "Select Student",
            students_df['ht_number'].tolist(),
            format_func=lambda x: (
                f"{x} - {students_df[students_df['ht_number'] == x]['name'].iloc[0]}"
            )
        )
        
        # Date range for student report
        # Use the new date range selector for student report
        student_start_date, student_end_date = date_range_selector("student_report")
        
        if st.button("Generate Student Report", type="primary"):
            student_df, attendance_df, daily_df = get_student_report(
                student_ht,
                student_start_date.strftime('%Y-%m-%d'),
                student_end_date.strftime('%Y-%m-%d')
            )
            
            if not student_df.empty:
                # Student Information
                st.write("### Student Information")
                st.write(f"Name: {student_df['name'].iloc[0]}")
                st.write(f"Section: {student_df['original_section'].iloc[0]}")
                
                if not attendance_df.empty:
                    # Format attendance data
                    attendance_df['Classes'] = attendance_df.apply(
                        lambda x: f"{int(x['present_count'])}/{int(x['total_classes'])}",
                        axis=1
                    )
                    
                    # Display subject-wise attendance
                    st.write("### Subject-wise Attendance")
                    st.dataframe(attendance_df[
                        ['subject', 'Classes', 'attendance_percentage']
                    ].rename(columns={
                        'Classes': 'Classes (Attended/Conducted)',
                        'attendance_percentage': 'Attendance %'
                    }))
                
                if not daily_df.empty:
                    st.write("### Daily Attendance Log")
                    st.dataframe(daily_df)
                    
                    # Export student report to Excel
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        daily_df.to_excel(
                            writer,
                            sheet_name='Daily Log',
                            index=False
                        )
                        attendance_df.to_excel(
                            writer,
                            sheet_name='Subject Summary',
                            index=False
                        )
                    
                    st.download_button(
                        label="📥 Download Student Report",
                        data=output.getvalue(),
                        file_name=(
                            f"student_report_{student_ht}_"
                            f"{datetime.now().strftime('%Y%m%d')}.xlsx"
                        ),
                        mime=(
                            "application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet"
                        )
                    )
                else:
                    st.info("No attendance records found for the selected date range.")
            else:
                st.error("Student not found in the database.")
    
    with tab3:
        st.subheader("Subject Analysis")
        
        # Date range for subject analysis
        # Use the new date range selector for subject analysis
        subject_start_date, subject_end_date = date_range_selector("subject_analysis")
        
        # Section selection for subject analysis
        subject_section = st.selectbox(
            "Select Section",
            ['All'] + list(students_df['original_section'].unique()),
            key="section_subject"
        )
        
        if st.button("Generate Subject Analysis", type="primary"):
            subject_df = get_subject_analysis(
                subject_start_date.strftime('%Y-%m-%d'),
                subject_end_date.strftime('%Y-%m-%d'),
                None if subject_section == 'All' else subject_section
            )
            
            if not subject_df.empty:
                # Display subject-wise analysis
                st.write("### Subject-wise Attendance Analysis")
                
                # Create bar chart
                st.bar_chart(subject_df.set_index('subject')['attendance_percentage'])
                
                # Display detailed statistics
                st.write("### Detailed Subject Statistics")
                st.dataframe(subject_df)
                
                # Export subject analysis to Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    subject_df.to_excel(
                        writer,
                        sheet_name='Subject Analysis',
                        index=False
                    )
                
                st.download_button(
                    label="📥 Download Subject Analysis",
                    data=output.getvalue(),
                    file_name=(
                        f"subject_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
                    ),
                    mime=(
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    )
                )
            else:
                st.info("No attendance records found for subject analysis.")




def reset_credentials_page():
    st.subheader("Reset Credentials")
    
    with st.form("reset_credentials_form"):
        current_password = st.text_input("Current Password", type="password")
        new_password = st.text_input("New Password", type="password")
        confirm_password = st.text_input("Confirm New Password", type="password")
        
        submitted = st.form_submit_button("Reset Password", type="primary")
        
        if submitted:
            if not all([current_password, new_password, confirm_password]):
                st.error("All fields are required")
                return
                
            if new_password != confirm_password:
                st.error("New passwords do not match")
                return
                
            conn = sqlite3.connect('attendance.db')
            try:
                cursor = conn.cursor()
                
                # Verify current password
                cursor.execute("""
                    SELECT * FROM faculty 
                    WHERE username = ? AND password = ?
                """, (st.session_state.username, current_password))
                
                if cursor.fetchone():
                    # Update password
                    cursor.execute("""
                        UPDATE faculty 
                        SET password = ? 
                        WHERE username = ?
                    """, (new_password, st.session_state.username))
                    
                    conn.commit()
                    st.success("Password updated successfully! Please login again.")
                    
                    # Clear session state to force re-login
                    for key in list(st.session_state.keys()):
                        del st.session_state[key]
                    st.rerun()
                else:
                    st.error("Current password is incorrect")
                    
            except Exception as e:
                st.error(f"Error resetting password: {str(e)}")
            finally:
                conn.close()
                

def get_faculty_classload(username, start_date=None, end_date=None):
    """Modified to handle cases where created_at might not exist"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Get faculty name
        cursor.execute("SELECT name FROM faculty WHERE username = ?", (username,))
        faculty_name = cursor.fetchone()[0]
        
        # Base parameters
        params = [faculty_name]
        
        # Prepare date filters if provided
        date_filter = ""
        if start_date and end_date:
            date_filter = "AND a.date >= ? AND a.date <= ?"
            params.extend([start_date, end_date])
        
        # Modified query to handle missing created_at
        query = f"""
            WITH PeriodGroups AS (
                SELECT 
                    date,
                    period,
                    COUNT(DISTINCT subject) as subject_count
                FROM (
                    SELECT DISTINCT
                        a.date,
                        a.period,
                        a.subject
                    FROM attendance a
                    WHERE a.faculty = ?
                    {date_filter}
                )
                GROUP BY date, period
            ),
            DetailedClasses AS (
                SELECT 
                    a.date,
                    a.period,
                    a.subject,
                    GROUP_CONCAT(DISTINCT s.merged_section) as sections,
                    COUNT(DISTINCT s.merged_section) as section_count,
                    a.lesson_plan,
                    MIN(a.id) as first_entry_id,
                    COUNT(DISTINCT a.ht_number) as student_count,
                    pg.subject_count,
                    COALESCE(MIN(a.created_at), datetime('now', 'localtime')) as submission_time
                FROM attendance a
                JOIN students s ON a.ht_number = s.ht_number
                JOIN PeriodGroups pg ON a.date = pg.date AND a.period = pg.period
                WHERE a.faculty = ?
                {date_filter}
                GROUP BY a.date, a.period, a.subject
            )
            SELECT 
                d.date,
                d.period,
                d.subject,
                d.sections,
                d.section_count,
                ROUND(1.0/d.subject_count, 2) as distributed_load,
                d.lesson_plan,
                d.student_count,
                time(d.submission_time) as submission_time
            FROM DetailedClasses d
            ORDER BY d.date DESC, d.period, d.subject
        """
        
        cursor.execute(query, params + params)
        rows = cursor.fetchall()
        
        if not rows:
            return pd.DataFrame()
            
        # Create DataFrame with submission times
        classload_data = []
        for row in rows:
            date, period, subject, sections, section_count, dist_load, lesson_plan, student_count, submission_time = row
            
            entry = {
                'Date': date,
                'Period': period,
                'Subject': subject,
                'Section': sections,
                'Combined Sections': sections,
                'Distributed Load': dist_load,
                'Student Count': student_count,
                'Submission Time': submission_time or 'Not recorded',
                'Lesson Plan': lesson_plan
            }
            classload_data.append(entry)
        
        df = pd.DataFrame(classload_data)
        
        if not df.empty:
            df['Total Load'] = df.groupby(['Date', 'Period'])['Distributed Load'].transform('sum')
            df['Day'] = pd.to_datetime(df['Date']).dt.strftime('%A')
            df = df.sort_values(['Date', 'Period', 'Subject'])
            df['Cumulative Load'] = df['Distributed Load'].cumsum()
        
        return df
        
    except Exception as e:
        print(f"Error calculating faculty classload: {str(e)}")
        return pd.DataFrame()
    finally:
        if conn:
            conn.close()

def get_missing_attendance_sections(date, period):
    """Get sections that haven't had attendance marked for a given period and date"""
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    try:
        # Modified query to remove is_active check and use proper table structure
        cursor.execute("""
            WITH AllSections AS (
                SELECT DISTINCT section
                FROM section_subject_mapping
            ),
            MarkedSections AS (
                SELECT DISTINCT s.merged_section
                FROM attendance a
                JOIN students s ON a.ht_number = s.ht_number
                WHERE a.date = ? AND a.period = ?
            )
            SELECT section 
            FROM AllSections 
            WHERE section NOT IN (SELECT * FROM MarkedSections)
            ORDER BY section
        """, (date, period))
        
        missing_sections = [row[0] for row in cursor.fetchall()]
        return missing_sections
    except Exception as e:
        print(f"Error getting missing sections: {str(e)}")
        return []
    finally:
        conn.close()



def view_class_timetable():
    st.title("Class Timetable")
    
    # Date selection
    col1, col2 = st.columns([3, 1])
    with col1:
        selected_date = st.date_input("Select Date", datetime.now())
    with col2:
        view_type = st.selectbox("View Type", ["Single Day", "Week View"], index=0)
    
    # Convert date to string format
    date_str = selected_date.strftime('%Y-%m-%d')
    
    # Get timetable data
    conn = sqlite3.connect('attendance.db')
    df = pd.read_sql_query("""
        SELECT 
            a.date,
            a.period,
            a.faculty,
            a.subject,
            s.merged_section as section,
            COUNT(DISTINCT a.ht_number) as students,
            MAX(a.created_at) as time,
            'Completed' as status
        FROM attendance a
        JOIN students s ON a.ht_number = s.ht_number
        WHERE a.date = ?
        GROUP BY a.date, a.period, a.faculty, a.subject, s.merged_section
        ORDER BY a.period, s.merged_section
    """, conn, params=[date_str])
    
    # Daily summary
    st.header("Daily Summary")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total Classes", len(df))
    with col2:
        st.metric("Faculty Engaged", df['faculty'].nunique())
    with col3:
        st.metric("Sections Covered", df['section'].nunique())
    
    # Class Timetable with Missing Attendance Dropdowns
    st.header("Class Timetable")
    
    # Get all periods
    periods = ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']
    
    # Create tabs for each period
    for period in periods:
        # Filter data for current period
        period_data = df[df['period'] == period]
        
        # Create expander for the period
        with st.expander(f"Period {period}", expanded=False):
            # Create two columns
            left_col, right_col = st.columns([3, 2])
            
            with left_col:
                if not period_data.empty:
                    st.write("📚 Completed Classes:")
                    st.dataframe(
                        period_data[['faculty', 'subject', 'section', 'students', 'time', 'status']],
                        hide_index=True,
                        use_container_width=True
                    )
                else:
                    st.info("No completed classes for this period")
            
            with right_col:
                # Get and show missing attendance sections
                missing_sections = get_missing_attendance_sections(date_str, period)
                if missing_sections:
                    st.write("⚠️ Missing Attendance:")
                    for section in missing_sections:
                        st.warning(section, icon="⚠️")
                else:
                    st.success("✅ All sections marked!", icon="✅")
    
    conn.close()






def show_class_timetable_page():
    """Modified to use the new date range selector"""
    st.subheader("Class Timetable")
    
    # Use the new date range selector
    selected_date, end_date = date_range_selector("class_timetable")
    
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Modified query to handle missing created_at
        query = """
            SELECT DISTINCT
                a.date,
                a.period,
                a.faculty,
                a.subject,
                s.merged_section as section,
                MIN(a.id) as first_entry_id,
                COUNT(DISTINCT a.ht_number) as student_count,
                COALESCE(MIN(a.created_at), datetime('now', 'localtime')) as submission_time
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE a.date BETWEEN ? AND ?
            GROUP BY a.date, a.period, a.faculty, a.subject, s.merged_section
            ORDER BY a.date, a.period, s.merged_section
        """
        
        cursor.execute(query, (
            selected_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d')
        ))
        
        classes_data = []
        for row in cursor.fetchall():
            date, period, faculty, subject, section, entry_id, student_count, submission_time = row
            
            # Format submission time
            try:
                time_str = datetime.strptime(submission_time, '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
            except (ValueError, TypeError):
                time_str = 'Not recorded'
            
            classes_data.append({
                'Date': date,
                'Period': period,
                'Faculty': faculty,
                'Subject': subject,
                'Section': section,
                'Students': student_count,
                'Submission Time': time_str,
                'Status': 'Completed'
            })
        
        if classes_data:
            df_classes = pd.DataFrame(classes_data)
            
            # Display summary metrics
            st.write("### Daily Summary")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Classes", len(df_classes))
            with col2:
                st.metric("Faculty Engaged", df_classes['Faculty'].nunique())
            with col3:
                st.metric("Sections Covered", df_classes['Section'].nunique())
            
            # Display timetable with submission time
            st.write("### Class Timetable")
            st.dataframe(
                df_classes,
                column_config={
                    'Date': st.column_config.TextColumn('Date', width=100),
                    'Period': st.column_config.TextColumn('Period', width=70),
                    'Faculty': st.column_config.TextColumn('Faculty', width=150),
                    'Subject': st.column_config.TextColumn('Subject', width=120),
                    'Section': st.column_config.TextColumn('Section', width=120),
                    'Students': st.column_config.NumberColumn('Students', width=100),
                    'Submission Time': st.column_config.TextColumn('Time', width=100),
                    'Status': st.column_config.TextColumn('Status', width=100)
                },
                hide_index=True,
                use_container_width=True
            )
            
        else:
            if view_type == "Date Range":
                st.info(f"No classes recorded between {selected_date.strftime('%Y-%m-%d')} and {end_date.strftime('%Y-%m-%d')}")
            else:
                st.info(f"No classes recorded on {selected_date.strftime('%Y-%m-%d')}")
            
    except Exception as e:
        st.error(f"Error loading class timetable: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

    


def show_faculty_classload():
    """Enhanced faculty classload page with worksheet capabilities"""
    st.subheader("Faculty Workload Dashboard")
    
    # Create tabs for different views
    tab1, tab2 = st.tabs(["📊 Overall Workload", "📝 Daily Worksheet"])
    
    # Single date range selector for both tabs
    # selected_date, end_date = date_range_selector("faculty_workload")
    
    # Tab 1: Overall Workload
    with tab1:
        selected_date, end_date = date_range_selector("faculty_workload")
        if st.button("Generate Workload Report", type="primary"):
            df = get_faculty_classload(
                st.session_state.username,
                selected_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            )
            
            if not df.empty:
                # Calculate summary metrics
                total_load = df['Distributed Load'].sum()
                unique_subjects = df['Subject'].nunique()
                unique_sections = df['Section'].nunique()
                
                # Display summary metrics
                st.write("### Workload Summary")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Net Classes", f"{total_load:.2f}")
                with col2:
                    st.metric("Unique Subjects", unique_subjects)
                with col3:
                    st.metric("Unique Sections", unique_sections)
                
                # Subject-wise distribution
                st.write("### Subject-wise Distribution")
                subject_dist = df.groupby('Subject')['Distributed Load'].sum().reset_index()
                subject_dist = subject_dist.sort_values('Distributed Load', ascending=False)
                st.bar_chart(subject_dist.set_index('Subject'))
                
                # Detailed workload table
                st.write("### Detailed Workload Report")
                st.dataframe(
                    df.sort_values(['Date', 'Period']),
                    column_config={
                        'Date': st.column_config.TextColumn('Date', width=100),
                        'Period': st.column_config.TextColumn('Period', width=80),
                        'Subject': st.column_config.TextColumn('Subject', width=150),
                        'Section': st.column_config.TextColumn('Section', width=120),
                        'Combined Sections': st.column_config.TextColumn('Combined Sections', width=200),
                        'Distributed Load': st.column_config.NumberColumn('Load', format="%.2f", width=80),
                        'Submission Time': st.column_config.TextColumn('Time', width=100),
                        'Lesson Plan': st.column_config.TextColumn('Lesson Plan', width=300)
                    },
                    hide_index=True
                )
                
                # Export options
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Detailed Workload', index=False)
                    
                    # Write summary sheet
                    summary_df = pd.DataFrame([{
                        'Metric': 'Net Classes',
                        'Value': total_load
                    }, {
                        'Metric': 'Unique Subjects',
                        'Value': unique_subjects
                    }, {
                        'Metric': 'Unique Sections',
                        'Value': unique_sections
                    }])
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Write subject distribution
                    subject_dist.to_excel(writer, sheet_name='Subject Distribution', index=False)
                
                st.download_button(
                    label="📥 Download Complete Report",
                    data=buffer.getvalue(),
                    file_name=f"faculty_workload_{st.session_state.username}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("No workload data found for the selected date range")
    
    # Tab 2: Daily Worksheet
    with tab2:
        selected_date, end_date = date_range_selector("faculty_worksheet1")
        worksheet_data = get_faculty_worksheet_data(
            st.session_state.username,
            selected_date,
            end_date
        )
        
        if worksheet_data:
            # Display summary metrics
            st.write("### Daily Summary")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Net Class Conducted", worksheet_data['total_classes'])
            with col2:
                st.metric("Other Activities", worksheet_data['total_hours'])
            with col3:
                total_load = worksheet_data['total_classes'] + worksheet_data['total_hours']
                st.metric("Total Load", f"{total_load}/6")
            
            # Display period-wise activities
            st.write("### Period-wise Activities")
            for period, data in worksheet_data['periods'].items():
                with st.expander(f"Period {period[1]} - {data['time']}", expanded=False):
                    if data['activity_type'] == 'Class':
                        # Display class details
                        st.info(f"""
                            **Subject:** {data['subject']}  
                            **Section(s):** {data['section']}  
                            **Topic:** {data['topic']}  
                            **Load:** {data.get('load', 1.0):.2f}
                        """)
                    else:
                        # Only show activity input form if it's today's date
                        if selected_date == datetime.now().date():
                            with st.form(key=f"activity_form_{period}"):
                                activity_type = st.selectbox(
                                    "Activity Type",
                                    [
                                        "Select Activity",
                                        "Diploma Classwork",
                                        "MCA Claswork",
                                        "BCA Classwork",
                                        "Absentees Phone Follow-up",
                                        "Administrative Work",
                                        "AICTE",
                                        "APSCHE",
                                        "B.Tech / MCA / M.Tech Project / Internship",
                                        "Department Work",
                                        "Department/HoD's Meeting",
                                        "DST",
                                        "Exam Duty",
                                        "Floor Duty",
                                        "IQAC",
                                        "JNTUK",
                                        "MSME",
                                        "NAAC",
                                        "NBA",
                                        "NIRF",
                                        "Parent's Meeting",
                                        "Placement",
                                        "PMKVY",
                                        "SBTET",
                                        "Student Counselling / Mentoring",
                                        "UGC",
                                        "Other"
                                    ],
                                    key=f"activity_type_{period}"
                                )
                                
                                description = st.text_area(
                                    "Description",
                                    value=data['topic'] if data['topic'] != '-' else '',
                                    key=f"description_{period}",
                                    height=100
                                )
                                
                                submitted = st.form_submit_button("Save Activity")
                                if submitted:
                                    if activity_type != "Select Activity" and description:
                                        if update_faculty_worksheet(
                                            st.session_state.username,
                                            selected_date,
                                            period,
                                            activity_type,
                                            description
                                        ):
                                            st.success("Activity updated successfully!")
                                            st.rerun()
                                    else:
                                        st.error("Please fill in all fields")
                        else:
                            if data['activity_type'] != 'None':
                                st.info(f"""
                                    **Activity:** {data['subject']}  
                                    **Description:** {data['topic']}  
                                    **Load:** {data.get('load', 1.0):.2f}
                                """)
            
            # Download worksheet
            if st.button("Download Daily Worksheet"):
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    worksheet_df = pd.DataFrame([
                        {
                            'Period': period,
                            'Time': data['time'],
                            'Activity': data['subject'],
                            'Section': data['section'],
                            'Topic/Description': data['topic'],
                            'Type': data['activity_type'],
                            'Load': data.get('load', 1.0)
                        }
                        for period, data in worksheet_data['periods'].items()
                    ])
                    
                    worksheet_df.to_excel(writer, sheet_name='Daily Worksheet', index=False)
                    
                st.download_button(
                    label="📥 Download Worksheet",
                    data=buffer.getvalue(),
                    file_name=f"daily_worksheet_{selected_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("No worksheet data found for the selected date")





def init_db():
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()

    # Create Students table
    c.execute('''CREATE TABLE IF NOT EXISTS students
                 (ht_number TEXT PRIMARY KEY, name TEXT, original_section TEXT, merged_section TEXT)''')
    
    # Create Faculty table
    c.execute('''CREATE TABLE IF NOT EXISTS faculty
                 (id INTEGER PRIMARY KEY, name TEXT, username TEXT UNIQUE, password TEXT)''')
    
    # Create Attendance table with created_at column
    # c.execute('DROP TABLE IF EXISTS faculty_worksheet')
    c.execute('''CREATE TABLE IF NOT EXISTS attendance
                 (id INTEGER PRIMARY KEY, 
                  ht_number TEXT, 
                  date TEXT, 
                  period TEXT, 
                  status TEXT, 
                  faculty TEXT, 
                  subject TEXT, 
                  lesson_plan TEXT,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    # Create Section-Subject-Mapping table
    c.execute('''CREATE TABLE IF NOT EXISTS section_subject_mapping
                 (id INTEGER PRIMARY KEY, section TEXT, subject_name TEXT)''')
    
    # Drop and recreate faculty_worksheet table with additional columns

    c.execute('''CREATE TABLE IF NOT EXISTS faculty_worksheet
                 (id INTEGER PRIMARY KEY,
                  faculty TEXT,
                  date TEXT,
                  period TEXT,
                  activity_type TEXT,
                  description TEXT,
                  subject TEXT,
                  section TEXT,
                  student_count INTEGER DEFAULT 0,
                  load REAL DEFAULT 1.0,
                  source TEXT DEFAULT 'Worksheet',
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    # Add default admin user if not exists
    c.execute("INSERT OR IGNORE INTO faculty (username, password, name) VALUES (?, ?, ?)", 
             ('admin', 'admin', 'Administrator'))
             
    conn.commit()
    conn.close()





def get_faculty_worksheet_data(username, start_date, end_date=None, view_type="Daily"):
    """Get faculty worksheet data with improved load calculation"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Get faculty name
        cursor.execute("SELECT name FROM faculty WHERE username = ?", (username,))
        faculty_name = cursor.fetchone()[0]
        
        # Initialize worksheet data structure
        worksheet_data = {
            'total_classes': 0,
            'total_hours': 0,
            'unique_sections': set(),
            'periods': {
                f'P{i}': {
                    'time': f'Period {i}',
                    'subject': 'No Activity',
                    'section': '-',
                    'topic': '-',
                    'status': 'Pending',
                    'activity_type': 'None',
                    'load': 0
                } for i in range(1, 7)
            }
        }
        
        # Convert dates to proper format
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date and isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        if not end_date:
            end_date = start_date
            
        # Get class data with proper load calculation
        query = """
            WITH PeriodGroups AS (
                SELECT 
                    date,
                    period,
                    COUNT(DISTINCT subject) as subject_count
                FROM (
                    SELECT DISTINCT
                        a.date,
                        a.period,
                        a.subject
                    FROM attendance a
                    WHERE a.faculty = ?
                    AND a.date BETWEEN ? AND ?
                )
                GROUP BY date, period
            ),
            DetailedClasses AS (
                SELECT 
                    a.date,
                    a.period,
                    a.subject,
                    GROUP_CONCAT(DISTINCT s.merged_section) as sections,
                    COUNT(DISTINCT s.merged_section) as section_count,
                    a.lesson_plan,
                    pg.subject_count,
                    ROUND(1.0/pg.subject_count, 2) as distributed_load
                FROM attendance a
                JOIN students s ON a.ht_number = s.ht_number
                JOIN PeriodGroups pg ON a.date = pg.date AND a.period = pg.period
                WHERE a.faculty = ?
                AND a.date BETWEEN ? AND ?
                GROUP BY a.date, a.period, a.subject
            )
            SELECT 
                period,
                subject,
                sections,
                lesson_plan,
                distributed_load
            FROM DetailedClasses
        """
        
        cursor.execute(query, (
            faculty_name, 
            start_date.strftime('%Y-%m-%d'), 
            end_date.strftime('%Y-%m-%d'),
            faculty_name,
            start_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d')
        ))
        
        # Process class data
        class_load = 0
        for row in cursor.fetchall():
            period, subject, sections, lesson_plan, dist_load = row
            worksheet_data['periods'][period] = {
                'time': f'Period {period[1]}',
                'subject': subject,
                'section': sections,
                'topic': lesson_plan,
                'status': 'Completed',
                'activity_type': 'Class',
                'load': dist_load
            }
            class_load += dist_load
            worksheet_data['unique_sections'].update(sections.split(','))
            
        worksheet_data['total_classes'] = class_load
            
        # Get other activities from faculty_worksheet
        query = """
            SELECT period, activity_type, description
            FROM faculty_worksheet
            WHERE faculty = ? 
            AND date BETWEEN ? AND ?
        """
        
        cursor.execute(query, (
            faculty_name, 
            start_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d')
        ))
        
        # Process activity data
        activity_load = 0
        for row in cursor.fetchall():
            period, activity_type, description = row
            if period in worksheet_data['periods']:
                # Only count if no class is already recorded for this period
                if worksheet_data['periods'][period]['activity_type'] == 'None':
                    worksheet_data['periods'][period] = {
                        'time': f'Period {period[1]}',
                        'subject': activity_type,
                        'section': '-',
                        'topic': description,
                        'status': 'Completed',
                        'activity_type': 'Other',
                        'load': 1.0
                    }
                    activity_load += 1
        
        worksheet_data['total_hours'] = activity_load
        
        return worksheet_data
        
    except Exception as e:
        st.error(f"Error getting worksheet data: {str(e)}")
        return None
    finally:
        if 'conn' in locals():
            conn.close()




def generate_template_excel(table_type):
    """Generate template Excel file for data import"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if table_type == 'students':
            # Create template for students
            df = pd.DataFrame(columns=['ht_number', 'name', 'original_section', 'merged_section'])
            df.loc[0] = ['21B01A0501', 'John Doe', 'CSE-A', 'CSE-1']  # Example row
            df.to_excel(writer, sheet_name='Students', index=False)
            
        elif table_type == 'faculty':
            # Create template for faculty
            df = pd.DataFrame(columns=['name', 'username', 'password'])
            df.loc[0] = ['John Smith', 'jsmith', 'password123']  # Example row
            df.to_excel(writer, sheet_name='Faculty', index=False)
            
        elif table_type == 'section_subjects':
            # Create template for section-subject mapping
            df = pd.DataFrame(columns=['section', 'subject_name'])
            df.loc[0] = ['CSE-A', 'Data Structures']  # Example row
            df.to_excel(writer, sheet_name='Section-Subject Mapping', index=False)
        
        # Format the template
        worksheet = writer.sheets[list(writer.sheets.keys())[0]]
        for column in worksheet.columns:
            length = max(len(str(cell.value) if cell.value else '') for cell in column)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = length + 2
            
        # Add header formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    return output.getvalue()

def import_excel_data(uploaded_file, table_type):
    """Import data from Excel file and update database"""
    try:
        df = pd.read_excel(uploaded_file)
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        if table_type == 'students':
            # Validate required columns
            required_columns = ['ht_number', 'name', 'original_section', 'merged_section']
            if not all(col in df.columns for col in required_columns):
                return False, "Missing required columns in the Excel file"
            
            # Update students table
            for _, row in df.iterrows():
                cursor.execute("""
                    INSERT OR REPLACE INTO students (ht_number, name, original_section, merged_section)
                    VALUES (?, ?, ?, ?)
                """, (row['ht_number'], row['name'], row['original_section'], row['merged_section']))
                
        elif table_type == 'faculty':
            # Validate required columns
            required_columns = ['name', 'username', 'password']
            if not all(col in df.columns for col in required_columns):
                return False, "Missing required columns in the Excel file"
            
            # Update faculty table
            for _, row in df.iterrows():
                cursor.execute("""
                    INSERT OR REPLACE INTO faculty (name, username, password)
                    VALUES (?, ?, ?)
                """, (row['name'], row['username'], row['password']))
                
        elif table_type == 'section_subjects':
            # Validate required columns
            required_columns = ['section', 'subject_name']
            if not all(col in df.columns for col in required_columns):
                return False, "Missing required columns in the Excel file"
            
            # Update section_subject_mapping table
            for _, row in df.iterrows():
                cursor.execute("""
                    INSERT OR REPLACE INTO section_subject_mapping (section, subject_name)
                    VALUES (?, ?)
                """, (row['section'], row['subject_name']))
        
        conn.commit()
        return True, "Data imported successfully"
        
    except Exception as e:
        return False, f"Error importing data: {str(e)}"
    finally:
        if 'conn' in locals():
            conn.close()

def get_all_tables():
    """Get list of all tables in the database"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Query to get all table names
        cursor.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name NOT LIKE 'sqlite_%'
        """)
        
        return [table[0] for table in cursor.fetchall()]
    finally:
        if 'conn' in locals():
            conn.close()

def get_table_data(table_name):
    """Get all data from specified table"""
    try:
        conn = sqlite3.connect('attendance.db')
        return pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
    finally:
        if 'conn' in locals():
            conn.close()

def mark_attendance(ht_number, date, period, status, faculty, subject, lesson_plan):
    """Modified to ensure correct IST date handling"""
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    
    try:
        # Get current time in IST
        current_time_ist = get_current_time_ist()
        current_time_str = format_ist_time(current_time_ist)
        
        # Debug logging
        print(f"Debug: Marking attendance at IST time: {current_time_str}")
        print(f"Debug: For date: {date}")
        
        # First insert into attendance table
        c.execute("""
            INSERT INTO attendance 
            (ht_number, date, period, status, faculty, subject, lesson_plan, created_at) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (ht_number, date, period, status, faculty, subject, lesson_plan, current_time_str))
        
        # Get the section for this student
        c.execute("SELECT merged_section FROM students WHERE ht_number = ?", (ht_number,))
        section = c.fetchone()[0]
        
        # Count total sections for load calculation
        c.execute("""
            WITH UniqueSections AS (
                SELECT DISTINCT s.merged_section
                FROM attendance a
                JOIN students s ON a.ht_number = s.ht_number
                WHERE a.faculty = ? 
                AND a.date = ? 
                AND a.period = ?
            )
            SELECT COUNT(*) FROM UniqueSections
        """, (faculty, date, period))
        
        total_sections = c.fetchone()[0]
        distributed_load = 1.0 / total_sections if total_sections > 0 else 1.0
        
        # Get student count
        c.execute("""
            SELECT COUNT(DISTINCT a.ht_number)
            FROM attendance a
            JOIN students s ON a.ht_number = s.ht_number
            WHERE a.faculty = ? 
            AND a.date = ? 
            AND a.period = ?
            AND a.subject = ?
            AND s.merged_section = ?
        """, (faculty, date, period, subject, section))
        
        student_count = c.fetchone()[0]
        
        # Check existing entry
        c.execute("""
            SELECT id 
            FROM faculty_worksheet 
            WHERE faculty = ? 
            AND date = ? 
            AND period = ? 
            AND subject = ? 
            AND section = ?
            AND source = 'Class'
        """, (faculty, date, period, subject, section))
        
        existing = c.fetchone()
        
        if existing:
            c.execute("""
                UPDATE faculty_worksheet 
                SET student_count = ?,
                    description = ?,
                    load = ?,
                    created_at = ?
                WHERE id = ?
            """, (student_count, lesson_plan, distributed_load, current_time_str, existing[0]))
        else:
            c.execute("""
                INSERT INTO faculty_worksheet 
                (faculty, date, period, activity_type, description, subject, section, 
                 student_count, load, source, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Class', ?)
            """, (faculty, date, period, subject, lesson_plan, subject, section, 
                 student_count, distributed_load, current_time_str))
        
        # Update loads for all sections
        c.execute("""
            UPDATE faculty_worksheet
            SET load = ?
            WHERE faculty = ?
            AND date = ?
            AND period = ?
            AND source = 'Class'
        """, (distributed_load, faculty, date, period))
        
        conn.commit()
        print(f"Successfully updated faculty_worksheet with distributed load at {current_time_str}")
        
    except Exception as e:
        print(f"Error marking attendance: {str(e)}")
        conn.rollback()
        raise e
    finally:
        conn.close()


def daily_worksheet():
    """Enhanced daily worksheet view with proper date-based editing permissions"""
    st.title("Faculty Workload Dashboard")
    
    # Get current time in IST
    current_time_ist = get_current_time_ist()
    current_date_ist = current_time_ist.date()
    
    # Create tabs for different views
    tab1, tab2 = st.tabs(["📊 Overall Workload", "📝 Daily Worksheet"])
    
    with tab2:
        st.subheader("Daily Summary")
        
        # Date selection with IST timezone consideration
        col1, col2 = st.columns(2)
        with col1:
            selected_date = st.date_input(
                "Select Date",
                value=current_date_ist,
                min_value=current_date_ist - timedelta(days=30),
                max_value=current_date_ist + timedelta(days=30)
            )
        
        with col2:
            view_type = st.selectbox(
                "View Type",
                ["Single Day", "Week View"],
                index=0
            )
        
        # Check if selected date is today
        is_today = selected_date == current_date_ist
        if not is_today:
            st.warning("⚠️ You can only edit worksheets for the current day. Past worksheets are view-only.")
        
        # Get faculty worksheet data
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        try:
            # Calculate loads
            c.execute("""
                SELECT COALESCE(SUM(load), 0)
                FROM faculty_worksheet
                WHERE faculty = ?
                AND date = ?
                AND source = 'Class'
            """, (st.session_state.username, selected_date.strftime('%Y-%m-%d')))
            
            net_class_load = c.fetchone()[0]
            
            c.execute("""
                SELECT COALESCE(SUM(load), 0)
                FROM faculty_worksheet
                WHERE faculty = ?
                AND date = ?
                AND source != 'Class'
            """, (st.session_state.username, selected_date.strftime('%Y-%m-%d')))
            
            other_activities_load = c.fetchone()[0]
            
            # Display summary
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Net Class Conducted", f"{net_class_load:.2f}")
            with col2:
                st.metric("Other Activities", f"{other_activities_load:.2f}")
            with col3:
                st.metric("Total Load", f"{net_class_load + other_activities_load:.2f}/6")
            
            st.subheader("Period-wise Activities")
            
            # Get existing activities for the selected date
            c.execute("""
                SELECT period, activity_type, description, subject, section, 
                       student_count, load, source
                FROM faculty_worksheet
                WHERE faculty = ?
                AND date = ?
                ORDER BY period
            """, (st.session_state.username, selected_date.strftime('%Y-%m-%d')))
            
            existing_activities = c.fetchall()
            activities_by_period = {row[0]: row for row in existing_activities}
            
            # Show all periods (1-6)
            for period in range(1, 7):
                period_str = f"Period {period}"
                with st.expander(f"{period_str} - {period_str}"):
                    if period_str in activities_by_period:
                        activity = activities_by_period[period_str]
                        st.write(f"**Activity Type:** {activity[1]}")
                        st.write(f"**Description:** {activity[2]}")
                        st.write(f"**Subject:** {activity[3]}")
                        st.write(f"**Section:** {activity[4]}")
                        st.write(f"**Student Count:** {activity[5]}")
                        st.write(f"**Load:** {activity[6]}")
                        st.write(f"**Source:** {activity[7]}")
                        
                        # Only show edit/delete buttons for today's worksheet
                        if is_today:
                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button(f"Edit Activity", key=f"edit_{period}"):
                                    st.session_state.editing_activity = True
                                    st.session_state.editing_period = period_str
                                    st.session_state.activity_data = activity
                            with col2:
                                if st.button(f"Delete Activity", key=f"delete_{period}"):
                                    if st.warning("Are you sure you want to delete this activity?"):
                                        c.execute("""
                                            DELETE FROM faculty_worksheet
                                            WHERE faculty = ?
                                            AND date = ?
                                            AND period = ?
                                        """, (st.session_state.username, 
                                             selected_date.strftime('%Y-%m-%d'),
                                             period_str))
                                        conn.commit()
                                        st.success("Activity deleted successfully!")
                                        st.rerun()
                    else:
                        st.info("No activities recorded for this period")
                        
                        # Only show add button for today's worksheet
                        if is_today:
                            if st.button(f"Add Activity", key=f"add_{period}"):
                                st.session_state.adding_activity = True
                                st.session_state.selected_period = period_str
            
            # Download button
            if st.button("Download Daily Worksheet"):
                df = pd.read_sql_query("""
                    SELECT period, activity_type, description, subject, section, 
                           student_count, load, source, created_at
                    FROM faculty_worksheet
                    WHERE faculty = ?
                    AND date = ?
                    ORDER BY period
                """, conn, params=(st.session_state.username, 
                                 selected_date.strftime('%Y-%m-%d')))
                
                # Convert timestamps to IST
                df['created_at'] = pd.to_datetime(df['created_at']).apply(
                    lambda x: format_ist_time(x)
                )
                
                # Create Excel file
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Daily Worksheet')
                    worksheet = writer.sheets['Daily Worksheet']
                    
                    # Format headers
                    for idx, col in enumerate(df.columns, 1):
                        cell = worksheet.cell(1, idx)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color='D3D3D3',
                            end_color='D3D3D3',
                            fill_type='solid'
                        )
                        worksheet.column_dimensions[get_column_letter(idx)].width = max(len(col) + 5, 15)
                
                # Offer download
                st.download_button(
                    label="📥 Download Excel",
                    data=output.getvalue(),
                    file_name=f"daily_worksheet_{selected_date.strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
        except Exception as e:
            st.error(f"Error loading worksheet: {str(e)}")
            print(f"Debug - Error in daily_worksheet: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()

def view_timetable():
    """Complete timetable view with proper variable handling and IST timezone"""
    st.title("Class Timetable")
    
    # Get current date in IST
    current_date = get_current_date_ist()
    
    # Date selection
    col1, col2 = st.columns(2)
    with col1:
        selected_date = st.date_input(
            "Select Date",
            value=current_date,
            min_value=current_date - timedelta(days=30),
            max_value=current_date + timedelta(days=30)
        )
    
    with col2:
        view_type = st.selectbox(
            "View Type",
            ["Single Day", "Date Range"],
            index=0,
            key="timetable_view_type"  # Added unique key
        )
    
    # Initialize end_date based on view type
    end_date = None
    if view_type == "Date Range":
        end_date = st.date_input(
            "End Date",
            value=selected_date + timedelta(days=1),
            min_value=selected_date,
            max_value=selected_date + timedelta(days=30),
            key="end_date"  # Added unique key
        )
    else:
        end_date = selected_date
    
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Calculate summary metrics
        cursor.execute("""
            WITH DailyStats AS (
                SELECT 
                    COUNT(DISTINCT period) as total_classes,
                    COUNT(DISTINCT faculty) as faculty_count,
                    COUNT(DISTINCT section) as section_count
                FROM faculty_worksheet
                WHERE date BETWEEN ? AND ?
                AND faculty = ?
            )
            SELECT 
                COALESCE(total_classes, 0) as total_classes,
                COALESCE(faculty_count, 0) as faculty_count,
                COALESCE(section_count, 0) as section_count
            FROM DailyStats
        """, (selected_date.strftime('%Y-%m-%d'), 
              end_date.strftime('%Y-%m-%d'),
              st.session_state.username))
        
        stats = cursor.fetchone()
        
        # Display metrics in columns
        st.subheader("Daily Summary")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Classes", stats[0])
        with col2:
            st.metric("Faculty Engaged", stats[1])
        with col3:
            st.metric("Sections Covered", stats[2])
        
        # Fetch timetable data
        query = """
            SELECT 
                date,
                period,
                faculty,
                subject,
                section,
                student_count,
                strftime('%H:%M', created_at) as time,
                CASE 
                    WHEN student_count > 0 THEN 'Completed'
                    ELSE 'Pending'
                END as status
            FROM faculty_worksheet
            WHERE date BETWEEN ? AND ?
            AND faculty = ?
            ORDER BY date, period
        """
        
        df = pd.read_sql_query(
            query, 
            conn, 
            params=(selected_date.strftime('%Y-%m-%d'),
                   end_date.strftime('%Y-%m-%d'),
                   st.session_state.username)
        )
        
        if not df.empty:
            st.subheader("Class Timetable")
            
            # Format the dataframe
            df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
            
            # Add styling
            def highlight_status(val):
                if val == 'Completed':
                    return 'background-color: #90EE90'
                return 'background-color: #FFB6C1'
            
            # Display the styled dataframe
            st.dataframe(
                df.style.apply(lambda x: [''] * len(x) if x.name != 'status' 
                             else [highlight_status(val) for val in x], axis=0),
                use_container_width=True,
                hide_index=True
            )
            
            # Export functionality
            if st.button("Export Timetable"):
                # Create Excel file
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Timetable')
                    worksheet = writer.sheets['Timetable']
                    
                    # Format headers
                    for idx, col in enumerate(df.columns, 1):
                        cell = worksheet.cell(1, idx)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color='D3D3D3',
                            end_color='D3D3D3',
                            fill_type='solid'
                        )
                        worksheet.column_dimensions[get_column_letter(idx)].width = max(len(col) + 5, 15)
                
                # Offer download
                st.download_button(
                    label="📥 Download Excel",
                    data=output.getvalue(),
                    file_name=f"timetable_{selected_date.strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("No classes scheduled for the selected date range")
        
    except Exception as e:
        st.error(f"Error loading class timetable: {str(e)}")
        print(f"Debug - Error in view_timetable: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()
    
    # Add refresh button
    if st.button("🔄 Refresh Timetable"):
        st.rerun()





def view_faculty_worksheet_stats():
    """Modified to use faculty_worksheet table as single source of truth"""
    st.subheader("Faculty Worksheet Statistics")
    
    # Use the date range selector
    selected_date, end_date = date_range_selector("faculty_stats")
    
    if st.button("Generate Report", type="primary"):
        try:
            conn = sqlite3.connect('attendance.db')
            cursor = conn.cursor()
            
            # Updated query to use faculty_worksheet table
            query = """
            WITH FacultyStats AS (
                SELECT 
                    f.name as faculty_name,
                    -- Classes from attendance records
                    COALESCE(SUM(CASE WHEN fw.source = 'Class' THEN fw.load ELSE 0 END), 0) as net_classes_conducted,
                    -- Other activities
                    COALESCE(SUM(CASE WHEN fw.source = 'Worksheet' THEN fw.load ELSE 0 END), 0) as other_activities,
                    -- Unique sections
                    COUNT(DISTINCT fw.section) as unique_sections,
                    -- Unique subjects
                    COUNT(DISTINCT fw.subject) as unique_subjects,
                    -- Unique activity types
                    COUNT(DISTINCT CASE WHEN fw.source = 'Worksheet' THEN fw.activity_type END) as unique_activity_types,
                    -- Total students
                    SUM(COALESCE(fw.student_count, 0)) as total_students,
                    -- Class dates
                    COUNT(DISTINCT CASE WHEN fw.source = 'Class' THEN fw.date END) as class_dates,
                    -- Activity dates
                    COUNT(DISTINCT CASE WHEN fw.source = 'Worksheet' THEN fw.date END) as activity_dates
                FROM faculty f
                LEFT JOIN faculty_worksheet fw ON f.name = fw.faculty 
                    AND fw.date BETWEEN ? AND ?
                GROUP BY f.name
            )
            SELECT 
                faculty_name,
                net_classes_conducted,
                other_activities,
                unique_sections,
                unique_subjects,
                unique_activity_types,
                ROUND(net_classes_conducted + other_activities, 2) as total_load,
                total_students,
                class_dates,
                activity_dates
            FROM FacultyStats
            ORDER BY total_load DESC;
            """
            
            # Format dates properly for SQLite
            start_date_str = selected_date.strftime('%Y-%m-%d')
            end_date_str = end_date.strftime('%Y-%m-%d')
            
            cursor.execute(query, (start_date_str, end_date_str))
            results = cursor.fetchall()
            
            if results:
                # Create DataFrame with proper column names
                df = pd.DataFrame(results, columns=[
                    'Faculty Name',
                    'Net Classes Conducted',
                    'Other Activities',
                    'Unique Sections',
                    'Unique Subjects',
                    'Unique Activity Types',
                    'Total Load',
                    'Total Students',
                    'Class Dates',
                    'Activity Dates'
                ])
                
                # Display summary metrics
                st.write("### Overall Summary")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Faculty", len(df))
                with col2:
                    st.metric("Net Classes", f"{df['Net Classes Conducted'].sum():.2f}")
                with col3:
                    st.metric("Total Activities", df['Other Activities'].sum())
                with col4:
                    st.metric("Avg Load", f"{df['Total Load'].mean():.2f}")
                
                # Display the statistics table
                st.write("### Faculty-wise Statistics")
                st.dataframe(df)
                
                # Get detailed activity log
                detailed_query = """
                SELECT 
                    faculty,
                    date,
                    period,
                    COALESCE(subject, activity_type) as activity,
                    section,
                    description,
                    student_count,
                    load,
                    source,
                    created_at
                FROM faculty_worksheet
                WHERE date BETWEEN ? AND ?
                ORDER BY date DESC, period ASC, faculty
                """
                
                detailed_df = pd.read_sql_query(detailed_query, conn, params=(start_date_str, end_date_str))
                
                if not detailed_df.empty:
                    st.write("### Detailed Activity Log")
                    st.dataframe(
                        detailed_df,
                        column_config={
                            'faculty': st.column_config.TextColumn('Faculty', width=150),
                            'date': st.column_config.TextColumn('Date', width=100),
                            'period': st.column_config.TextColumn('Period', width=80),
                            'activity': st.column_config.TextColumn('Activity/Subject', width=150),
                            'section': st.column_config.TextColumn('Section', width=100),
                            'description': st.column_config.TextColumn('Description/Plan', width=300),
                            'student_count': st.column_config.NumberColumn('Students', width=80),
                            'load': st.column_config.NumberColumn('Load', width=80, format="%.2f"),
                            'source': st.column_config.TextColumn('Type', width=100),
                            'created_at': st.column_config.TextColumn('Timestamp', width=150)
                        }
                    )
                
                # Export functionality
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Faculty Statistics', index=False)
                    if not detailed_df.empty:
                        detailed_df.to_excel(writer, sheet_name='Detailed Log', index=False)
                
                st.download_button(
                    label="📥 Download Complete Report",
                    data=buffer.getvalue(),
                    file_name=f"faculty_worksheet_stats_{selected_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            else:
                st.info("No data found for the selected date range")
                
        except Exception as e:
            st.error(f"Error generating faculty statistics: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()






def update_faculty_worksheet(username, date, period, activity_type, description):
    """Update faculty worksheet with enhanced error handling and proper data saving"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Get faculty name
        cursor.execute("SELECT name FROM faculty WHERE username = ?", (username,))
        faculty_name = cursor.fetchone()
        
        if not faculty_name:
            raise Exception("Faculty not found")
            
        faculty_name = faculty_name[0]
        
        # Format date properly
        if isinstance(date, datetime):
            date = date.strftime('%Y-%m-%d')
            
        # Check if worksheet entry exists
        cursor.execute("""
            SELECT id FROM faculty_worksheet 
            WHERE faculty = ? AND date = ? AND period = ?
        """, (faculty_name, date, period))
        
        existing = cursor.fetchone()
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if existing:
            # Update existing worksheet entry
            cursor.execute("""
                UPDATE faculty_worksheet 
                SET activity_type = ?, 
                    description = ?,
                    load = 1.0,
                    created_at = ?
                WHERE id = ?
            """, (activity_type, description, current_time, existing[0]))
        else:
            # Insert new worksheet entry
            cursor.execute("""
                INSERT INTO faculty_worksheet 
                (faculty, date, period, activity_type, description, load, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (faculty_name, date, period, activity_type, description, 1.0, current_time))
            
        conn.commit()
        return True
        
    except Exception as e:
        st.error(f"Error updating worksheet: {str(e)}")
        return False
    finally:
        if 'conn' in locals():
            conn.close()




def generate_table_template(table_name):
    """Generate template Excel file for any table"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Get column information
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = [col[1] for col in cursor.fetchall()]
        
        # Create template DataFrame
        df = pd.DataFrame(columns=columns)
        
        # Add example row based on table structure
        example_row = {}
        for col in columns:
            if col == 'id':  # Skip ID column in template
                continue
            elif 'date' in col.lower():
                example_row[col] = datetime.now().strftime('%Y-%m-%d')
            elif 'created_at' in col.lower():
                continue  # Skip created_at as it's auto-generated
            elif col == 'status':
                example_row[col] = 'P'
            elif col == 'load':
                example_row[col] = 1.0
            elif col == 'student_count':
                example_row[col] = 0
            elif col == 'source':
                example_row[col] = 'Class'
            else:
                example_row[col] = f'Example {col}'
        
        # Add example row if we have data
        if example_row:
            df.loc[0] = example_row
        
        # Create Excel writer object
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=table_name, index=False)
            
            # Format the template
            worksheet = writer.sheets[table_name]
            for idx, col in enumerate(df.columns, 1):
                worksheet.column_dimensions[get_column_letter(idx)].width = max(len(col) + 5, 15)
            
            # Add header formatting
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        return output.getvalue()
        
    except Exception as e:
        print(f"Error generating template: {str(e)}")
        return None
    finally:
        if 'conn' in locals():
            conn.close()

def bulk_upload_data(table_name, df):
    """Handle bulk data upload for any table"""
    try:
        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        
        # Get column information
        cursor.execute(f"PRAGMA table_info({table_name})")
        table_columns = [col[1] for col in cursor.fetchall()]
        
        # Remove id and created_at columns if they exist
        upload_columns = [col for col in table_columns if col not in ['id', 'created_at']]
        
        # Verify DataFrame columns match table columns
        missing_cols = set(upload_columns) - set(df.columns)
        if missing_cols:
            return False, f"Missing required columns: {', '.join(missing_cols)}"
        
        # Prepare data for insertion
        df = df[upload_columns]  # Only keep relevant columns
        
        # Begin transaction
        conn.execute("BEGIN TRANSACTION")
        
        # Prepare INSERT statement
        placeholders = ','.join(['?' for _ in upload_columns])
        columns = ','.join(upload_columns)
        insert_sql = f"INSERT OR REPLACE INTO {table_name} ({columns}) VALUES ({placeholders})"
        
        # Insert data
        cursor.executemany(insert_sql, df.values.tolist())
        
        # Commit transaction
        conn.commit()
        return True, f"Successfully uploaded {len(df)} records to {table_name}"
        
    except Exception as e:
        conn.rollback()
        return False, f"Error uploading data: {str(e)}"
    finally:
        if 'conn' in locals():
            conn.close()

def view_data_tab():
    """Enhanced view data tab with multi-sheet workbook bulk upload capability"""
    st.subheader("View Database Tables")
    
    tables = get_all_tables()
    selected_table = st.selectbox(
        "Select Table to View",
        tables,
        format_func=lambda x: x.replace('_', ' ').title()
    )
    
    if selected_table:
        # Create tabs for different operations
        tab1, tab2 = st.tabs(["View & Edit Data", "Bulk Upload"])
        
        with tab1:
            # View & Edit functionality remains unchanged
            df = get_table_data(selected_table)
            
            if not df.empty:
                if f"original_{selected_table}" not in st.session_state:
                    st.session_state[f"original_{selected_table}"] = df.copy()
                
                edited_df = st.data_editor(
                    df,
                    num_rows="dynamic",
                    use_container_width=True,
                    hide_index=True,
                    key=f"editor_{selected_table}"
                )
                
                if st.button("Save Changes", type="primary"):
                    try:
                        conn = sqlite3.connect('attendance.db')
                        cursor = conn.cursor()
                        
                        primary_key_col = df.columns[0]
                        original_df = st.session_state[f"original_{selected_table}"]
                        modified_mask = (edited_df != original_df).any(axis=1)
                        modified_rows = edited_df[modified_mask]
                        
                        if not modified_rows.empty:
                            conn.execute("BEGIN TRANSACTION")
                            
                            for idx, row in modified_rows.iterrows():
                                columns = edited_df.columns
                                set_clause = ", ".join([f"{col} = ?" for col in columns])
                                query = f"UPDATE {selected_table} SET {set_clause} WHERE {primary_key_col} = ?"
                                
                                values = [row[col] for col in columns]
                                values.append(row[primary_key_col])
                                
                                cursor.execute(query, values)
                            
                            conn.commit()
                            st.success(f"Successfully updated {len(modified_rows)} rows!")
                            st.session_state[f"original_{selected_table}"] = edited_df.copy()
                            st.rerun()
                        else:
                            st.info("No changes detected")
                            
                    except Exception as e:
                        if 'conn' in locals():
                            conn.rollback()
                        st.error(f"Error saving changes: {str(e)}")
                    finally:
                        if 'conn' in locals():
                            conn.close()
        
        with tab2:
            st.write("### Bulk Upload Data")
            st.info("You can either upload data for a single table or upload a workbook containing multiple tables as separate sheets.")
            
            # Option to download complete template
            if st.button("📥 Download Complete Template"):
                try:
                    conn = sqlite3.connect('attendance.db')
                    cursor = conn.cursor()
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for table in tables:
                            # Get column information for each table
                            cursor.execute(f"PRAGMA table_info({table})")
                            columns = [col[1] for col in cursor.fetchall() 
                                     if col[1] not in ['id', 'created_at']]
                            
                            # Create example data
                            example_data = {}
                            for col in columns:
                                if 'date' in col.lower():
                                    example_data[col] = datetime.now().strftime('%Y-%m-%d')
                                elif col == 'status':
                                    example_data[col] = 'P'
                                elif col == 'load':
                                    example_data[col] = 1.0
                                elif col == 'student_count':
                                    example_data[col] = 0
                                elif col == 'source':
                                    example_data[col] = 'Class'
                                else:
                                    example_data[col] = f'Example {col}'
                            
                            template_df = pd.DataFrame([example_data])
                            
                            # Write to sheet
                            template_df.to_excel(writer, sheet_name=table, index=False)
                            worksheet = writer.sheets[table]
                            
                            # Format header
                            for idx, col in enumerate(template_df.columns, 1):
                                cell = worksheet.cell(1, idx)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(
                                    start_color='D3D3D3',
                                    end_color='D3D3D3',
                                    fill_type='solid'
                                )
                                worksheet.column_dimensions[get_column_letter(idx)].width = max(len(col) + 5, 15)
                    
                    st.download_button(
                        label="Download Complete Template Excel",
                        data=output.getvalue(),
                        file_name="database_template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except Exception as e:
                    st.error(f"Error generating template: {str(e)}")
                finally:
                    if 'conn' in locals():
                        conn.close()
            
            # Single table template download
            if st.button("📥 Download Template for Selected Table"):
                try:
                    conn = sqlite3.connect('attendance.db')
                    cursor = conn.cursor()
                    
                    cursor.execute(f"PRAGMA table_info({selected_table})")
                    columns = [col[1] for col in cursor.fetchall() 
                             if col[1] not in ['id', 'created_at']]
                    
                    example_data = {}
                    for col in columns:
                        if 'date' in col.lower():
                            example_data[col] = datetime.now().strftime('%Y-%m-%d')
                        elif col == 'status':
                            example_data[col] = 'P'
                        elif col == 'load':
                            example_data[col] = 1.0
                        elif col == 'student_count':
                            example_data[col] = 0
                        elif col == 'source':
                            example_data[col] = 'Class'
                        else:
                            example_data[col] = f'Example {col}'
                    
                    template_df = pd.DataFrame([example_data])
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        template_df.to_excel(writer, index=False, sheet_name='Template')
                        worksheet = writer.sheets['Template']
                        
                        for idx, col in enumerate(template_df.columns, 1):
                            cell = worksheet.cell(1, idx)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color='D3D3D3',
                                end_color='D3D3D3',
                                fill_type='solid'
                            )
                            worksheet.column_dimensions[get_column_letter(idx)].width = max(len(col) + 5, 15)
                    
                    st.download_button(
                        label="Download Template Excel",
                        data=output.getvalue(),
                        file_name=f"{selected_table}_template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except Exception as e:
                    st.error(f"Error generating template: {str(e)}")
                finally:
                    if 'conn' in locals():
                        conn.close()
            
            uploaded_file = st.file_uploader(
                "Upload Excel File",
                type=['xlsx'],
                help="Upload either a single table template or a complete workbook with multiple sheets"
            )
            
            if uploaded_file is not None:
                try:
                    # Read all sheets in the workbook
                    excel_file = pd.ExcelFile(uploaded_file)
                    sheet_names = excel_file.sheet_names
                    
                    st.write("### Preview of uploaded data:")
                    
                    # Display preview for each sheet
                    for sheet in sheet_names:
                        if sheet in tables:  # Only process sheets that match table names
                            df = pd.read_excel(excel_file, sheet_name=sheet)
                            st.write(f"**Table: {sheet}**")
                            st.dataframe(df.head())
                    
                    if st.button("Confirm Upload", type="primary"):
                        if st.warning("⚠️ This will DELETE ALL existing data in the affected tables and replace it with the uploaded data. Are you sure?"):
                            conn = sqlite3.connect('attendance.db')
                            cursor = conn.cursor()
                            
                            try:
                                conn.execute("BEGIN TRANSACTION")
                                
                                for sheet in sheet_names:
                                    if sheet in tables:
                                        df = pd.read_excel(excel_file, sheet_name=sheet)
                                        
                                        # Get column information
                                        cursor.execute(f"PRAGMA table_info({sheet})")
                                        table_info = cursor.fetchall()
                                        table_columns = [col[1] for col in table_info 
                                                       if col[1] not in ['id', 'created_at']]
                                        
                                        # Verify columns match
                                        missing_cols = set(table_columns) - set(df.columns)
                                        if missing_cols:
                                            raise ValueError(f"Missing required columns in {sheet}: {', '.join(missing_cols)}")
                                        
                                        # Process data types
                                        for col in df.columns:
                                            if pd.api.types.is_datetime64_any_dtype(df[col]):
                                                df[col] = df[col].dt.strftime('%Y-%m-%d')
                                            elif 'timestamp' in str(df[col].dtype).lower():
                                                df[col] = pd.to_datetime(df[col]).dt.strftime('%Y-%m-%d %H:%M:%S')
                                        
                                        # Clear existing data
                                        cursor.execute(f"DELETE FROM {sheet}")
                                        
                                        # Insert new data
                                        columns = ','.join(table_columns)
                                        placeholders = ','.join(['?' for _ in table_columns])
                                        insert_sql = f"""
                                            INSERT INTO {sheet} 
                                            ({columns}) VALUES ({placeholders})
                                        """
                                        
                                        data_to_insert = df[table_columns].values.tolist()
                                        cursor.executemany(insert_sql, data_to_insert)
                                        
                                        st.success(f"Successfully replaced {len(df)} records in {sheet}")
                                
                                conn.commit()
                                st.rerun()
                                
                            except Exception as e:
                                conn.rollback()
                                st.error(f"Error uploading data: {str(e)}")
                            finally:
                                conn.close()
                                
                except Exception as e:
                    st.error(f"Error reading file: {str(e)}")
    else:
        st.info("Please select a table to view")




def admin_page():
    st.title("Admin Dashboard")
    
    tab1, tab2, tab3= st.tabs([

        "Faculty Worksheet Stats",
        "View/Edit Data",
        "Downlaod Data"
    ])


    with tab1:
        view_faculty_worksheet_stats()
    
    with tab2:
        view_data_tab()
    
    
    
    with tab3:
        download_data_tab()




def download_data_tab():
    """Enhanced download data tab with latest database file handling"""
    st.subheader("Download Database and Export Data")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Create a copy of the current database file
        if st.button("💾 Download Database File"):
            try:
                # Close any existing connections to ensure all changes are saved
                conn = sqlite3.connect('attendance.db')
                conn.close()
                
                # Create a copy of the database in memory
                current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
                temp_db_path = f'temp_db_{current_time}.db'
                
                # Copy the database file
                with open('attendance.db', 'rb') as src, open(temp_db_path, 'wb') as dst:
                    dst.write(src.read())
                
                # Read the temporary file and serve it for download
                with open(temp_db_path, 'rb') as f:
                    db_data = f.read()
                
                # Clean up the temporary file
                import os
                os.remove(temp_db_path)
                
                # Offer the file for download with timestamp in filename
                st.download_button(
                    label="📥 Click to Download Database",
                    data=db_data,
                    file_name=f"attendance_db_{current_time}.db",
                    mime="application/x-sqlite3"
                )
                
            except Exception as e:
                st.error(f"Error preparing database for download: {str(e)}")
    
    with col2:
        if st.button("📊 Export All Data to Excel (Including Sensitive Info)"):
            try:
                # Create Excel file in memory
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Get all tables
                    conn = sqlite3.connect('attendance.db')
                    cursor = conn.cursor()
                    
                    # Get list of all tables
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = cursor.fetchall()
                    
                    # Export each table to a separate sheet
                    for table in tables:
                        table_name = table[0]
                        df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
                        
                        # Write to Excel
                        df.to_excel(writer, sheet_name=table_name, index=False)
                        
                        # Format the sheet
                        worksheet = writer.sheets[table_name]
                        for idx, col in enumerate(df.columns, 1):
                            cell = worksheet.cell(1, idx)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color='D3D3D3',
                                end_color='D3D3D3',
                                fill_type='solid'
                            )
                            worksheet.column_dimensions[get_column_letter(idx)].width = max(len(str(col)) + 5, 15)
                    
                    conn.close()
                
                # Offer the Excel file for download
                current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
                st.download_button(
                    label="📥 Download Complete Excel Export",
                    data=output.getvalue(),
                    file_name=f"attendance_system_export_{current_time}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"Error exporting data: {str(e)}")


def get_current_time_ist():
    """Gets the current time in IST with proper timezone handling"""
    ist = pytz.timezone('Asia/Kolkata')
    utc_now = datetime.now(pytz.UTC)
    ist_now = utc_now.astimezone(ist)
    return ist_now

def get_current_date_ist():
    """Gets the current date in IST"""
    return get_current_time_ist().date()

def format_ist_time(dt):
    """Formats a datetime object to IST time with proper timezone conversion"""
    ist = pytz.timezone('Asia/Kolkata')
    if dt.tzinfo is None:
        # If datetime is naive, assume it's in UTC
        dt = pytz.UTC.localize(dt)
    ist_time = dt.astimezone(ist)
    return ist_time.strftime('%Y-%m-%d %H:%M:%S')

def main():
    # Initialize the database
    init_db()
    
    # Set page config
    st.set_page_config(
        page_title="RVIT - Attendance Management System",
        page_icon="https://rvit.edu.in/rvitlogo_f.jpg",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Add custom CSS
    st.markdown("""
        <style>
        .stApp {
            max-width: 100%;
            padding: 1rem;
        }
        .stButton button {
            width: 100%;
            padding: 0.8rem !important;
            border-radius: 10px !important;
            margin: 0.5rem 0 !important;
        }
        .student-row {
            background-color: #f0f2f6;
            padding: 1rem;
            border-radius: 10px;
            margin: 0.5rem 0;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Session state management
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
    
    # Main application logic
    if not st.session_state.logged_in:
        login()
    else:
        # Sidebar for navigation
        with st.sidebar:
            st.image("https://rvit.edu.in/rvitlogo_f.jpg", width=100)
            st.write(f"Welcome, {st.session_state.faculty_name}")
            st.divider()
            
            if st.session_state.get('is_admin', False):
                page = "Admin Dashboard"
            else:
                page = st.radio("Navigation", ["Mark Attendance", "View Statistics", 
                                               "Classes Timetable", "My Work Tracker", "Reset Credentials"])
            
            if st.button("Logout", type="primary"):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
        
        # Main content
        if page == "Admin Dashboard" and st.session_state.get('is_admin', False):
            admin_page()
        elif page == "Mark Attendance":
            mark_attendance_page()
        elif page == "View Statistics":
            view_statistics_page()
        elif page == "Classes Timetable":
            view_class_timetable()
        elif page == "My Work Tracker":
            show_faculty_classload()
        elif page == "Reset Credentials":
            reset_credentials_page()
        else:
            st.error("You don't have permission to access this page.")


if __name__ == "__main__":
    main()

