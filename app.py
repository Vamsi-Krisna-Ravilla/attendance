import streamlit as st
import pandas as pd
import shutil
import glob
import os
import hashlib
import openpyxl
from openpyxl.styles import Font
import io
import numpy as np
from datetime import datetime, timezone
import pytz
from openpyxl.styles import Border, Side, Alignment
import pdfkit
from datetime import datetime
import base64
from io import BytesIO
from jinja2 import Template


def generate_pdf_report(df, title, filters=None):
    """
    Generate a PDF report with automatic column adjustment and header image
    
    Args:
        df: DataFrame containing the report data
        title: Report title
        filters: Dictionary containing any filter criteria applied
    """
    # HTML template for the report
    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            @page {
                size: landscape;
                margin: 1cm;
            }
            body {
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
            }
            .header {
                text-align: center;
                margin-bottom: 20px;
            }
            .header img {
                max-width: 100%;
                height: auto;
            }
            .report-title {
                font-size: 20px;
                font-weight: bold;
                margin: 15px 0;
                color: #2c3e50;
                text-align: center;
            }
            .filters {
                margin-bottom: 15px;
                font-size: 12px;
                color: #666;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
                font-size: 10px;
            }
            th {
                background-color: #f8f9fa;
                color: #2c3e50;
                padding: 8px;
                border: 1px solid #ddd;
                text-align: left;
            }
            td {
                padding: 6px;
                border: 1px solid #ddd;
            }
            tr:nth-child(even) {
                background-color: #f9f9f9;
            }
            .timestamp {
                text-align: right;
                font-size: 10px;
                color: #666;
                margin-top: 10px;
            }
            .footer {
                position: fixed;
                bottom: 0;
                width: 100%;
                text-align: center;
                font-size: 10px;
                color: #666;
                padding: 10px 0;
            }
        </style>
    </head>
    <body>
        <div class="header">
            <img src="data:image/png;base64,{{ header_image }}" alt="Header">
        </div>
        
        <div class="report-title">{{ title }}</div>
        
        {% if filters %}
        <div class="filters">
            <strong>Filters:</strong>
            {% for key, value in filters.items() %}
                {{ key }}: {{ value }}{% if not loop.last %} | {% endif %}
            {% endfor %}
        </div>
        {% endif %}
        
        <table>
            <thead>
                <tr>
                    {% for column in columns %}
                    <th>{{ column }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for row in data %}
                <tr>
                    {% for value in row %}
                    <td>{{ value }}</td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
        
        <div class="timestamp">
            Generated on: {{ timestamp }}
        </div>
        
        <div class="footer">
            Page <span class="pageNumber"></span> of <span class="totalPages"></span>
        </div>
    </body>
    </html>
    """
    
    # Read and encode the header image
    with open('header.png', 'rb') as img_file:
        header_image = base64.b64encode(img_file.read()).decode('utf-8')
    
    # Remove unnecessary columns
    if 'Section' in df.columns:
        df = df.drop('Section', axis=1)
    if 'Student Name' in df.columns:
        df = df.drop('Student Name', axis=1)
    
    # Prepare template data
    template_data = {
        'header_image': header_image,
        'title': title,
        'filters': filters,
        'columns': df.columns.tolist(),
        'data': df.values.tolist(),
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # Generate HTML
    template = Template(html_template)
    html_content = template.render(**template_data)
    
    # PDF options for landscape A4
    pdf_options = {
        'page-size': 'A4',
        'orientation': 'Landscape',
        'margin-top': '10mm',
        'margin-right': '10mm',
        'margin-bottom': '10mm',
        'margin-left': '10mm',
        'encoding': 'UTF-8',
        'no-outline': None,
        'enable-local-file-access': None
    }
    
    # Generate PDF
    pdf = pdfkit.from_string(html_content, False, options=pdf_options)
    
    return pdf

# Example usage in the existing code:
def download_attendance_report(stats_df, filters):
    try:
        pdf = generate_pdf_report(
            df=stats_df,
            title="Attendance Report",
            filters=filters
        )
        
        return pdf
    except Exception as e:
        st.error(f"Error generating PDF: {str(e)}")
        return None

# Single consolidated page config at the very start
st.set_page_config(
    page_title="Attendance Management System",
    layout="wide",
    initial_sidebar_state="collapsed",  # Better for mobile
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': "Attendance Management System"
    }
)

# Rest of your custom CSS for mobile-friendly styling
st.markdown("""
    <style>
        /* Mobile-friendly containers */
        .stApp {
            max-width: 100%;
            padding: 1rem;
        }
        
        /* Improved button styling */
        .stButton button {
            width: 100%;
            padding: 0.8rem !important;
            border-radius: 10px !important;
            font-size: 1rem !important;
            font-weight: 500 !important;
            margin: 0.5rem 0 !important;
        }
        
        /* Card-like containers */
        .css-1r6slb0 {  /* Streamlit container class */
            background-color: white;
            padding: 1.5rem;
            border-radius: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.05);
            margin: 0.5rem 0;
        }
        
        /* Responsive inputs */
        .stTextInput input, .stSelectbox select {
            width: 100%;
            padding: 0.8rem !important;
            border-radius: 10px !important;
            border: 1px solid #e0e0e0 !important;
        }
        
        /* Mobile-friendly metrics */
        .css-1xarl3l {  /* Metric container class */
            padding: 1rem !important;
            border-radius: 10px;
            background: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }
        
        /* Student list styling */
        .student-card {
            background-color: white;
            padding: 1rem;
            border-radius: 10px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            margin: 0.8rem 0;
        }
        
        /* Improved table responsiveness */
        .stDataFrame {
            overflow-x: auto;
        }
        
        /* Better spacing for mobile */
        @media (max-width: 768px) {
            .stApp {
                padding: 0.5rem;
            }
            
            .row-widget {
                margin: 0.5rem 0 !important;
            }
            
            /* Stack columns on mobile */
            .css-12w0qpk {
                flex-direction: column;
            }
            
            .css-1d391kg {
                width: 100% !important;
            }
        }
        
        /* Floating action button for submit */
        .submit-button {
            position: fixed;
            bottom: 20px;
            right: 20px;
            z-index: 999;
            width: auto !important;
            padding: 1rem 2rem !important;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }
        .attendance-status {
            background: #f8fafc;
            border-left: 1px solid #e2e8f0;
            padding: 0.75rem;
            min-width: 120px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
    </style>
""", unsafe_allow_html=True)

# Admin credentials
ADMIN_CREDENTIALS = {
    "a": hashlib.sha256("a".encode()).hexdigest()
}

def format_fraction(value):
    """Format attendance fractions to prevent Excel from converting them to dates"""
    if isinstance(value, str) and '/' in value:
        # Add apostrophe prefix to force Excel to treat it as text
        return f"'{value}"
    return value

def format_text_for_excel(value):
    """Format value as text to prevent Excel conversion"""
    if isinstance(value, str):
        # Add space after forward slash to prevent Excel from treating it as a date
        if '/' in value:
            num, den = value.split('/')
            return f"{num} /{den}"  # Add space before the slash
    return value

def view_statistics_page():
    """Page for viewing attendance statistics with enhanced download options"""
    st.subheader("View Attendance Statistics")
    
    # First select course
    courses = get_courses(for_attendance=False)  
    selected_course = st.selectbox("Select Course", options=[''] + courses)
    
    if selected_course:
        sections = get_sections_by_course(selected_course, for_attendance=False)
        sections = [f"(O){s}" if not s.startswith('(O)') else s for s in sections]
        selected_sections = st.multiselect("Select Sections", options=sections)
        
        if selected_sections:
            # Date range selection
            col1, col2 = st.columns(2)
            with col1:
                from_date = st.date_input("From Date")
            with col2:
                to_date = st.date_input("To Date")
            
            all_stats = []
            for section in selected_sections:
                query_section = section.replace("(O)", "") if section.startswith("(O)") else section
                stats_df = get_attendance_stats(query_section, from_date, to_date)
                if stats_df is not None and not stats_df.empty:
                    stats_df['Section'] = section
                    all_stats.append(stats_df)
            
            if all_stats:
                combined_stats = pd.concat(all_stats, ignore_index=True)
                
                st.write("### Overall Statistics")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Total Students", len(combined_stats))
                with col2:
                    avg_attendance = combined_stats['Overall %'].mean()
                    st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                with col3:
                    below_75 = len(combined_stats[combined_stats['Overall %'] < 75])
                    st.metric("Students Below 75%", below_75)
                
                # Configure column display
                column_config = {
                    'HT Number': st.column_config.TextColumn('HT Number', width=120),
                    'Student Name': st.column_config.TextColumn('Student Name', width=180),
                    'Section': st.column_config.TextColumn('Section', width=150),
                    'Overall %': st.column_config.NumberColumn(
                        'Overall %',
                        format="%.2f%%",
                        width=100
                    )
                }
                
                for col in combined_stats.columns:
                    if 'Attended/Conducted' in col:
                        # Replace column header with A/C notation
                        new_col = col.replace('Attended/Conducted', 'A/C')
                        column_config[col] = st.column_config.TextColumn(
                            new_col,
                            width=150
                        )
                
                st.write("### Student-wise Statistics")
                st.dataframe(
                    combined_stats,
                    column_config=column_config,
                    use_container_width=True,
                    hide_index=True
                    
                )
                
                # Download options
                # Download options with more explicit buttons
                # Download options
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("📊 Download Excel"):
                        # Create Excel file
                        buffer = io.BytesIO()
                        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                            combined_stats.to_excel(writer, sheet_name='Attendance_Stats', index=False)
                            
                            # Format worksheet
                            worksheet = writer.sheets['Attendance_Stats']
                            
                            # Add legend at the top
                            worksheet.insert_rows(0)
                            legend_cell = worksheet.cell(row=1, column=1)
                            legend_cell.value = "A: Attended, C: Conducted"
                            legend_cell.font = Font(italic=True)
                            
                            # Format cells
                            for column in worksheet.columns:
                                max_length = max(len(str(cell.value or '')) for cell in column)
                                worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                                
                            # Add borders and alignment    
                            thin_border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))
                                            
                            for row in worksheet.iter_rows(min_row=2):  # Start from second row due to legend
                                for cell in row:
                                    cell.border = thin_border
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        st.download_button(
                            label="📥 Download Excel File",
                            data=buffer.getvalue(),
                            file_name=f"attendance_stats_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                with col2:
                    if st.button("📄 Download PDF"):
                        try:
                            # Prepare data for PDF
                            pdf_df = combined_stats.copy()
                            
                            # Replace column headers with A/C notation
                            pdf_df.columns = [col.replace('Attended/Conducted', 'A/C') for col in pdf_df.columns]
                            
                            # Add filters info
                            filters = {
                                "Course": selected_course,
                                "Sections": ", ".join(selected_sections),
                                "Date Range": f"{from_date} to {to_date}"
                            }
                            
                            # Generate PDF
                            pdf_data = generate_statistics_pdf(
                                df=pdf_df,
                                title="Attendance Statistics Report",
                                filters=filters
                            )
                            
                            if pdf_data:
                                st.download_button(
                                    label="📥 Download PDF Report",
                                    data=pdf_data,
                                    file_name=f"attendance_stats_{datetime.now().strftime('%Y%m%d')}.pdf",
                                    mime="application/pdf"
                                )
                            else:
                                st.error("Error generating PDF report")
                        except Exception as e:
                            st.error(f"Error generating PDF report: {str(e)}")
            else:
                st.info("No attendance records found for the selected criteria")
                
                
                
def generate_statistics_pdf(df, title, filters=None):
    """
    Generate a PDF report for attendance statistics with column header optimization
    """
    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            @page {
                size: landscape;
                margin: 1.5cm;
            }
            body {
                font-family: Arial, sans-serif;
                background-color: #1a1a1a;
                color: #ffffff;
                margin: 0;
                padding: 20px;
            }
            .legend {
                text-align: center;
                margin-bottom: 20px;
                color: #888;
                font-size: 12px;
            }
            .metrics {
                display: flex;
                justify-content: space-between;
                margin-bottom: 30px;
            }
            .metric {
                text-align: center;
                flex: 1;
                margin: 0 15px;
            }
            .metric-value {
                font-size: 36px;
                font-weight: bold;
                margin: 10px 0;
            }
            .metric-label {
                font-size: 14px;
                color: #888;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                background-color: #1a1a1a;
                margin-top: 20px;
                font-size: 10px;
            }
            th {
                background-color: #333;
                color: #fff;
                padding: 8px;
                text-align: left;
                font-weight: normal;
            }
            td {
                padding: 8px;
                border-top: 1px solid #333;
                color: #fff;
            }
            tr:nth-child(even) {
                background-color: #222;
            }
            .warning-cell {
                color: #ff4444;
            }
            .timestamp {
                text-align: right;
                font-size: 10px;
                color: #666;
                margin-top: 20px;
            }
        </style>
    </head>
    <body>
        <div class="legend">
            A: Attended, C: Conducted
        </div>
        
        <div class="metrics">
            <div class="metric">
                <div class="metric-value">{{ metrics.total_students }}</div>
                <div class="metric-label">Total Students</div>
            </div>
            <div class="metric">
                <div class="metric-value">{{ "%.2f"|format(metrics.avg_attendance) }}%</div>
                <div class="metric-label">Average Attendance</div>
            </div>
            <div class="metric">
                <div class="metric-value">{{ metrics.below_75 }}</div>
                <div class="metric-label">Students Below 75%</div>
            </div>
        </div>

        <table>
            <thead>
                <tr>
                    {% for column in columns %}
                    <th>{{ column | replace('(Attended/Conducted)', '(A/C)') }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for row in data %}
                <tr>
                    {% for value in row %}
                    <td {% if loop.last and value|float < 75 %}class="warning-cell"{% endif %}>
                        {{ value }}
                    </td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
        
        <div class="timestamp">
            Generated on: {{ timestamp }}
        </div>
    </body>
    </html>
    """
    
    # Calculate metrics
    metrics = {
        'total_students': len(df),
        'avg_attendance': df['Overall %'].mean(),
        'below_75': len(df[df['Overall %'] < 75])
    }
    
    # Prepare template data
    template_data = {
        'title': title,
        'filters': filters,
        'metrics': metrics,
        'columns': df.columns.tolist(),
        'data': df.values.tolist(),
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # Generate HTML
    template = Template(html_template)
    html_content = template.render(**template_data)
    
    # PDF options for landscape A4
    pdf_options = {
        'page-size': 'A4',
        'orientation': 'Landscape',
        'margin-top': '15mm',
        'margin-right': '15mm',
        'margin-bottom': '15mm',
        'margin-left': '15mm',
        'encoding': 'UTF-8',
        'no-outline': None,
        'enable-local-file-access': None,
        'background': True,
        'print-media-type': None,
        'no-background': None
    }
    
    # Generate PDF
    pdf = pdfkit.from_string(html_content, False, options=pdf_options)
    
    return pdf



def create_template_df(sheet_name):
    """Create template DataFrame with updated structure"""
    if sheet_name == 'Faculty':
        current_month = datetime.now().strftime('%b%Y')
        return pd.DataFrame(columns=['Faculty Name', 'Username', 'Password', current_month])
    elif sheet_name == 'Section-Subject-Mapping':
        return pd.DataFrame(columns=['Section', 'Subject Names'])
    elif sheet_name == 'Students':
        return pd.DataFrame(columns=[
            'HT Number', 'Student Name', 'Original Section', 'Merged Section',
            'P1', 'P2', 'P3', 'P4', 'P5', 'P6'
        ])
    else:
        st.error(f"Unknown sheet type: {sheet_name}")
        return pd.DataFrame()



def get_subject_analysis(section, subject, from_date=None, to_date=None):
    """Get subject-wise attendance analysis with improved date filtering"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name='Students')
        # Filter students by merged section
        df = df[df['Merged Section'] == section]
        
        analysis = []
        for _, row in df.iterrows():
            if pd.isna(row['HT Number']) or pd.isna(row['Student Name']):
                continue
                
            present = 0
            total = 0
            
            for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                if pd.notna(row[period]) and str(row[period]).strip():
                    entries = str(row[period]).split('\n')
                    for entry in entries:
                        # Check if entry contains the subject
                        if subject in entry:
                            # Parse date from entry
                            try:
                                parts = entry.split('_')
                                if len(parts) >= 5:  # Ensure we have enough parts
                                    date_str = parts[0]
                                    date_obj = datetime.strptime(date_str, '%d/%m/%Y').date()
                                    
                                    # Convert input dates to datetime.date objects if needed
                                    filter_from = from_date
                                    filter_to = to_date
                                    if isinstance(from_date, datetime):
                                        filter_from = from_date.date()
                                    if isinstance(to_date, datetime):
                                        filter_to = to_date.date()
                                    
                                    # Apply date filter if dates are provided
                                    if filter_from and filter_to:
                                        if not (filter_from <= date_obj <= filter_to):
                                            continue
                                    
                                    total += 1
                                    if '_P_' in entry:
                                        present += 1
                            except (ValueError, IndexError) as e:
                                continue
            
            if total > 0:  # Only include students who have classes in the selected period
                percentage = (present / total) * 100
                analysis.append({
                    'HT Number': str(row['HT Number']),
                    'Student Name': str(row['Student Name']),
                    'Original Section': row['Original Section'],
                    'Classes Attended': present,
                    'Total Classes': total,
                    'Attendance %': round(percentage, 2)
                })
        
        if not analysis:
            return pd.DataFrame()
            
        return pd.DataFrame(analysis)
        
    except Exception as e:
        st.error(f"Error in subject analysis: {str(e)}")
        return pd.DataFrame()

def subject_analysis_page():
    """Page for subject-wise analysis with improved date filtering and error handling"""
    st.subheader("Subject-wise Analysis")
    
    try:
        # First select course
        courses = get_courses(for_attendance=True)
        selected_course = st.selectbox("Select Course", options=[''] + courses)
        
        if selected_course:
            # Then show filtered sections
            sections = get_sections_by_course(selected_course, for_attendance=True)
            if not sections:
                st.info("No sections available for selected course")
                return
                
            section = st.selectbox("Select Section", options=sections)
            
            if section:
                # Get subjects for merged section
                subjects = get_section_subjects(section, for_subject_analysis=True)
                if subjects:
                    subject = st.selectbox("Select Subject", options=subjects)
                    
                    if subject:
                        # Add date range selection
                        st.write("### Select Date Range")
                        col1, col2 = st.columns(2)
                        with col1:
                            from_date = st.date_input(
                                "From Date",
                                datetime.now().replace(day=1)
                            )
                        with col2:
                            to_date = st.date_input(
                                "To Date",
                                datetime.now()
                            )
                        
                        # Get analysis with date filtering
                        analysis_df = get_subject_analysis(section, subject, from_date, to_date)
                        
                        if not analysis_df.empty:
                            st.write("### Subject Statistics")
                            col1, col2, col3 = st.columns(3)
                            
                            with col1:
                                avg_attendance = analysis_df['Attendance %'].mean()
                                st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                            with col2:
                                total_classes = analysis_df['Total Classes'].max()
                                st.metric("Total Classes", total_classes)
                            with col3:
                                below_75 = len(analysis_df[analysis_df['Attendance %'] < 75])
                                st.metric("Students Below 75%", below_75)
                            
                            st.write("### Student-wise Analysis")
                            st.dataframe(
                                analysis_df.sort_values('Attendance %', ascending=False),
                                column_config={
                                    'HT Number': st.column_config.TextColumn('HT Number', width=120),
                                    'Student Name': st.column_config.TextColumn('Student Name', width=180),
                                    'Original Section': st.column_config.TextColumn('Original Section', width=150),
                                    'Classes Attended': st.column_config.NumberColumn('Classes Attended', width=130),
                                    'Total Classes': st.column_config.NumberColumn('Total Classes', width=120),
                                    'Attendance %': st.column_config.NumberColumn('Attendance %', format="%.2f%%", width=120)
                                },
                                hide_index=True,
                                use_container_width=True
                            )
                            
                            if st.button("Download Analysis"):
                                buffer = io.BytesIO()
                                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                    analysis_df.to_excel(writer, sheet_name='Subject Analysis', index=False)
                                    
                                    worksheet = writer.sheets['Subject Analysis']
                                    
                                    for column in worksheet.columns:
                                        max_length = max(len(str(cell.value or '')) for cell in column)
                                        worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                                        
                                    thin_border = Border(
                                        left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin')
                                    )
                                    
                                    for row in worksheet.iter_rows():
                                        for cell in row:
                                            cell.border = thin_border
                                            cell.alignment = Alignment(horizontal='center', vertical='center')
                                
                                st.download_button(
                                    label="Download Excel",
                                    data=buffer.getvalue(),
                                    file_name=f"subject_analysis_{section}_{subject}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        else:
                            st.info(f"No attendance records found for {subject} in {section} for the selected date range")
                    else:
                        st.info("Please select a subject to continue")
                else:
                    st.error(f"No subjects found for section '{section}' in Section-Subject-Mapping sheet.")
            else:
                st.info("Please select a section to continue")
        else:
            st.info("Please select a course to continue")
    except Exception as e:
        st.error(f"Error in subject analysis: {str(e)}")


def student_reports_page():
    """Page for individual student reports with original section names"""
    st.subheader("Individual Student Reports")
    
    # Get original sections for reports - keeping original names
    original_sections = get_sections(for_attendance=False)
    selected_sections = st.multiselect("Select Sections", options=original_sections)
    
    if selected_sections:
        try:
            df = pd.read_excel('attendance.xlsx', sheet_name='Students')
            df_filtered = df[df['Original Section'].isin(selected_sections)]
            
            if not df_filtered.empty:
                # Add the date range input fields
                col1, col2 = st.columns(2)
                with col1:
                    from_date = st.date_input("From Date")
                with col2:
                    to_date = st.date_input("To Date")
                
                student = st.selectbox(
                    "Select Student",
                    df_filtered['HT Number'].tolist(),
                    format_func=lambda x: f"{x} - {df_filtered[df_filtered['HT Number']==x]['Student Name'].iloc[0]} ({df_filtered[df_filtered['HT Number']==x]['Original Section'].iloc[0]})"
                )
                
                if student:
                    student_data = df_filtered[df_filtered['HT Number'] == student].iloc[0]
                    
                    st.write(f"### Attendance Report for {student}")
                    st.write(f"**Name:** {student_data['Student Name']}")
                    st.write(f"**Section:** {student_data['Original Section']}")
                    
                    attendance_data = get_student_attendance_details(student_data['Original Section'], student, from_date, to_date)
                    
                    if attendance_data is not None and not attendance_data.empty:
                        # Configure column display
                        column_config = {
                            'Date': st.column_config.TextColumn('Date', width=100),
                            'Time': st.column_config.TextColumn('Time', width=100),
                            'Period': st.column_config.TextColumn('Period', width=80),
                            'Status': st.column_config.TextColumn('Status', width=80),
                            'Faculty': st.column_config.TextColumn('Faculty', width=150),
                            'Subject': st.column_config.TextColumn('Subject', width=150)
                        }
                        
                        st.dataframe(
                            attendance_data.sort_values('Date', ascending=False),
                            column_config=column_config,
                            hide_index=True,
                            use_container_width=True
                        )
                        
                        # In student_reports_page(), replace the download button section with:
                        if st.button("Download Student Report"):
                            # Create Excel file
                            buffer = io.BytesIO()
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                attendance_data.to_excel(writer, sheet_name='Student Report', index=False)
                                
                                # Format worksheet
                                worksheet = writer.sheets['Student Report']
                                
                                # Set column widths
                                for column in worksheet.columns:
                                    max_length = max(len(str(cell.value or '')) for cell in column)
                                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                                    
                                # Add borders and alignment
                                thin_border = Border(left=Side(style='thin'), 
                                                right=Side(style='thin'), 
                                                top=Side(style='thin'), 
                                                bottom=Side(style='thin'))
                                                
                                for row in worksheet.rows:
                                    for cell in row:
                                        cell.border = thin_border
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Offer download as Excel
                            st.download_button(
                                label="Download Excel",
                                data=buffer.getvalue(),
                                file_name=f"student_report_{student}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.info("No attendance records found")
            else:
                st.info("No students found in selected sections")
                
        except Exception as e:
            st.error(f"Error loading student data: {str(e)}")

def get_student_attendance_details(section, student_id, from_date=None, to_date=None):
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name='Students')
        student_row = df[df['HT Number'] == student_id]
        
        if student_row.empty:
            return None
            
        student_row = student_row.iloc[0]
        attendance_data = []
        
        for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
            if pd.notna(student_row[period]) and student_row[period]:
                entries = str(student_row[period]).split('\n')
                for entry in entries:
                    if entry.strip():
                        try:
                            # More flexible splitting to handle different formats
                            parts = entry.split('_')
                            if len(parts) >= 4:  # Minimum required parts
                                date = parts[0]
                                time = parts[1]
                                status = parts[2]
                                faculty = parts[3]
                                # Join remaining parts as subject and plan
                                subject_and_plan = '_'.join(parts[4:]) if len(parts) > 4 else ''
                                
                                # Convert date string to datetime object for filtering
                                date_obj = datetime.strptime(date, '%d/%m/%Y')
                                
                                # Apply date range filter if dates are provided
                                if from_date and to_date:
                                    filter_from = from_date if isinstance(from_date, datetime) else datetime.combine(from_date, datetime.min.time())
                                    filter_to = to_date if isinstance(to_date, datetime) else datetime.combine(to_date, datetime.max.time())
                                    
                                    if not (filter_from <= date_obj <= filter_to):
                                        continue
                                
                                attendance_data.append({
                                    'Date': date,
                                    'Time': time,
                                    'Period': period,
                                    'Status': status,
                                    'Faculty': faculty,
                                    'Subject': subject_and_plan
                                })
                        except Exception as e:
                            st.error(f"Error processing entry: {entry}\nError: {str(e)}")
                            continue
        
        if not attendance_data:
            return pd.DataFrame()
            
        df_attendance = pd.DataFrame(attendance_data)
        
        # Sort by date and time
        df_attendance['DateObj'] = pd.to_datetime(df_attendance['Date'], format='%d/%m/%Y')
        df_attendance = df_attendance.sort_values(['DateObj', 'Time'], ascending=[False, False])
        df_attendance = df_attendance.drop('DateObj', axis=1)
        
        return df_attendance
        
    except Exception as e:
        st.error(f"Error getting attendance details: {str(e)}")
        return None

def get_courses(for_attendance=False):
    """Get unique courses directly from Course column"""
    try:
        # Read the Excel file
        df = pd.read_excel('attendance.xlsx', sheet_name='Students', dtype={
            'Course': str, 
            'Original Section': str,
            'Merged Section': str
        })
        
        # Get unique, non-null courses directly from Course column
        courses = df['Course'].dropna().unique().tolist()
        
        # Remove any empty strings and sort
        courses = sorted([c for c in courses if str(c).strip()])
        
        return courses
        
    except Exception as e:
        st.error(f"Error getting courses: {str(e)}")
        return []

def verify_excel_file():
    """Verify Excel file integrity"""
    try:
        with open('attendance.xlsx', 'rb') as file:
            # Read file content
            content = file.read()
            # Basic integrity check
            if len(content) < 100:  # Too small to be valid
                return False
            # Try opening with pandas
            pd.read_excel('attendance.xlsx', sheet_name=None)
            return True
    except Exception:
        return False

def backup_excel_file():
    """Create backup of Excel file"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = f'backup/attendance_{timestamp}.xlsx'
        os.makedirs('backup', exist_ok=True)
        shutil.copy2('attendance.xlsx', backup_path)
        # Keep only last 5 backups
        backups = sorted(glob.glob('backup/attendance_*.xlsx'))
        if len(backups) > 5:
            os.remove(backups[0])
    except Exception as e:
        st.error(f"Backup failed: {str(e)}")

def get_sections_by_course(course, for_attendance=False):
    """Get sections filtered by course with enhanced error handling"""
    try:
        if not course:  # If no course selected, return empty list
            return []
            
        # Add file verification before reading
        if not os.path.exists('attendance.xlsx'):
            st.error("Attendance file not found")
            return []
            
        # Try reading with repair mode
        try:
            df = pd.read_excel('attendance.xlsx', sheet_name='Students', dtype={
                'Course': str,
                'Original Section': str,
                'Merged Section': str
            })
        except Exception as e:
            # If normal read fails, try to repair
            st.warning("Attempting to repair Excel file...")
            # Create backup
            backup_excel_file()
            # Try alternate reading method
            df = pd.read_excel('attendance.xlsx', sheet_name='Students', engine='openpyxl')
        
        # Filter by course
        df = df[df['Course'] == course]
        if for_attendance:
            # For attendance marking: use merged sections
            sections = df['Merged Section'].dropna().unique().tolist()
        else:
            # For analytics: use original sections
            sections = df['Original Section'].dropna().unique().tolist()
            sections = [f"(O){s}" if not s.startswith('(O)') else s for s in sections]
        
        # Remove any empty strings and sort
        sections = sorted([s for s in sections if str(s).strip()])
        
        return sections
        
    except Exception as e:
        st.error(f"Error getting sections for course: {str(e)}")
        return []


def get_attendance_stats(section, from_date=None, to_date=None):
    """Calculate attendance statistics with attended/conducted format"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name='Students')
        # Filter students by original section without modifying section name
        df = df[df['Original Section'] == section]
        
        if df.empty:
            return None
        
        # Get subjects for this section from Section-Subject-Mapping
        subjects = get_section_subjects(section)
        if not subjects:
            st.error(f"No subjects found for section {section}")
            return None

        # Convert date inputs to datetime.date objects if they're strings
        if isinstance(from_date, str):
            from_date = datetime.strptime(from_date, '%d/%m/%Y').date()
        if isinstance(to_date, str):
            to_date = datetime.strptime(to_date, '%d/%m/%Y').date()
        
        stats = []
        for _, row in df.iterrows():
            if pd.isna(row['HT Number']) or pd.isna(row['Student Name']):
                continue
                
            student_stats = {
                'HT Number': str(row['HT Number']),
                'Student Name': str(row['Student Name'])
            }
            
            total_present = 0
            total_classes = 0
            
            # Calculate for each mapped subject
            for subject in subjects:
                present = 0
                total = 0
                
                for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                    if pd.notna(row[period]) and str(row[period]).strip():
                        entries = str(row[period]).split('\n')
                        for entry in entries:
                            try:
                                parts = entry.split('_')
                                if len(parts) >= 5:
                                    date_str = parts[0]
                                    entry_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                                    
                                    # Check if entry is within date range
                                    if from_date and to_date:
                                        if not (from_date <= entry_date <= to_date):
                                            continue
                                    
                                    # Check if this entry is for the current subject
                                    subject_name = parts[4]
                                    if subject in subject_name:
                                        total += 1
                                        total_classes += 1
                                        if '_P_' in entry:
                                            present += 1
                                            total_present += 1
                            except (ValueError, IndexError) as e:
                                continue
                
                # Only add subject column if there are classes for this subject
                # Inside get_attendance_stats function, when creating fraction strings:
                if total > 0:
                    fraction = f"{present}/{total}"  # No space initially
                    student_stats[f"{subject}\n(Attended/Conducted)"] = fraction

                # And for total:
                if total_classes > 0:
                    total_fraction = f"{total_present}/{total_classes}"  # No space initially
                    student_stats[f"Total\n(Attended/Conducted)"] = total_fraction
            
            # Only add student stats if they attended any classes
            if total_classes > 0:
                student_stats[f"Total\n(Attended/Conducted)"] = f"{total_present}/{total_classes}"
                student_stats['Overall %'] = round((total_present / total_classes * 100), 2)
                stats.append(student_stats)
        
        if not stats:
            return pd.DataFrame()
        
        # Convert to DataFrame and ensure correct column order
        stats_df = pd.DataFrame(stats)
        base_columns = ['HT Number', 'Student Name']
        subject_columns = [col for col in stats_df.columns 
                         if '(Attended/Conducted)' in col and 'Total' not in col]
        total_column = [col for col in stats_df.columns if col.startswith('Total')]
        percentage_column = ['Overall %']
        
        ordered_columns = base_columns + subject_columns + total_column + percentage_column
        return stats_df[ordered_columns]
        
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return None


def get_section_subjects(section, for_subject_analysis=False):
    """Get subjects for a section from Section-Subject-Mapping sheet
    
    Args:
        section: Section name
        for_subject_analysis: If True, use section name as is; if False, get merged section first
    """
    try:
        # First read the students sheet to get the merged section if needed
        lookup_section = section
        if not for_subject_analysis and '(O)' in section:
            df_students = pd.read_excel('attendance.xlsx', sheet_name='Students')
            # Find the first student from this original section to get their merged section
            student_row = df_students[df_students['Original Section'] == section].iloc[0]
            lookup_section = student_row['Merged Section']
        
        # Now lookup the subjects using the correct section name
        df_mapping = pd.read_excel('attendance.xlsx', sheet_name='Section-Subject-Mapping')
        matching_rows = df_mapping[df_mapping['Section'] == lookup_section]
        
        if matching_rows.empty:
            st.error(f"No subject mapping found for section: {lookup_section}")
            return []
        
        # Get subjects from the first matching row
        subjects_str = str(matching_rows.iloc[0]['Subject Names'])
        # Split subjects and clean up (handle both \n and regular spaces)
        # First split by newline, then by multiple spaces
        subjects = []
        for line in subjects_str.split('\n'):
            # Clean and add each subject
            line = line.strip()
            if line:
                subjects.append(line)
        
        return subjects
        
    except Exception as e:
        st.error(f"Error getting subjects: {str(e)}")
        return []








def get_sections(for_attendance=False):
    """Get sections based on context
    for_attendance=True: Returns merged sections for attendance marking
    for_attendance=False: Returns original sections"""
    try:
        # Read the Excel file with string dtype for relevant columns
        df = pd.read_excel('attendance.xlsx', sheet_name='Students', dtype={
            'Original Section': str,
            'Merged Section': str,
            'Course': str
        })
        
        if for_attendance:
            # For attendance marking: return unique merged sections
            sections = df['Merged Section'].dropna().unique().tolist()
        else:
            # For analytics: return unique original sections with (O) prefix
            sections = df['Original Section'].dropna().unique().tolist()
            sections = [f"(O){s}" if not s.startswith('(O)') else s for s in sections]
        
        # Filter out any empty strings or whitespace
        sections = [s for s in sections if s and str(s).strip()]
        
        # Sort the sections
        return sorted(sections)
        
    except Exception as e:
        st.error(f"Error getting sections: {str(e)}")
        return []

# Modify get_student_data function to keep original section names
def get_student_data(section, for_attendance=True):
    """Get student data for a section"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name='Students')
        if for_attendance:
            # For attendance marking, get students based on merged section
            students_df = df[df['Merged Section'] == section][['HT Number', 'Student Name', 'Original Section']].fillna('')
        else:
            # For analytics, get students based on original section
            students_df = df[df['Original Section'] == section][['HT Number', 'Student Name', 'Original Section']].fillna('')
        
        return students_df
    except Exception as e:
        st.error(f"Error getting student data: {str(e)}")
        return None



def check_duplicate_attendance(section, period, date_str):
    """
    Check if attendance is already marked for given section, period and date
    Returns tuple: (bool, str) - (is_duplicate, faculty_name who marked it)
    """
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name='Students')
        # Filter for students in the given merged section
        section_df = df[df['Merged Section'] == section]
        
        if period in section_df.columns:
            for _, row in section_df.iterrows():
                if pd.notna(row[period]):
                    entries = str(row[period]).split('\n')
                    for entry in entries:
                        if entry.strip():
                            # Extract date and faculty from the entry
                            try:
                                parts = entry.split('_')
                                if len(parts) >= 5:
                                    entry_date = parts[0]
                                    # Only check if dates match exactly
                                    if entry_date == date_str:
                                        faculty_name = parts[3]
                                        return True, faculty_name
                            except:
                                continue
            return False, None
        return False, None
    except Exception as e:
        st.error(f"Error checking duplicate attendance: {str(e)}")
        return False, None

def check_existing_attendance(section, period):
    """
    Check if any attendance exists for this merged section and period for today
    Returns tuple: (bool, str, str) - (has_existing, faculty_name, date)
    """
    try:
        current_date = datetime.now().strftime('%d/%m/%Y')
        df = pd.read_excel('attendance.xlsx', sheet_name='Students', dtype=str)
        
        # Filter for students in the merged section
        section_df = df[df['Merged Section'] == section]
        
        if period in section_df.columns:
            for _, row in section_df.iterrows():
                if pd.notna(row[period]):
                    entries = str(row[period]).split('\n')
                    for entry in entries:
                        if entry.strip():
                            try:
                                parts = entry.split('_')
                                if len(parts) >= 4:
                                    entry_date = parts[0]
                                    # Only block if trying to mark attendance twice on same day
                                    if entry_date == current_date:
                                        faculty_name = parts[3]
                                        return True, faculty_name, entry_date
                            except:
                                continue
            return False, None, None
        return False, None, None
    except Exception as e:
        st.error(f"Error checking attendance: {str(e)}")
        return False, None, None




def get_column_width(col_name, values):
    try:
        max_length = max(
            max(len(str(val)) for val in values if val is not None),
            len(str(col_name))
        )
        return min(max_length * 10, 300)
    except:
        return 150








def get_all_faculty_workload(from_date=None, to_date=None):
    """Get workload statistics for all faculty members with combined section handling"""
    try:
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        faculty_stats = []
        
        for _, faculty_row in df_faculty.iterrows():
            faculty_name = faculty_row['Faculty Name']
            workload_data = []
            combined_classes = {}  # Track classes by date and period
            
            # Process each month column
            for col in faculty_row.index:
                if col not in ['Faculty Name', 'Password', 'Username']:
                    entries = str(faculty_row[col]).split('\n') if pd.notna(faculty_row[col]) else []
                    
                    for entry in entries:
                        if entry.strip():
                            try:
                                parts = entry.strip().split('_')
                                if len(parts) >= 5:
                                    date_str = parts[0]
                                    time_str = parts[1]
                                    period = parts[2]
                                    subject = parts[3]
                                    section = parts[4]
                                    
                                    # Convert date for filtering
                                    date_obj = pd.to_datetime(date_str, format='%d/%m/%Y')
                                    
                                    # Apply date filter if provided
                                    if from_date and to_date:
                                        if not (pd.to_datetime(from_date) <= date_obj <= pd.to_datetime(to_date)):
                                            continue
                                    
                                    # Create unique key for date-period combination
                                    date_period_key = f"{date_str}_{period}"
                                    
                                    # Track combined sections
                                    if date_period_key not in combined_classes:
                                        combined_classes[date_period_key] = {
                                            'sections': [],
                                            'time': time_str,
                                            'subject': subject
                                        }
                                    combined_classes[date_period_key]['sections'].append(section)
                                    
                            except Exception:
                                continue
            
            # Process combined classes into workload data
            for date_period_key, class_info in combined_classes.items():
                date_str, period = date_period_key.split('_')
                num_sections = len(class_info['sections'])
                
                # Create one entry per section with distributed workload
                for section in class_info['sections']:
                    workload_data.append({
                        'Date': date_str,
                        'Time': class_info['time'],
                        'Period': period,
                        'Subject': class_info['subject'],
                        'Section': section,
                        'Combined Sections': ', '.join(class_info['sections']),
                        'Workload': 1/num_sections  # Distribute workload among sections
                    })
            
            # Calculate statistics for this faculty
            if workload_data:
                df_workload = pd.DataFrame(workload_data)
                unique_days = len(df_workload['Date'].unique())
                unique_subjects = len(df_workload['Subject'].unique())
                unique_sections = len(df_workload['Section'].unique())
                total_classes = len(combined_classes)  # Count unique date-period combinations
                actual_workload = df_workload['Workload'].sum()  # Total distributed workload
                
                # Get subject and section distribution with workload
                subject_dist = df_workload.groupby('Subject')['Workload'].sum().round(1).to_dict()
                section_dist = df_workload.groupby('Section')['Workload'].sum().round(1).to_dict()
                
                faculty_stats.append({
                    'Faculty Name': faculty_name,
                    'Total Classes': total_classes,
                    'Actual Workload': actual_workload,
                    'Days Engaged': unique_days,
                    'Daily Average': round(actual_workload / max(unique_days, 1), 2),
                    'Unique Subjects': unique_subjects,
                    'Unique Sections': unique_sections,
                    'Subject Distribution': subject_dist,
                    'Section Distribution': section_dist,
                    'Detailed Records': df_workload
                })
        
        return faculty_stats
    except Exception as e:
        st.error(f"Error calculating faculty workload: {str(e)}")
        return []


def admin_page():
    """Updated admin page with course filtering"""
    st.title("Admin Dashboard")
    
    with st.sidebar:
        st.header("Navigation")
        page = st.radio(
            "Select",
            ["Faculty Workload","Data Management",  "Reset Credentials"]
        )
    
    if page == "Reset Credentials":
        st.subheader("Reset Password")
        current_password = st.text_input("Current Password", type="password", key="current_pwd")
        new_password = st.text_input("New Password", type="password", key="new_pwd")
        confirm_password = st.text_input("Confirm New Password", type="password", key="confirm_pwd")
        
        if st.button("Update Password", type="primary"):
            try:
                if not all([current_password, new_password, confirm_password]):
                    st.error("All fields are required")
                    return
                    
                # Read faculty data
                df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
                
                # Convert credentials columns to string and strip whitespace
                df_faculty['Username'] = df_faculty['Username'].astype(str).str.strip()
                df_faculty['Password'] = df_faculty['Password'].astype(str).str.strip()
                
                # Check current credentials
                user_mask = (df_faculty['Username'] == st.session_state.username) & \
                           (df_faculty['Password'] == current_password)
                           
                if not user_mask.any():
                    st.error("Current password is incorrect")
                    return
                
                # Verify new passwords match
                if new_password != confirm_password:
                    st.error("New passwords do not match")
                    return
                
                # Update password
                df_faculty.loc[user_mask, 'Password'] = new_password
                
                # Save changes
                with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                    df_faculty.to_excel(writer, sheet_name='Faculty', index=False)
                    
                    # Format worksheet
                    worksheet = writer.sheets['Faculty']
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    for column in worksheet.columns:
                        max_length = max(len(str(cell.value or '')) for cell in column)
                        worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                
                st.success("Password updated successfully! Please login again.")
                
                # Clear session state to force re-login
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
                
            except Exception as e:
                st.error(f"Error updating password: {str(e)}")
        return
    
    elif page == "Data Management":
        try:
            with st.sidebar:
                st.header("Data Management")
                sheet = st.selectbox(
                    "Select Sheet",
                    ["Students", "Faculty", "Section-Subject-Mapping"]
                )
                
                st.write("### Download Options")
                
                # Add course filter for data management
                if sheet == "Students":
                    courses = get_courses(for_attendance=False)
                    selected_course = st.selectbox(
                        "Filter by Course",
                        options=['All'] + courses,
                        key="dm_course_filter"
                    )
                
                # Upload complete workbook button
                uploaded_workbook = st.file_uploader("Upload Complete Workbook", type=['xlsx'])
                if uploaded_workbook is not None:
                    try:
                        xls = pd.ExcelFile(uploaded_workbook)
                        st.write("Sheets found in workbook:", xls.sheet_names)
                        
                        if st.button("Confirm Upload"):
                            with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                                for sheet_name in xls.sheet_names:
                                    df = pd.read_excel(uploaded_workbook, sheet_name=sheet_name)
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                                    
                                    # Format worksheet
                                    worksheet = writer.sheets[sheet_name]
                                    for row in worksheet.iter_rows():
                                        for cell in row:
                                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                                    
                                    for column in worksheet.columns:
                                        max_length = max(len(str(cell.value or '')) for cell in column)
                                        worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                                        
                            st.success("Workbook uploaded successfully!")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Error uploading workbook: {str(e)}")
                
                # Download complete workbook
                if st.button("Download Complete Workbook"):
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        sheets = ["Students", "Faculty", "Section-Subject-Mapping"]
                        for sheet_name in sheets:
                            df = pd.read_excel('attendance.xlsx', sheet_name=sheet_name)
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            # Format worksheet
                            worksheet = writer.sheets[sheet_name]
                            for column in worksheet.columns:
                                max_length = max(len(str(cell.value or '')) for cell in column)
                                worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                    
                    st.download_button(
                        label="📥 Download Complete Workbook",
                        data=buffer.getvalue(),
                        file_name="attendance_workbook.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            # Add tabs for different operations
            tab1, tab2 = st.tabs(["Edit Data", "Bulk Upload"])
            
            with tab1:
                st.subheader(f"Edit {sheet}")
                if sheet == "Students" and "dm_course_filter" in st.session_state:
                    show_data_editor(sheet, course_filter=st.session_state.dm_course_filter)
                else:
                    show_data_editor(sheet)
            
            with tab2:
                st.subheader("Bulk Upload")
                show_bulk_upload(sheet)
                
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
    
    else:  # Faculty Workload
        show_faculty_workload_admin()
    
    # Logout button at the bottom of sidebar
    with st.sidebar:
        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()



def show_data_editor(sheet, course_filter='All'):
    """Show the data editor component with improved layout and course filtering"""
    try:
        if sheet == 'Students':
            # Load student data with proper column configuration
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            df = df.fillna('')
            
            # Apply course filter if specified
            if course_filter != 'All':
                df = df[df['Original Section'].str.startswith(course_filter) | 
                       df['Merged Section'].str.startswith(course_filter)]
            
            # Configure columns with appropriate widths
            column_config = {
                'HT Number': st.column_config.TextColumn('HT Number', width=120),
                'Student Name': st.column_config.TextColumn('Student Name', width=180),
                'Original Section': st.column_config.TextColumn('Original Section', width=150),
                'Merged Section': st.column_config.TextColumn('Merged Section', width=150),
                'P1': st.column_config.TextColumn('P1', width=300),
                'P2': st.column_config.TextColumn('P2', width=300),
                'P3': st.column_config.TextColumn('P3', width=300),
                'P4': st.column_config.TextColumn('P4', width=300),
                'P5': st.column_config.TextColumn('P5', width=300),
                'P6': st.column_config.TextColumn('P6', width=300)
            }
        elif sheet == 'Faculty':
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            df = df.fillna('')
            
            column_config = {
                'Faculty Name': st.column_config.TextColumn('Faculty Name', width=180),
                'Username': st.column_config.TextColumn('Username', width=120),
                'Password': st.column_config.TextColumn('Password', width=120)
            }
        else:  # Section-Subject-Mapping
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            df = df.fillna('')
            
            if course_filter != 'All':
                df = df[df['Section'].str.startswith(course_filter)]
            
            column_config = {col: st.column_config.TextColumn(col, width=150) for col in df.columns}
        
        # Get actual number of rows
        data_rows = len(df)
        
        # Display editor with minimal extra rows
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            column_config=column_config,
            hide_index=True,
            height=min(600, (data_rows + 2) * 35)
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("Save Changes", type="primary"):
                try:
                    with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                        edited_df.to_excel(writer, sheet_name=sheet, index=False)
                        
                        worksheet = writer.sheets[sheet]
                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        for column in worksheet.columns:
                            max_length = max(len(str(cell.value or '')) for cell in column)
                            worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                    
                    st.success("Changes saved successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error saving changes: {str(e)}")
        
        with col2:
            # Download current filtered view
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                edited_df.to_excel(writer, sheet_name=sheet, index=False)
            
            st.download_button(
                label="Download Current View",
                data=buffer.getvalue(),
                file_name=f"{sheet}_{course_filter}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")

def show_bulk_upload(sheet):
    """Show bulk upload interface with course-aware template structure"""
    if st.button("Download Template"):
        template_df = create_template_df(sheet)
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            template_df.to_excel(writer, sheet_name=sheet, index=False)
            worksheet = writer.sheets[sheet]
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2
        
        st.download_button(
            label="📥 Download Template",
            data=buffer.getvalue(),
            file_name=f"{sheet}_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    uploaded_file = st.file_uploader("Upload Excel file", type=['xlsx'])
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            st.write("Preview of uploaded data:")
            st.dataframe(df)
            
            if st.button("Confirm Upload"):
                if validate_upload_data(df, sheet):
                    # If it's the Students sheet, get the course information
                    if sheet == 'Students':
                        courses = get_courses(for_attendance=False)
                        selected_course = st.selectbox(
                            "Select Course for Upload",
                            options=courses
                        )
                        
                        # Validate that sections match the selected course
                        if not all(row['Original Section'].startswith(selected_course) for _, row in df.iterrows()):
                            st.error(f"Some sections don't match the selected course {selected_course}")
                            return
                    
                    with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, sheet_name=sheet, index=False)
                        
                        worksheet = writer.sheets[sheet]
                        for column in worksheet.columns:
                            max_length = max(len(str(cell.value or '')) for cell in column)
                            worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                    
                    st.success("Data uploaded successfully!")
                    st.rerun()
                else:
                    st.error("Invalid data format. Please use the template.")
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")


# Update the display code in show_faculty_workload_admin()
def show_faculty_workload_admin():
    """Enhanced display function for admin view with additional analytics"""
    st.subheader("Faculty Workload Overview")
    
    # Classes Section with Date Filters
    st.write("### Class Schedule")
    col1, col2 = st.columns(2)
    with col1:
        selected_date = st.date_input(
            "Select Date",
            datetime.now(),
            key="class_schedule_date"
        )
    with col2:
        view_type = st.selectbox(
            "View Type",
            ["Single Day", "Date Range"],
            key="class_view_type"
        )
    
    if view_type == "Date Range":
        end_date = st.date_input(
            "End Date",
            selected_date,
            key="class_schedule_end_date"
        )
    else:
        end_date = selected_date
    
    try:
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        classes_data = []
        
        # Convert dates to string format for comparison
        start_date_str = selected_date.strftime('%d/%m/%Y')
        end_date_str = end_date.strftime('%d/%m/%Y')
        
        for _, faculty_row in df_faculty.iterrows():
            faculty_name = faculty_row['Faculty Name']
            for col in faculty_row.index:
                if col not in ['Faculty Name', 'Username', 'Password']:
                    entries = str(faculty_row[col]).split('\n') if pd.notna(faculty_row[col]) else []
                    for entry in entries:
                        if pd.notna(entry) and entry.strip():
                            try:
                                parts = entry.strip().split('_')
                                if len(parts) >= 6:
                                    entry_date = parts[0]
                                    # Check if date is within selected range
                                    if start_date_str <= entry_date <= end_date_str:
                                        classes_data.append({
                                            'Date': entry_date,
                                            'Time': parts[1],
                                            'Period': parts[2],
                                            'Faculty': faculty_name,
                                            'Subject': parts[3],
                                            'Section': parts[4],
                                            'Status': 'Completed'
                                        })
                            except Exception:
                                continue
        
        if classes_data:
            df_classes = pd.DataFrame(classes_data)
            df_classes = df_classes.sort_values(['Date', 'Time', 'Period'])
            
            if view_type == "Date Range":
                st.write(f"### Classes from {start_date_str} to {end_date_str}")
            else:
                st.write(f"### Classes on {start_date_str}")
                
            # Add summary metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Classes", len(df_classes))
            with col2:
                st.metric("Faculty Engaged", df_classes['Faculty'].nunique())
            with col3:
                st.metric("Sections Covered", df_classes['Section'].nunique())
            
            st.dataframe(
                df_classes,
                column_config={
                    'Date': st.column_config.TextColumn('Date', width=100),
                    'Time': st.column_config.TextColumn('Time', width=80),
                    'Period': st.column_config.TextColumn('Period', width=70),
                    'Faculty': st.column_config.TextColumn('Faculty', width=150),
                    'Subject': st.column_config.TextColumn('Subject', width=120),
                    'Section': st.column_config.TextColumn('Section', width=120),
                    'Status': st.column_config.TextColumn(
                        'Status',
                        width=100,
                        help="Shows if attendance has been marked"
                    )
                },
                hide_index=True,
                use_container_width=True
            )
            
            # Download option for filtered data
            if st.button("Download Schedule"):
                csv = df_classes.to_csv(index=False)
                st.download_button(
                    label="Download CSV",
                    data=csv,
                    file_name=f"class_schedule_{start_date_str}_to_{end_date_str}.csv",
                    mime="text/csv"
                )

            # Add Missing Attendance Analysis section
            st.write("### Missing Attendance Analysis")
            missing_data, all_sections = analyze_missing_attendance(start_date_str, df_classes)
            
            if missing_data:
                st.warning("⚠️ The following sections have not had attendance marked:")
                
                for period_data in missing_data:
                    with st.expander(f"Period {period_data['Period']}"):
                        # Create metrics for the period
                        total_missing = len(period_data['Missing Sections'])
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Missing Sections", total_missing)
                        
                        with col2:
                            coverage = ((len(all_sections) - total_missing) / len(all_sections)) * 100
                            st.metric("Attendance Coverage", f"{coverage:.1f}%")
                        
                        # Display missing sections in a organized way
                        st.write("##### Sections without Attendance:")
                        for section in period_data['Missing Sections']:
                            st.markdown(f"""
                                <div style='background: #FF575720; 
                                          padding: 8px 15px; 
                                          border-radius: 5px; 
                                          border-left: 4px solid #FF5757;
                                          margin: 5px 0;'>
                                    {section}
                                </div>
                            """, unsafe_allow_html=True)
            else:
                st.success("✅ All sections have attendance marked for all periods!")
                
        else:
            if view_type == "Date Range":
                st.info(f"No classes recorded between {start_date_str} and {end_date_str}")
            else:
                st.info(f"No classes recorded on {start_date_str}")
            
    except Exception as e:
        st.error(f"Error loading class schedule: {str(e)}")
    
    # Historical Analysis section
    st.write("### Historical Analysis")
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input(
            "From Date",
            datetime.now().replace(day=1),
            format="YYYY/MM/DD"
        )
    with col2:
        to_date = st.date_input(
            "To Date",
            datetime.now(),
            format="YYYY/MM/DD"
        )
    
    # Course filter
    courses = get_courses(for_attendance=True)
    selected_course = st.selectbox(
        "Filter by Course",
        options=['All'] + courses,
        key="admin_workload_course"
    )
    
    try:
        faculty_stats = get_all_faculty_workload(from_date, to_date)
        
        if faculty_stats:
            # Filter and process stats
            if selected_course != 'All':
                for stat in faculty_stats:
                    stat['Detailed Records'] = stat['Detailed Records'][
                        stat['Detailed Records']['Section'].str.startswith(selected_course)
                    ]
                    stat['Subject Distribution'] = stat['Detailed Records']['Subject'].value_counts().to_dict()
                    stat['Section Distribution'] = stat['Detailed Records']['Section'].value_counts().to_dict()
                    stat['Total Classes'] = len(stat['Detailed Records'])
                    if stat['Total Classes'] > 0:
                        stat['Days Engaged'] = stat['Detailed Records']['Date'].nunique()
                        stat['Daily Average'] = stat['Total Classes'] / stat['Days Engaged']
            
            faculty_stats = [stat for stat in faculty_stats if stat['Total Classes'] > 0]
            
            if faculty_stats:
                # Calculate additional metrics
                total_faculty = len(faculty_stats)
                total_classes = sum(stat['Total Classes'] for stat in faculty_stats)
                avg_classes = total_classes / total_faculty if total_faculty > 0 else 0
                max_daily_avg = max(stat['Daily Average'] for stat in faculty_stats)
                min_daily_avg = min(stat['Daily Average'] for stat in faculty_stats)
                
                # Display overall metrics
                st.write("### Overall Statistics")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Faculty", total_faculty)
                    st.metric("Maximum Daily Load", f"{max_daily_avg:.1f}")
                with col2:
                    st.metric("Total Classes", total_classes)
                    st.metric("Minimum Daily Load", f"{min_daily_avg:.1f}")
                with col3:
                    st.metric("Avg Classes per Faculty", f"{avg_classes:.1f}")
                    st.metric("Active Teaching Days", max(stat['Days Engaged'] for stat in faculty_stats))
                
                # Faculty Performance Metrics
                st.write("### Faculty Performance Analysis")
                
                # Create performance dataframe
                performance_data = []
                for stat in faculty_stats:
                    performance_data.append({
                        'Faculty Name': stat['Faculty Name'],
                        'Total Classes': stat['Total Classes'],
                        'Days Engaged': stat['Days Engaged'],
                        'Daily Average': round(stat['Daily Average'], 2),
                        'Subjects': stat['Unique Subjects'],
                        'Sections': stat['Unique Sections'],
                        'Coverage Score': round((stat['Unique Sections'] * stat['Unique Subjects']) / 2, 2)
                    })
                
                df_performance = pd.DataFrame(performance_data)
                df_performance = df_performance.sort_values('Total Classes', ascending=False)
                
                st.dataframe(
                    df_performance,
                    column_config={
                        'Faculty Name': st.column_config.TextColumn('Faculty Name', width=200),
                        'Total Classes': st.column_config.NumberColumn('Total Classes', width=100),
                        'Days Engaged': st.column_config.NumberColumn('Teaching Days', width=100),
                        'Daily Average': st.column_config.NumberColumn('Daily Avg', format="%.2f", width=100),
                        'Subjects': st.column_config.NumberColumn('Subjects', width=90),
                        'Sections': st.column_config.NumberColumn('Sections', width=90),
                        'Coverage Score': st.column_config.NumberColumn(
                            'Coverage Score',
                            help="Composite score based on subject and section coverage",
                            format="%.2f",
                            width=120
                        )
                    },
                    hide_index=True,
                    use_container_width=True
                )
                
                # Search functionality
                st.write("### Individual Faculty Details")
                search_query = st.text_input("🔍 Search Faculty", placeholder="Enter faculty name or ID...")
                
                # Filter faculty stats based on search query
                filtered_stats = faculty_stats
                if search_query:
                    filtered_stats = [
                        stat for stat in faculty_stats 
                        if search_query.lower() in stat['Faculty Name'].lower()
                    ]
                
                # Detailed Faculty Records
                for stat in filtered_stats:
                    with st.expander(f"📊 {stat['Faculty Name']}"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write("##### Subject Distribution")
                            subject_df = pd.DataFrame([
                                {'Subject': subject, 'Classes': count}
                                for subject, count in stat['Subject Distribution'].items()
                            ]).sort_values('Classes', ascending=False)
                            st.dataframe(subject_df, hide_index=True)
                            
                            # Add subject coverage metrics
                            total_subject_classes = subject_df['Classes'].sum()
                            if total_subject_classes > 0:
                                st.metric(
                                    "Subject Balance Score",
                                    f"{(1 - subject_df['Classes'].std() / subject_df['Classes'].mean()):.2f}",
                                    help="Score between 0-1 indicating how evenly distributed the subjects are"
                                )
                        
                        with col2:
                            st.write("##### Section Distribution")
                            section_df = pd.DataFrame([
                                {'Section': section, 'Classes': count}
                                for section, count in stat['Section Distribution'].items()
                            ]).sort_values('Classes', ascending=False)
                            st.dataframe(section_df, hide_index=True)
                            
                            # Add section coverage metrics
                            if len(section_df) > 0:
                                st.metric(
                                    "Section Coverage",
                                    f"{(len(section_df) / len(courses) * 100):.1f}%",
                                    help="Percentage of total sections covered"
                                )
                        
                        st.write("##### Detailed Class Records")
                        detailed_df = stat['Detailed Records'].sort_values(['Date', 'Time'], ascending=[False, False])
                        st.dataframe(
                            detailed_df,
                            column_config={
                                'Date': st.column_config.TextColumn('Date', width=100),
                                'Time': st.column_config.TextColumn('Time', width=100),
                                'Period': st.column_config.TextColumn('Period', width=80),
                                'Subject': st.column_config.TextColumn('Subject', width=150),
                                'Section': st.column_config.TextColumn('Section', width=150),
                                'Lesson Plan': st.column_config.TextColumn('Lesson Plan', width=300)
                            },
                            hide_index=True,
                            use_container_width=True
                        )
                        
                        # Download option for individual faculty
                        csv = detailed_df.to_csv(index=False)
                        st.download_button(
                            label=f"Download {stat['Faculty Name']}'s Records",
                            data=csv,
                            file_name=f"workload_{stat['Faculty Name']}_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
                
                # Download complete report
                st.write("### Download Options")
                if st.button("Download Complete Report"):
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        # Write performance summary
                        df_performance.to_excel(writer, sheet_name='Performance Summary', index=False)
                        
                        # Write individual faculty sheets
                        for stat in faculty_stats:
                            sheet_name = stat['Faculty Name'][:30]
                            stat['Detailed Records'].to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    st.download_button(
                        label="📥 Download Complete Analysis Report",
                        data=buffer.getvalue(),
                        file_name=f"faculty_workload_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.info(f"No faculty workload data found for {selected_course}")
        else:
            st.info("No faculty workload data available for the selected date range")
            
    except Exception as e:
        st.error(f"Error loading faculty workload: {str(e)}")


def analyze_missing_attendance(selected_date, df_classes):
    """Analyze which sections are missing attendance for each period"""
    try:
        # Get all merged sections from Students sheet
        df_students = pd.read_excel('attendance.xlsx', sheet_name='Students')
        all_sections = df_students['Merged Section'].unique()
        
        # Get all periods
        periods = ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']
        
        # Create record of covered sections per period
        covered_sections = {period: set() for period in periods}
        
        # Record sections that have attendance marked
        for _, row in df_classes.iterrows():
            covered_sections[row['Period']].add(row['Section'])
        
        # Find missing sections for each period
        missing_data = []
        for period in periods:
            missing = set(all_sections) - covered_sections[period]
            if missing:
                missing_data.append({
                    'Period': period,
                    'Missing Sections': sorted(list(missing))
                })
        
        return missing_data, all_sections
    except Exception as e:
        st.error(f"Error analyzing missing attendance: {str(e)}")
        return [], []


def get_faculty_workload(username, include_lesson_plans=True):
    """Calculate faculty workload with combined section handling and precise date filtering"""
    try:
        # Set timezone to IST
        ist = pytz.timezone('Asia/Kolkata')
        
        # Get faculty sheet data
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        
        # Get faculty name for workload lookup
        user_mask = df_faculty['Username'] == username
        if not user_mask.any():
            return 0, pd.DataFrame()
            
        workload_data = []
        combined_classes = {}  # Track classes by date and period
        
        # Process each month column
        for col in df_faculty.columns:
            if col not in ['Faculty Name', 'Username', 'Password']:
                entries = str(df_faculty.loc[user_mask, col].iloc[0]).split('\n') if pd.notna(df_faculty.loc[user_mask, col].iloc[0]) else []
                
                for entry in entries:
                    if pd.notna(entry) and entry.strip():
                        try:
                            parts = entry.strip().split('_')
                            if len(parts) >= 6:
                                date_str = parts[0]
                                time_str = parts[1]
                                period = parts[2]
                                subject = parts[3]
                                section = parts[4]
                                lesson_plan = '_'.join(parts[5:]) if include_lesson_plans else ''
                                
                                # Convert date for filtering with exact date match
                                date_obj = datetime.strptime(date_str, '%d/%m/%Y').date()
                                
                                # Check date range if provided
                                if 'from_date' in st.session_state and 'to_date' in st.session_state:
                                    from_date = st.session_state.from_date.date() if isinstance(st.session_state.from_date, datetime) else st.session_state.from_date
                                    to_date = st.session_state.to_date.date() if isinstance(st.session_state.to_date, datetime) else st.session_state.to_date
                                    
                                    # Use inclusive range comparison
                                    if not (from_date <= date_obj <= to_date):
                                        continue
                                
                                # Create unique key for date-period combination
                                date_period_key = f"{date_str}_{period}"
                                
                                # Initialize or update combined class tracking
                                if date_period_key not in combined_classes:
                                    combined_classes[date_period_key] = {
                                        'sections': [],
                                        'time': time_str,
                                        'subject': subject,
                                        'lesson_plan': lesson_plan
                                    }
                                
                                # Add section to the combination
                                combined_classes[date_period_key]['sections'].append(section)
                                
                        except Exception as e:
                            st.error(f"Error processing entry: {entry}")
                            continue
        
        # Process combined classes into workload data
        for date_period_key, class_info in combined_classes.items():
            date_str, period = date_period_key.split('_')
            num_sections = len(class_info['sections'])
            
            # Create one entry per section with distributed workload
            for section in class_info['sections']:
                entry_data = {
                    'Date': date_str,
                    'Time': class_info['time'],
                    'Period': period,
                    'Subject': class_info['subject'],
                    'Section': section,
                    'Combined Sections': ', '.join(class_info['sections']),
                    'Workload': 1/num_sections  # Distribute workload among sections
                }
                if include_lesson_plans:
                    entry_data['Lesson Plan'] = class_info['lesson_plan']
                workload_data.append(entry_data)
        
        if workload_data:
            # Convert to DataFrame and add Month column
            df_workload = pd.DataFrame(workload_data)
            df_workload['DateObj'] = pd.to_datetime(df_workload['Date'], format='%d/%m/%Y')
            df_workload['Month'] = df_workload['DateObj'].dt.strftime('%b%Y')
            
            # Calculate total workload (counting combined sections as one class)
            total_workload = len(combined_classes)  # Each unique date-period combination counts as one class
            
            # Sort by date and remove DateObj column
            df_workload = df_workload.sort_values('DateObj', ascending=False)
            df_workload = df_workload.drop('DateObj', axis=1)
            
            return total_workload, df_workload
            
        return 0, pd.DataFrame()
        
    except Exception as e:
        st.error(f"Error calculating workload: {str(e)}")
        return 0, pd.DataFrame()


def get_faculty_id(faculty_name):
    """Extract RVIT ID from faculty name"""
    if '(' in faculty_name and ')' in faculty_name:
        return faculty_name[faculty_name.index('(')+1:faculty_name.index(')')]
    return None





def workload_analysis_page():
    """Page for viewing faculty workload with unified sheet structure and accurate workload distribution"""
    st.subheader("My Workload Analysis")
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.from_date = st.date_input(
            "From Date",
            datetime.now().replace(day=1),
            format="YYYY/MM/DD"
        )
    with col2:
        st.session_state.to_date = st.date_input(
            "To Date",
            datetime.now(),
            format="YYYY/MM/DD"
        )
    
    # Course filter
    courses = get_courses(for_attendance=True)
    selected_course = st.selectbox("Select Course", options=['All'] + courses)
    
    # Get workload data
    total_workload, workload_df = get_faculty_workload(st.session_state.username)
    
    if not workload_df.empty:
        # Filter by selected course if not 'All'
        if selected_course != 'All':
            workload_df = workload_df[workload_df['Section'].str.startswith(selected_course)]
        
        if not workload_df.empty:
            # Summary metrics using actual workload
            col1, col2, col3 = st.columns(3)
            with col1:
                total_actual_classes = workload_df['Workload'].sum()
                st.metric("Total Classes", f"{total_actual_classes:.1f}")
            with col2:
                unique_days = workload_df['Date'].nunique()
                st.metric("Days Engaged", unique_days)
            with col3:
                avg_classes = total_actual_classes / max(unique_days, 1)
                st.metric("Daily Average", f"{avg_classes:.1f}")
            
            # Show distribution with adjusted workload
            st.subheader("Teaching Distribution")
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("##### Subject-wise Classes")
                # Aggregate workload by subject
                subject_counts = workload_df.groupby('Subject')['Workload'].sum().reset_index()
                subject_counts.columns = ['Subject', 'Classes']
                subject_counts = subject_counts.sort_values('Classes', ascending=False)
                # Round to 1 decimal place
                subject_counts['Classes'] = subject_counts['Classes'].round(1)
                st.dataframe(subject_counts, hide_index=True)
                
            with col2:
                st.write("##### Section-wise Classes")
                # Aggregate workload by section
                section_counts = workload_df.groupby('Section')['Workload'].sum().reset_index()
                section_counts.columns = ['Section', 'Classes']
                section_counts = section_counts.sort_values('Classes', ascending=False)
                # Round to 1 decimal place
                section_counts['Classes'] = section_counts['Classes'].round(1)
                st.dataframe(section_counts, hide_index=True)
            
            # Detailed records grouped by month with combined sections info
            st.subheader("Detailed Class Records")
            workload_df['Month'] = pd.to_datetime(workload_df['Date'], format='%d/%m/%Y').dt.strftime('%B %Y')
            
            for month in sorted(workload_df['Month'].unique(), reverse=True):
                with st.expander(f"### {month}"):
                    month_data = workload_df[workload_df['Month'] == month].copy()
                    month_data = month_data.drop('Month', axis=1)
                    st.dataframe(
                        month_data,
                        column_config={
                            'Date': st.column_config.TextColumn('Date', width=100),
                            'Time': st.column_config.TextColumn('Time', width=100),
                            'Period': st.column_config.TextColumn('Period', width=80),
                            'Section': st.column_config.TextColumn('Section', width=150),
                            'Subject': st.column_config.TextColumn('Subject', width=150),
                            'Combined Sections': st.column_config.TextColumn('Combined Sections', width=200),
                            'Workload': st.column_config.NumberColumn('Workload', format="%.1f", width=100),
                            'Lesson Plan': st.column_config.TextColumn('Lesson Plan', width=300)
                        },
                        hide_index=True,
                        use_container_width=True
                    )
                    
                    # Download option for monthly data
                    csv = month_data.to_csv(index=False)
                    st.download_button(
                        label=f"Download {month} Records",
                        data=csv,
                        file_name=f"workload_{month}_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv"
                    )
        else:
            st.info(f"No classes found for {selected_course}")
    else:
        st.info("No classes recorded in the selected date range")


def reset_password():
    """Function to handle password reset with improved data handling"""
    st.subheader("Reset Password")
    
    current_password = st.text_input("Current Password", type="password", key="current_pwd")
    new_password = st.text_input("New Password", type="password", key="new_pwd")
    confirm_password = st.text_input("Confirm New Password", type="password", key="confirm_pwd")
    
    if st.button("Update Password", type="primary"):
        try:
            if not all([current_password, new_password, confirm_password]):
                st.error("All fields are required")
                return
                
            # Read faculty data
            df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
            
            # Convert credentials columns to string and strip whitespace
            df_faculty['Username'] = df_faculty['Username'].astype(str).str.strip()
            df_faculty['Password'] = df_faculty['Password'].astype(str).str.strip()
            
            # Check current credentials
            user_mask = (df_faculty['Username'] == st.session_state.username) & \
                       (df_faculty['Password'] == current_password)
                       
            if not user_mask.any():
                st.error("Current password is incorrect")
                return
            
            # Verify new passwords match
            if new_password != confirm_password:
                st.error("New passwords do not match")
                return
            
            # Update password
            df_faculty.loc[user_mask, 'Password'] = new_password
            
            # Save changes
            with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                df_faculty.to_excel(writer, sheet_name='Faculty', index=False)
                
                # Format worksheet
                worksheet = writer.sheets['Faculty']
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or '')) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
            
            st.success("Password updated successfully! Please login again.")
            
            # Clear session state to force re-login
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
            
        except Exception as e:
            st.error(f"Error updating password: {str(e)}")

def get_last_class_attendance(section, period):
    """
    Get attendance pattern from the last class period for comparison, regardless of subject or faculty
    Args:
        section: The section to check
        period: Current period (e.g. 'P1', 'P2', etc.)
    Returns a dictionary of {HT Number: attendance_status}
    """
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name='Students')
        df_filtered = df[df['Merged Section'] == section]
        
        if df_filtered.empty:
            return None
            
        # Get current period's numeric value
        current_period_num = int(period[1])
        
        # Get previous period
        prev_period = f'P{current_period_num - 1}' if current_period_num > 1 else None
        
        if not prev_period:
            return None
            
        # Get attendance pattern from previous period
        latest_pattern = {}
        last_faculty = None
        last_subject = None
        
        for _, student_row in df_filtered.iterrows():
            if pd.notna(student_row[prev_period]):
                entries = str(student_row[prev_period]).split('\n')
                # Get the latest entry for this period (last line)
                if entries and entries[-1].strip():
                    try:
                        parts = entries[-1].split('_')
                        if len(parts) >= 6:
                            faculty = parts[3]
                            subject = parts[4]
                            # Get status regardless of subject/faculty
                            status = 'P' if '_P_' in entries[-1] else 'A'
                            latest_pattern[str(student_row['HT Number'])] = {
                                'status': status,
                                'faculty': faculty,
                                'subject': subject
                            }
                            last_faculty = faculty
                            last_subject = subject
                    except:
                        continue
        
        if latest_pattern:
            # Add faculty and subject information to the pattern
            latest_pattern['prev_faculty'] = last_faculty
            latest_pattern['prev_subject'] = last_subject
            
        return latest_pattern if latest_pattern else None
        
    except Exception as e:
        st.error(f"Error getting last class attendance: {str(e)}")
        return None

def mark_attendance_page():
    """Enhanced attendance marking page without constant refreshes"""
    section = st.session_state.sections[0] if st.session_state.sections else None
    subject = st.session_state.subject  
    period = st.session_state.period

    if section and period:
        # Check for duplicate attendance
        current_date = datetime.now().strftime('%d/%m/%Y')
        is_duplicate, existing_faculty = check_duplicate_attendance(section, period, current_date)
        
        if is_duplicate:
            if existing_faculty:
                st.error(f"⚠️ Attendance for this section and period has already been marked by {existing_faculty}")
            else:
                st.error("⚠️ Attendance for this section and period has already been marked")
            return

    # Only continue if there's no duplicate and we have all required fields
    if section and subject and period:
        # Session info card
        st.markdown(f"""
            <div style='background: linear-gradient(135deg, #6B46C1 0%, #805AD5 100%);
                      color: white;
                      padding: 1.2rem;
                      border-radius: 15px;
                      margin-bottom: 1.5rem;
                      box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);'>
                <h3 style='margin: 0; font-size: 1.2rem; font-weight: 600;'>Current Session</h3>
                <div style='margin-top: 1rem;'>
                    <div style='display: flex; align-items: center; margin: 0.5rem 0;'>
                        <span style='width: 24px; text-align: center; margin-right: 8px;'>📚</span>
                        <span style='font-size: 1rem;'>{section}</span>
                    </div>
                    <div style='display: flex; align-items: center; margin: 0.5rem 0;'>
                        <span style='width: 24px; text-align: center; margin-right: 8px;'>📖</span>
                        <span style='font-size: 1rem;'>{subject}</span>
                    </div>
                    <div style='display: flex; align-items: center; margin: 0.5rem 0;'>
                        <span style='width: 24px; text-align: center; margin-right: 8px;'>⏰</span>
                        <span style='font-size: 1rem;'>Period {period}</span>
                    </div>
                </div>
            </div>
        """, unsafe_allow_html=True)

        # Get students for this section
        df_students = get_student_data(section, for_attendance=True)
        if df_students is not None:
            # Initialize attendance data dictionary if not exists
            if 'attendance_data' not in st.session_state:
                st.session_state.attendance_data = {}
                
            # Initialize select_all in session state if not present
            if 'select_all' not in st.session_state:
                st.session_state.select_all = True
            
            # Use form to prevent constant refreshes
            with st.form(key='attendance_form'):
                # Initialize data structures
                current_pattern = {}
                attendance_data = {}
                
                # Quick action buttons in three columns
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    mark_all_present = st.form_submit_button("✓ Mark All Present", type="primary")
                with col2:
                    mark_all_absent = st.form_submit_button("✗ Mark All Absent")
                with col3:
                    use_last_pattern = st.form_submit_button("↺ Use Last Pattern")

                # Update status based on form actions
                if mark_all_present:
                    st.session_state.select_all = True
                    for _, student in df_students.iterrows():
                        st.session_state[f"attendance_{student['HT Number']}"] = True
                elif mark_all_absent:
                    st.session_state.select_all = False
                    for _, student in df_students.iterrows():
                        st.session_state[f"attendance_{student['HT Number']}"] = False
                elif use_last_pattern:
                    last_pattern = get_last_class_attendance(section, period)
                    if last_pattern:
                        st.session_state.last_pattern = last_pattern
                        for _, student in df_students.iterrows():
                            ht_number = str(student['HT Number'])
                            st.session_state[f"attendance_{ht_number}"] = last_pattern.get(ht_number, {}).get('status', 'P') == 'P'
                    else:
                        st.warning("No previous attendance pattern found")

                # Student list with clean cards
                for _, student in df_students.iterrows():
                    ht_number = str(student['HT Number'])
                    
                    # Get default value from session state or use select_all
                    default_value = st.session_state.get(f"attendance_{ht_number}", st.session_state.select_all)

                    with st.container():
                        col1, col2 = st.columns([7,3])
                        
                        with col1:
                            st.markdown(f"""
                                <div class="student-info" style="background: #1E1E1E; padding: 0.5rem; border-radius: 8px;">
                                    <div style="font-size: 1rem; font-weight: 500; color: #FF0099; margin-bottom: 0.2rem;">
                                        {student['Student Name']}
                                    </div>
                                    <div style="font-size: 1rem; color: #FF9900; margin-bottom: 0.2rem;">
                                        {ht_number}
                                    </div>
                                    <div style="font-size: 0.8rem; color: #888;">
                                        {student['Original Section']}
                                    </div>
                                </div>
                            """, unsafe_allow_html=True)
                        
                        with col2:
                            # Use form checkbox to prevent refresh
                            status = st.checkbox(
                                "Present",
                                value=default_value,
                                key=f"form_attendance_{ht_number}"
                            )
                            
                        current_pattern[ht_number] = 'P' if status else 'A'
                        attendance_data[ht_number] = {
                            'status': 'P' if status else 'A',
                            'original_section': student['Original Section']
                        }
                    
                    st.markdown("<hr style='margin: 0.5rem 0; border: none; border-top: 1px solid #333;'>", unsafe_allow_html=True)

                # Add lesson plan input
                lesson_plan = st.text_area(
                    "Enter Lesson Plan/Topic Name (Required)",
                    help="Please enter topic covered in this class",
                    height=100,
                    key="lesson_plan"
                )
                
                # Submit button at the bottom of the form
                submit_button = st.form_submit_button("📝 Submit Attendance", type="primary", use_container_width=True)
                
            # Handle form submission outside the form
            if submit_button:
                if not lesson_plan.strip():
                    st.error("⚠️ Please enter a lesson plan before submitting attendance")
                else:
                    success, unsuccessful_records = mark_attendance(
                        section, period, attendance_data,
                        st.session_state.username, subject,
                        lesson_plan
                    )
                    
                    if unsuccessful_records:
                        st.info(f"✅ Recorded {len(attendance_data) - len(unsuccessful_records)} students")
                        st.warning("⚠️ Issues found:")
                        for record in unsuccessful_records:
                            st.markdown(f"""
                                <div style='background: #FFF3CD;
                                          padding: 1rem;
                                          border-radius: 8px;
                                          margin: 0.5rem 0;
                                          border: 1px solid #FFE69C;'>
                                    <div style='font-weight: 500; color: #664D03;'>
                                        {record['Student Name']} ({record['HT Number']})
                                    </div>
                                    <div style='color: #997404; font-size: 0.9rem; margin-top: 0.3rem;'>
                                        {record['Reason']}
                                    </div>
                                </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.success(f"✅ Successfully recorded all {len(attendance_data)} students")
                        
                        # Compare with previous period's pattern
                        if period != 'P1':
                            last_pattern = get_last_class_attendance(section, period)
                            if last_pattern and 'prev_faculty' in last_pattern:
                                prev_faculty = last_pattern.pop('prev_faculty')
                                prev_subject = last_pattern.pop('prev_subject')
                                conflicts = []
                                for ht_number, current_status in current_pattern.items():
                                    if ht_number in last_pattern:
                                        last_status = last_pattern[ht_number]['status']
                                        last_faculty = last_pattern[ht_number]['faculty']
                                        last_subject = last_pattern[ht_number]['subject']
                                        if last_status != current_status:
                                            student = df_students[df_students['HT Number'] == ht_number].iloc[0]
                                            conflicts.append({
                                                'name': student['Student Name'],
                                                'ht_number': ht_number,
                                                'last_status': last_status,
                                                'current_status': current_status,
                                                'prev_period': f'P{int(period[1])-1}',
                                                'prev_faculty': last_faculty,
                                                'prev_subject': last_subject,
                                                'current_subject': subject
                                            })
                                
                                if conflicts:
                                    st.warning("🔍 Attendance Pattern Changes Detected:")
                                    present_to_absent = []
                                    absent_to_present = []
                                    
                                    for conflict in conflicts:
                                        if conflict['last_status'] == 'P' and conflict['current_status'] == 'A':
                                            present_to_absent.append(conflict)
                                        elif conflict['last_status'] == 'A' and conflict['current_status'] == 'P':
                                            absent_to_present.append(conflict)
                                    
                                    if present_to_absent:
                                        st.markdown("##### Present ➡️ Absent:")
                                        for change in present_to_absent:
                                            st.markdown(f"""
                                                <div style='background: #FFE8E8;
                                                          padding: 0.8rem;
                                                          border-radius: 8px;
                                                          margin: 0.5rem 0;
                                                          border: 1px solid #FFB4B4;'>
                                                    <div style='font-weight: 500; color: #CC0000;'>
                                                        {change['name']} ({change['ht_number']})
                                                    </div>
                                                    <div style='color: #990000; font-size: 0.9rem; margin-top: 0.3rem;'>
                                                        Previous Period ({change['prev_period']} - {change['prev_subject']} by {change['prev_faculty']}): Present
                                                        → Current ({period} - {change['current_subject']}): Absent
                                                    </div>
                                                </div>
                                            """, unsafe_allow_html=True)
                                    
                                    if absent_to_present:
                                        st.markdown("##### Absent ➡️ Present:")
                                        for change in absent_to_present:
                                            st.markdown(f"""
                                                <div style='background: #E8FFE8;
                                                          padding: 0.8rem;
                                                          border-radius: 8px;
                                                          margin: 0.5rem 0;
                                                          border: 1px solid #B4FFB4;'>
                                                    <div style='font-weight: 500; color: #006600;'>
                                                        {change['name']} ({change['ht_number']})
                                                    </div>
                                                    <div style='color: #004D00; font-size: 0.9rem; margin-top: 0.3rem;'>
                                                        Previous Period ({change['prev_period']} - {change['prev_subject']} by {change['prev_faculty']}): Absent
                                                        → Current ({period} - {change['current_subject']}): Present
                                                    </div>
                                                </div>
                                            """, unsafe_allow_html=True)
                                    
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.metric("Present → Absent", len(present_to_absent))
                                    with col2:
                                        st.metric("Absent → Present", len(absent_to_present))

                    # Clear form data and rerun only after successful submission
                    if success:
                        st.session_state.attendance_data = {}
                        st.rerun()


def reset_username():
    """Function to handle username reset with improved data handling"""
    st.subheader("Reset Username")
    
    current_username = st.text_input("Current Username", key="reset_user_current")
    password = st.text_input("Current Password", type="password", key="reset_user_pwd")
    new_username = st.text_input("New Username", key="reset_user_new")
    
    if st.button("Reset Username", key="reset_user_button", type="primary"):
        try:
            if not all([current_username, password, new_username]):
                st.error("All fields are required")
                return
                
            # Read faculty data
            df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
            
            # Convert credentials columns to string and strip whitespace
            df_faculty['Username'] = df_faculty['Username'].astype(str).str.strip()
            df_faculty['Password'] = df_faculty['Password'].astype(str).str.strip()
            
            # Clean input credentials
            current_username = str(current_username).strip()
            password = str(password).strip()
            new_username = str(new_username).strip()
            
            # Verify credentials
            user_mask = (df_faculty['Username'] == current_username) & \
                       (df_faculty['Password'] == password)
            if not user_mask.any():
                st.error("Invalid credentials")
                return
                
            # Check if new username already exists
            if (df_faculty['Username'] == new_username).any():
                st.error("Username already exists")
                return
                
            # Update username
            df_faculty.loc[user_mask, 'Username'] = new_username
            
            # Save changes while preserving all columns
            with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                df_faculty.to_excel(writer, sheet_name='Faculty', index=False)
                
                # Format worksheet
                worksheet = writer.sheets['Faculty']
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Set column widths
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or '')) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
            
            st.success("Username updated successfully! Please login again with your new username.")
            
            # Clear session state to force re-login
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
            
        except Exception as e:
            st.error(f"Error resetting username: {str(e)}")



def show_class_schedule_page():
    """Display class schedule and missing attendance analysis"""
    st.subheader("Class Schedule")
    
    # Classes Section with Date Filters
    col1, col2 = st.columns(2)
    with col1:
        selected_date = st.date_input(
            "Select Date",
            datetime.now(),
            key="class_schedule_date"
        )
    with col2:
        view_type = st.selectbox(
            "View Type",
            ["Single Day", "Date Range"],
            key="class_view_type"
        )
    
    if view_type == "Date Range":
        end_date = st.date_input(
            "End Date",
            selected_date,
            key="class_schedule_end_date"
        )
    else:
        end_date = selected_date
    
    try:
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        classes_data = []
        
        # Convert dates to string format for comparison
        start_date_str = selected_date.strftime('%d/%m/%Y')
        end_date_str = end_date.strftime('%d/%m/%Y')
        
        for _, faculty_row in df_faculty.iterrows():
            faculty_name = faculty_row['Faculty Name']
            for col in faculty_row.index:
                if col not in ['Faculty Name', 'Username', 'Password']:
                    entries = str(faculty_row[col]).split('\n') if pd.notna(faculty_row[col]) else []
                    for entry in entries:
                        if pd.notna(entry) and entry.strip():
                            try:
                                parts = entry.strip().split('_')
                                if len(parts) >= 6:
                                    entry_date = parts[0]
                                    # Check if date is within selected range
                                    if start_date_str <= entry_date <= end_date_str:
                                        classes_data.append({
                                            'Date': entry_date,
                                            'Time': parts[1],
                                            'Period': parts[2],
                                            'Faculty': faculty_name,
                                            'Subject': parts[3],
                                            'Section': parts[4],
                                            'Status': 'Completed'
                                        })
                            except Exception:
                                continue
        
        if classes_data:
            df_classes = pd.DataFrame(classes_data)
            df_classes = df_classes.sort_values(['Date', 'Time', 'Period'])
            
            if view_type == "Date Range":
                st.write(f"### Classes from {start_date_str} to {end_date_str}")
            else:
                st.write(f"### Classes on {start_date_str}")
                
            # Add summary metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Classes", len(df_classes))
            with col2:
                st.metric("Faculty Engaged", df_classes['Faculty'].nunique())
            with col3:
                st.metric("Sections Covered", df_classes['Section'].nunique())
            
            st.dataframe(
                df_classes,
                column_config={
                    'Date': st.column_config.TextColumn('Date', width=100),
                    'Time': st.column_config.TextColumn('Time', width=80),
                    'Period': st.column_config.TextColumn('Period', width=70),
                    'Faculty': st.column_config.TextColumn('Faculty', width=150),
                    'Subject': st.column_config.TextColumn('Subject', width=120),
                    'Section': st.column_config.TextColumn('Section', width=120),
                    'Status': st.column_config.TextColumn(
                        'Status',
                        width=100,
                        help="Shows if attendance has been marked"
                    )
                },
                hide_index=True,
                use_container_width=True
            )
            
            # Add Missing Attendance Analysis section
            st.write("### Missing Attendance Analysis")
            missing_data, all_sections = analyze_missing_attendance(start_date_str, df_classes)
            
            if missing_data:
                st.warning("⚠️ The following sections have not had attendance marked:")
                
                for period_data in missing_data:
                    with st.expander(f"Period {period_data['Period']}"):
                        # Create metrics for the period
                        total_missing = len(period_data['Missing Sections'])
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Missing Sections", total_missing)
                        
                        with col2:
                            coverage = ((len(all_sections) - total_missing) / len(all_sections)) * 100
                            st.metric("Attendance Coverage", f"{coverage:.1f}%")
                        
                        # Display missing sections in a organized way
                        st.write("##### Sections without Attendance:")
                        for section in period_data['Missing Sections']:
                            st.markdown(f"""
                                <div style='background: #FF575720; 
                                          padding: 8px 15px; 
                                          border-radius: 5px; 
                                          border-left: 4px solid #FF5757;
                                          margin: 5px 0;'>
                                    {section}
                                </div>
                            """, unsafe_allow_html=True)
            else:
                st.success("✅ All sections have attendance marked for all periods!")
                
        else:
            if view_type == "Date Range":
                st.info(f"No classes recorded between {start_date_str} and {end_date_str}")
            else:
                st.info(f"No classes recorded on {start_date_str}")
            
    except Exception as e:
        st.error(f"Error loading class schedule: {str(e)}")

def faculty_page():
    """Updated faculty page with course filtering and enhanced download options"""
    faculty_name = st.session_state.faculty_name
    
    st.title(f"Welcome, {faculty_name}")
    
    with st.sidebar:
        st.header("Navigation")
        page = st.radio(
            "Select", 
            ["Mark Attendance", "Class Schedule", "View Statistics", "Student Reports", 
             "Subject Analysis", "My Workload", "Reset Credentials"]
        )

    if page == "Reset Credentials":
        reset_password()
        return

    elif page == "Mark Attendance":
        st.subheader("Select Class Details")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            period = st.selectbox(
                "Select Period",
                options=[''] + ['P1', 'P2', 'P3', 'P4', 'P5', 'P6'],
                key="period_select"
            )
        
        with col2:
            courses = get_courses(for_attendance=True)
            selected_course = st.selectbox(
                "Select Course",
                options=[''] + courses,
                key="course_select"
            )
        
        with col3:
            filtered_sections = get_sections_by_course(selected_course, for_attendance=True) if selected_course else []
            selected_section = st.selectbox(
                "Select Section",
                options=[''] + filtered_sections,
                key="section_select"
            )

        if selected_section and period:
            with col4:
                subjects = get_section_subjects(selected_section, for_subject_analysis=True)
                unique_subjects = sorted(list(set(subjects)))
                subject = st.selectbox(
                    "Select Subject",
                    options=[''] + unique_subjects,
                    key="subject_select"
                )

            if subject:
                current_date = datetime.now().strftime('%d/%m/%Y')
                is_duplicate, existing_faculty = check_duplicate_attendance(selected_section, period, current_date)
                
                if is_duplicate:
                    if existing_faculty:
                        st.error(f"⚠️ Attendance for this section and period has already been marked by {existing_faculty}")
                        return
                    else:
                        st.error("⚠️ Attendance for this section and period has already been marked")
                        return

                st.session_state.period = period
                st.session_state.sections = [selected_section] if selected_section else []
                st.session_state.subject = subject

                mark_attendance_page()
            else:
                st.info("Please select subject to continue")
        else:
            st.info("Please select period and section")

    elif page == "Class Schedule":
        show_class_schedule_page()
            
    elif page == "View Statistics":
        st.subheader("View Attendance Statistics")
        
        courses = get_courses(for_attendance=False)
        selected_course = st.selectbox("Select Course", options=[''] + courses)
        
        if selected_course:
            sections = get_sections_by_course(selected_course, for_attendance=False)
            selected_sections = st.multiselect("Select Sections", options=sections)
            
            if selected_sections:
                col1, col2 = st.columns(2)
                with col1:
                    from_date = st.date_input("From Date")
                with col2:
                    to_date = st.date_input("To Date")
                
                all_stats = []
                for section in selected_sections:
                    stats_df = get_attendance_stats(section, from_date, to_date)
                    if stats_df is not None and not stats_df.empty:
                        stats_df['Section'] = section
                        all_stats.append(stats_df)
                
                if all_stats:
                    combined_stats = pd.concat(all_stats, ignore_index=True)
                    
                    st.write("### Overall Statistics")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Total Students", len(combined_stats))
                    with col2:
                        avg_attendance = combined_stats['Overall %'].mean()
                        st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                    with col3:
                        below_75 = len(combined_stats[combined_stats['Overall %'] < 75])
                        st.metric("Students Below 75%", below_75)
                    
                    # Configure column display
                    column_config = {
                        'HT Number': st.column_config.TextColumn('HT Number', width=120),
                        'Student Name': st.column_config.TextColumn('Student Name', width=180),
                        'Section': st.column_config.TextColumn('Section', width=150),
                        'Overall %': st.column_config.NumberColumn(
                            'Overall %',
                            format="%.2f%%",
                            width=100
                        )
                    }
                    
                    for col in combined_stats.columns:
                        if 'Attended/Conducted' in col:
                            new_col = col.replace('Attended/Conducted', 'A/C')
                            column_config[col] = st.column_config.TextColumn(
                                new_col,
                                width=150
                            )
                    
                    st.write("### Student-wise Statistics")
                    st.dataframe(
                        combined_stats,
                        column_config=column_config,
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Download options
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("📊 Download Excel"):
                            buffer = io.BytesIO()
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                combined_stats.to_excel(writer, sheet_name='Attendance_Stats', index=False)
                                
                                worksheet = writer.sheets['Attendance_Stats']
                                
                                # Add legend at the top
                                worksheet.insert_rows(0)
                                legend_cell = worksheet.cell(row=1, column=1)
                                legend_cell.value = "A: Attended, C: Conducted"
                                legend_cell.font = Font(italic=True)
                                
                                # Format cells
                                for column in worksheet.columns:
                                    max_length = max(len(str(cell.value or '')) for cell in column)
                                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                                    
                                thin_border = Border(left=Side(style='thin'), 
                                                   right=Side(style='thin'), 
                                                   top=Side(style='thin'), 
                                                   bottom=Side(style='thin'))
                                                   
                                for row in worksheet.iter_rows(min_row=2):
                                    for cell in row:
                                        cell.border = thin_border
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            st.download_button(
                                label="📥 Download Excel File",
                                data=buffer.getvalue(),
                                file_name=f"attendance_stats_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                    with col2:
                        if st.button("📄 Download PDF"):
                            try:
                                # Prepare data for PDF
                                pdf_df = combined_stats.copy()
                                
                                # Replace column headers with A/C notation
                                pdf_df.columns = [col.replace('Attended/Conducted', 'A/C') for col in pdf_df.columns]
                                
                                filters = {
                                    "Course": selected_course,
                                    "Sections": ", ".join(selected_sections),
                                    "Date Range": f"{from_date} to {to_date}"
                                }
                                
                                pdf_data = generate_statistics_pdf(
                                    df=pdf_df,
                                    title="Attendance Statistics Report",
                                    filters=filters
                                )
                                
                                if pdf_data:
                                    st.download_button(
                                        label="📥 Download PDF Report",
                                        data=pdf_data,
                                        file_name=f"attendance_stats_{datetime.now().strftime('%Y%m%d')}.pdf",
                                        mime="application/pdf"
                                    )
                                else:
                                    st.error("Error generating PDF report")
                            except Exception as e:
                                st.error(f"Error generating PDF report: {str(e)}")
                else:
                    st.info("No attendance records found for the selected criteria")

    elif page == "Student Reports":
        student_reports_page()

    elif page == "Subject Analysis":
        subject_analysis_page()

    elif page == "My Workload":
        workload_analysis_page()

    # Logout button
    with st.sidebar:
        st.markdown("<br>" * 5, unsafe_allow_html=True)
        if st.button("Logout", key="logout_button", type="primary", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

def check_login(username, password, is_admin=False):
    """Verify login credentials with improved data handling"""
    try:
        # Read faculty data with string type conversion
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        
        # Convert credentials columns to string and strip whitespace
        df_faculty['Username'] = df_faculty['Username'].astype(str).str.strip()
        df_faculty['Password'] = df_faculty['Password'].astype(str).str.strip()
        df_faculty['Faculty Name'] = df_faculty['Faculty Name'].astype(str).str.strip()
        
        # Convert input credentials to string and strip whitespace
        username = str(username).strip()
        password = str(password).strip()
        
        # First verify basic credentials
        user_exists = any((df_faculty['Username'] == username) & 
                         (df_faculty['Password'] == password))
        
        if not user_exists:
            return False
            
        if is_admin:
            # For admin login, check if the user has admin in their faculty name
            faculty_name = df_faculty[df_faculty['Username'] == username]['Faculty Name'].iloc[0]
            return '(admin)' in faculty_name.lower()
        
        return True
                      
    except Exception as e:
        st.error(f"Login error: {str(e)}")
        return False

def get_current_time_ist():
    """Get current time in IST timezone"""
    ist = pytz.timezone('Asia/Kolkata')
    # Convert UTC to IST
    utc_now = datetime.now(timezone.utc)
    current_time = utc_now.astimezone(ist)
    return current_time

def mark_attendance(section, period, attendance_data, username, subject, lesson_plan):
    try:
        # Get current time in IST and ensure timezone awareness
        current_time = get_current_time_ist()
        # Format the time with IST
        date_str = current_time.strftime('%d/%m/%Y')
        time_str = current_time.strftime('%I:%M%p')
        if time_str.startswith('0'):
            time_str = time_str[1:]

        # Check for duplicate attendance
        exists, marked_by, marked_date = check_existing_attendance(section, period)
        if exists:
            return False, [{
                'HT Number': 'N/A',
                'Student Name': 'N/A',
                'Original Section': section,
                'Reason': f"Attendance for this section and period has already been marked by {marked_by} on {marked_date}. Multiple entries per period are not allowed."
            }]

        # Get faculty name
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty', dtype=str)
        user_row = df_faculty[df_faculty['Username'] == username].iloc[0]
        faculty_name = user_row['Faculty Name']
        
        unsuccessful_records = []
        
        # Process attendance data
        # Read with string dtype to avoid type conversion issues
        df_students = pd.read_excel('attendance.xlsx', sheet_name='Students', dtype=str)
        
        # Convert NaN values to empty strings
        df_students = df_students.fillna('')
        
        success = True
        
        # Update attendance in the unified Students sheet
        with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            for ht_number, status in attendance_data.items():
                try:
                    row_mask = df_students['HT Number'] == str(ht_number)
                    if not row_mask.any():
                        student_name = "Unknown"
                        orig_section = "Unknown"
                        unsuccessful_records.append({
                            'HT Number': ht_number,
                            'Student Name': student_name,
                            'Original Section': orig_section,
                            'Reason': "Student not found"
                        })
                        continue
                    
                    # Create attendance entry
                    attendance_value = f"{date_str}_{time_str}_{status['status']}_{faculty_name}_{subject}_{lesson_plan}"
                    
                    # Get current value and append new entry
                    current_value = df_students.loc[row_mask, period].iloc[0]
                    
                    # Set the new value, handling empty cells properly
                    if current_value and str(current_value).strip():
                        df_students.loc[row_mask, period] = str(current_value) + '\n' + attendance_value
                    else:
                        df_students.loc[row_mask, period] = attendance_value
                    
                except Exception as e:
                    unsuccessful_records.append({
                        'HT Number': ht_number,
                        'Student Name': "Unknown",
                        'Original Section': "Unknown",
                        'Reason': f"Error processing attendance: {str(e)}"
                    })
            
            # Save updated data
            df_students.to_excel(writer, sheet_name='Students', index=False)
            
            # Format worksheet
            worksheet = writer.sheets['Students']
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
        
        # Update faculty log if successful
        if success:
            update_faculty_log(faculty_name, section, period, subject, lesson_plan, time_str, date_str)
        
        return success, unsuccessful_records
    
    except Exception as e:
        st.error(f"Error marking attendance: {str(e)}")
        return False, []

def get_faculty_data(sheet):
    """Get faculty data with proper empty value handling"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
        
        # Convert all columns to string type and replace NaN with empty string
        for col in df.columns:
            df[col] = df[col].fillna('')  # Replace NaN with empty string
            df[col] = df[col].astype(str)  # Convert to string type
            df[col] = df[col].replace('nan', '')  # Replace any 'nan' strings with empty string
            
        return df
    except Exception as e:
        st.error(f"Error getting faculty data: {str(e)}")
        return None

def update_faculty_log(faculty_name, section, period, subject, lesson_plan, time_str=None, date_str=None):
    """Update faculty log with improved empty value handling"""
    try:
        # Get current time in IST with proper timezone handling
        current_time = get_current_time_ist()
        
        # Ensure we're working with timezone-aware datetime
        if not current_time.tzinfo:
            ist = pytz.timezone('Asia/Kolkata')
            current_time = ist.localize(current_time)
        
        # Read faculty sheet with improved empty value handling
        df = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        
        # Replace NaN with empty strings
        df = df.fillna('')
        
        # Convert all columns to string type
        for col in df.columns:
            df[col] = df[col].astype(str)
            df[col] = df[col].replace('nan', '')
        
        # Get current month-year for column
        month_year = current_time.strftime('%b%Y')
        
        # Use provided time and date if available, otherwise generate new ones
        if time_str is None or date_str is None:
            date_str = current_time.strftime('%d/%m/%Y')
            time_str = current_time.strftime('%I:%M%p')
            if time_str.startswith('0'):
                time_str = time_str[1:]
        
        # Create log entry
        log_entry = f"{date_str}_{time_str}_{period}_{subject}_{section}_{lesson_plan}"
        
        # Check if month-year column exists, if not create it
        if month_year not in df.columns:
            # Get existing columns
            existing_cols = list(df.columns)
            # Find the position after 'Password' column
            password_idx = existing_cols.index('Password')
            # Insert new column after Password with empty string
            df.insert(password_idx + 1, month_year, '')
        
        # Update the log for the faculty
        faculty_mask = df['Faculty Name'] == faculty_name
        if faculty_mask.any():
            current_log = df.loc[faculty_mask, month_year].iloc[0]
            # Add new entry with proper newline handling
            if current_log and str(current_log).strip() and str(current_log).strip() != 'nan':
                new_log = f"{current_log}\n{log_entry}"
            else:
                new_log = log_entry
                
            # Update with new log entry
            df.loc[faculty_mask, month_year] = new_log
        
        # Save the updated sheet
        with pd.ExcelWriter('attendance.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name='Faculty', index=False)
            
            # Format worksheet
            worksheet = writer.sheets['Faculty']
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
        
        return True
    except Exception as e:
        st.error(f"Error updating faculty log: {str(e)}")
        return False



def create_template_df(sheet_name):
    """Create template DataFrame with updated structure and course-aware templates"""
    if sheet_name == 'Faculty':
        # Get current month-year
        current_month = datetime.now().strftime('%b%Y')
        return pd.DataFrame(columns=['Faculty Name', 'Username', 'Password', current_month])
    elif sheet_name == 'Section-Subject-Mapping':
        return pd.DataFrame(columns=['Section', 'Subject Names'])
    elif sheet_name == 'Students':
        return pd.DataFrame(columns=[
            'HT Number', 'Student Name', 'Original Section', 'Merged Section',
            'P1', 'P2', 'P3', 'P4', 'P5', 'P6'
        ])
    else:
        st.error(f"Unknown sheet type: {sheet_name}")
        return pd.DataFrame()

def validate_upload_data(df, sheet_name):
    """Validate uploaded data against expected format with course validation"""
    try:
        template_df = create_template_df(sheet_name)
        
        # Basic column validation
        if not all(col in df.columns for col in template_df.columns):
            return False
            
        if sheet_name == 'Students':
            # Validate section formats
            if not all(pd.notna(df['Original Section'])):
                st.error("Original Section cannot be empty")
                return False
                
            if not all(pd.notna(df['Merged Section'])):
                st.error("Merged Section cannot be empty")
                return False
                
            # Validate HT Numbers
            if not all(pd.notna(df['HT Number'])):
                st.error("HT Number cannot be empty")
                return False
                
            # Validate Student Names
            if not all(pd.notna(df['Student Name'])):
                st.error("Student Name cannot be empty")
                return False
                
        elif sheet_name == 'Section-Subject-Mapping':
            # Validate section and subject names
            if not all(pd.notna(df['Section'])):
                st.error("Section cannot be empty")
                return False
                
            if not all(pd.notna(df['Subject Names'])):
                st.error("Subject Names cannot be empty")
                return False
                
        elif sheet_name == 'Faculty':
            # Validate faculty data
            if not all(pd.notna(df['Faculty Name'])):
                st.error("Faculty Name cannot be empty")
                return False
                
            if not all(pd.notna(df['Username'])):
                st.error("Username cannot be empty")
                return False
                
            if not all(pd.notna(df['Password'])):
                st.error("Password cannot be empty")
                return False
        
        return True
        
    except Exception as e:
        st.error(f"Error validating data: {str(e)}")
        return False

def initialize_excel():
    """Initialize Excel file with improved error handling"""
    try:
        # Check if file exists and is valid
        if os.path.exists('attendance.xlsx'):
            if verify_excel_file():
                return True
            else:
                st.warning("Existing Excel file is corrupted. Creating new file...")
                # Backup corrupted file
                if os.path.exists('attendance.xlsx'):
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    shutil.move('attendance.xlsx', f'attendance_corrupted_{timestamp}.xlsx')

        # Create new Excel file with required sheets
        with pd.ExcelWriter('attendance.xlsx', engine='openpyxl') as writer:
            # Create Students sheet with updated structure
            students_df = pd.DataFrame(columns=[
                'HT Number', 'Student Name', 'Original Section', 'Merged Section',
                'P1', 'P2', 'P3', 'P4', 'P5', 'P6'
            ])
            students_df.to_excel(writer, sheet_name='Students', index=False)

            # Create Faculty sheet with current month
            current_month = datetime.now().strftime('%b%Y')
            faculty_df = pd.DataFrame(columns=['Faculty Name', 'Username', 'Password', current_month])
            faculty_df.to_excel(writer, sheet_name='Faculty', index=False)

            # Create Section-Subject-Mapping sheet
            mapping_df = pd.DataFrame(columns=['Section', 'Subject Names'])
            mapping_df.to_excel(writer, sheet_name='Section-Subject-Mapping', index=False)

            # Format all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or '')) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))

        return True
    except Exception as e:
        st.error(f"Error initializing Excel file: {str(e)}")
        return False

def main():
    """Main function with enhanced Excel file handling"""
    # Verify Excel file integrity
    if not verify_excel_file():
        st.error("Excel file is corrupted. Attempting to restore from backup...")
        # Try to restore from latest backup
        latest_backup = max(glob.glob('backup/attendance_*.xlsx'), default=None)
        if latest_backup:
            shutil.copy2(latest_backup, 'attendance.xlsx')
            st.success("Restored from backup")
        else:
            st.error("No backup available. Reinitializing file...")
            if not initialize_excel():
                st.error("Error initializing the application. Please check the error above.")
                return
    
    # Create backup before starting
    backup_excel_file()
    
    # Rest of your existing main function code...
    if 'logged_in' not in st.session_state:
        st.title("Login")
        
        # Single login interface
        login_type = st.radio("Select Login Type", ["Faculty", "Admin"], key="login_type")
        username = st.text_input("Username", key="login_username")
        password = st.text_input("Password", type="password", key="login_password")
        
        if st.button("Login", key="login_button", type="primary"):
            try:
                # Add file verification before reading
                if not verify_excel_file():
                    st.error("Database error. Please try again.")
                    return
                    
                # Read faculty data with enhanced error handling
                try:
                    df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
                except Exception as e:
                    st.error("Error reading faculty data. Please try again.")
                    return
                    
                df_faculty['Username'] = df_faculty['Username'].astype(str).str.strip()
                df_faculty['Password'] = df_faculty['Password'].astype(str).str.strip()
                
                # Clean input credentials
                username = str(username).strip()
                password = str(password).strip()
                
                # Check if user exists
                user_mask = (df_faculty['Username'] == username) & (df_faculty['Password'] == password)
                
                if not user_mask.any():
                    st.error("Invalid credentials")
                    return
                    
                # Get user info and row index
                user_row = df_faculty[user_mask].iloc[0]
                user_index = df_faculty[user_mask].index[0]
                faculty_name = user_row['Faculty Name']
                
                # For Admin login, check if it's the first row (index 0)
                if login_type == "Admin":
                    if user_index != 0:  # If not first row
                        st.error("Invalid admin credentials")
                        return
                else:  # For Faculty login
                    if user_index == 0:  # If first row
                        st.error("Please use Admin login for admin credentials")
                        return
                
                # Create backup after successful login
                backup_excel_file()
                
                # If we get here, credentials are valid
                st.session_state.logged_in = True
                st.session_state.is_admin = (login_type == "Admin")
                st.session_state.username = username
                st.session_state.faculty_name = faculty_name
                st.rerun()
                    
            except Exception as e:
                st.error(f"Login error: {str(e)}")
    else:
        # Create periodic backup during session
        if 'last_backup_time' not in st.session_state:
            st.session_state.last_backup_time = datetime.now()
        elif (datetime.now() - st.session_state.last_backup_time).seconds > 300:  # 5 minutes
            backup_excel_file()
            st.session_state.last_backup_time = datetime.now()
            
        if st.session_state.is_admin:
            admin_page()
        else:
            faculty_page()


if __name__ == "__main__":
    main()